"""只读:dump 招标采购清单 Excel + 招标 PDF 文字层,供人工/程序比对。

输出 UTF-8 文件,避免 Windows 控制台中文乱码。
从仓库根运行:  python scripts/dump_tender.py
"""
from __future__ import annotations
import os
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TEST = ROOT / "docs" / "test"
OUT = ROOT / "outputs" / "tender_reconcile"
OUT.mkdir(parents=True, exist_ok=True)

XLSX = TEST / "金桥地铁上盖J9A-03地块（浦发上城科创智谷）研发及商业项目（阀门）招标清单.xlsx"
PDF = TEST / "金桥地体上盖招标文件.pdf"

# ---- 1. dump Excel ----
import openpyxl
wb = openpyxl.load_workbook(str(XLSX), data_only=True)
lines = []
for ws in wb.worksheets:
    lines.append(f"=== SHEET: {ws.title}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column} ===")
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = ["" if v is None else str(v) for v in row]
        if any(c.strip() for c in cells):
            lines.append(f"[{r:>3}] " + " | ".join(cells))
(OUT / "excel_dump.txt").write_text("\n".join(lines), encoding="utf-8")
print(f"excel_dump.txt written: {len(lines)} lines")

# ---- 2. dump PDF text per page ----
import pypdfium2 as pdfium
doc = pdfium.PdfDocument(str(PDF))
plines = []
for i in range(len(doc)):
    txt = doc[i].get_textpage().get_text_range()
    plines.append(f"\n===== PAGE {i+1} ({len(txt.strip())} chars) =====")
    plines.append(txt)
doc.close()
(OUT / "pdf_text_dump.txt").write_text("\n".join(plines), encoding="utf-8")
print(f"pdf_text_dump.txt written: {len(doc) if False else ''} pages")
print("done ->", OUT)
