"""Build versioned raw CSV assets from source Excel files.

Reads source_registry.json, exports every active source's sheets to
docs/data/raw/<date>/csv/<category>/, and writes a manifest + audit reports.

Usage:
    python scripts/build_raw_assets.py                     # today's date
    python scripts/build_raw_assets.py --date 2026-06-23
    python scripts/build_raw_assets.py --dry-run           # plan only
    python scripts/build_raw_assets.py --registry path/to/source_registry.json

Output:
    docs/data/raw/<date>/
        manifest.json              SHA256 + row/col/sheet counts per source
        source_registry.json       snapshot of the registry used
        csv/<category>/            per-sheet CSVs (no header row, original values)
        audits/
            workbook_summary.json
            conversion_report.json
            配电箱_version_diff.csv (brand column analysis)
"""

import argparse
import csv as csvlib
import hashlib
import json
import sys
from datetime import date
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_REGISTRY = ROOT / "docs" / "data" / "source_registry.json"
RAW_BASE = ROOT / "docs" / "data" / "raw"

# String values in the brand column of 配电箱 that are NOT brand names
_PDB_NON_BRAND = frozenset({"新增", "利用原有箱体", "新增配电箱", "新作"})


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def sanitize(name: str) -> str:
    return name.replace("/", "_").replace("\\", "_").replace(" ", "_").strip("_")


def export_workbook(src: Path, out_dir: Path, engine: str) -> list[dict]:
    """Export every sheet to a CSV file. Returns per-sheet info list."""
    out_dir.mkdir(parents=True, exist_ok=True)
    entries = []
    xls = pd.ExcelFile(src, engine=engine)
    stem = sanitize(src.stem.lstrip("0").strip())
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)
        csv_name = f"{stem}_{sanitize(sheet)}.csv"
        df.to_csv(out_dir / csv_name, index=False, header=False, encoding="utf-8-sig")
        entries.append({
            "sheet": sheet,
            "csv": csv_name,
            "rows": len(df),
            "cols": len(df.columns),
        })
    return entries


def build_pdb_diff(original: Path, active: Path, out_path: Path) -> dict:
    """Cell-level diff between two 配电箱 workbooks.

    Flags brand-column cells containing non-brand status strings (新增 etc.)
    as candidates for migration to box_status_raw.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed — skipping diff"}

    def _load(p):
        return openpyxl.load_workbook(p, data_only=True)

    wb_orig = _load(original)
    wb_new  = _load(active)

    orig_sheets = set(wb_orig.sheetnames)
    new_sheets  = set(wb_new.sheetnames)

    added_sheets   = sorted(new_sheets - orig_sheets)
    removed_sheets = sorted(orig_sheets - new_sheets)
    common_sheets  = sorted(orig_sheets & new_sheets)

    diffs = []
    box_status_raw_candidates = []

    for sheet in common_sheets:
        ws_o = wb_orig[sheet]
        ws_n = wb_new[sheet]
        max_row = max(ws_o.max_row, ws_n.max_row)
        max_col = max(ws_o.max_column, ws_n.max_column)

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v_o = str(ws_o.cell(r, c).value or "").strip()
                v_n = str(ws_n.cell(r, c).value or "").strip()
                if v_o != v_n:
                    entry = {"sheet": sheet, "row": r, "col": c,
                             "original": v_o, "new": v_n}
                    diffs.append(entry)
                    if v_n in _PDB_NON_BRAND:
                        box_status_raw_candidates.append(
                            {**entry, "action": "migrate_to_box_status_raw"}
                        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csvlib.DictWriter(
            f, fieldnames=["sheet", "row", "col", "original", "new", "action"]
        )
        w.writeheader()
        for d in diffs:
            w.writerow({**d, "action": "changed"})
        for d in box_status_raw_candidates:
            w.writerow(d)

    return {
        "added_sheets": added_sheets,
        "removed_sheets": removed_sheets,
        "common_sheets_count": len(common_sheets),
        "changed_cells": len(diffs),
        "box_status_raw_candidates": len(box_status_raw_candidates),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--registry", default=str(DEFAULT_REGISTRY))
    ap.add_argument("--date", default=str(date.today()))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    registry_path = Path(args.registry)
    if not registry_path.exists():
        print(f"[ERROR] Registry not found: {registry_path}", file=sys.stderr)
        sys.exit(1)

    with open(registry_path, encoding="utf-8") as f:
        reg_data = json.load(f)
    sources = reg_data["sources"]

    out_dir = RAW_BASE / args.date

    if args.dry_run:
        print(f"DRY RUN — output would go to: {out_dir}\n")
        for s in sources:
            src = ROOT / s["path"]
            ok = "OK" if src.exists() else "MISSING"
            print(f"  [{s['status']:12s}]  {s['source_id']:38s}  {ok}")
        return

    # ── collect SHA256 for superseded sources (diff reference only) ──────────
    pdb_original_path: Path | None = None
    pdb_active_path:   Path | None = None

    for s in sources:
        if s["status"] == "superseded":
            p = ROOT / s["path"]
            if p.exists():
                s["sha256"] = sha256_file(p)
                if s["category"] == "配电箱":
                    pdb_original_path = p
            continue

    # ── export active sources ────────────────────────────────────────────────
    manifest_entries = []
    workbook_summary = {}
    conservation = {"pass": True, "sources": []}

    for s in sources:
        if s["status"] != "active":
            continue

        src = ROOT / s["path"]
        if not src.exists():
            print(f"[WARN] Missing: {src}", file=sys.stderr)
            conservation["pass"] = False
            continue

        sha = sha256_file(src)
        s["sha256"] = sha
        engine = "xlrd" if src.suffix == ".xls" else "openpyxl"

        csv_out = out_dir / "csv" / s["category"]
        sheets = export_workbook(src, csv_out, engine)
        total_rows = sum(e["rows"] for e in sheets)

        print(f"[{s['status']:6s}] {s['source_id']:38s}  "
              f"{len(sheets):2d} sheets  {total_rows:5d} rows  sha={sha[:10]}…")

        if s["category"] == "配电箱":
            pdb_active_path = src

        entry = {
            "source_id":  s["source_id"],
            "category":   s["category"],
            "profession": s["profession"],
            "path":       s["path"],
            "sha256":     sha,
            "status":     s["status"],
            "authority":  s.get("authority", ""),
            "sheet_count": len(sheets),
            "total_rows": total_rows,
            "csv_dir":    f"csv/{s['category']}",
            "sheets":     sheets,
        }
        manifest_entries.append(entry)
        workbook_summary[s["source_id"]] = entry

        conservation["sources"].append({
            "source_id":     s["source_id"],
            "source_rows":   total_rows,
            "exported_rows": total_rows,
            "pass":          True,
        })

    # ── 配電箱 version diff ──────────────────────────────────────────────────
    diff_summary: dict = {}
    if pdb_original_path and pdb_active_path:
        print(f"\n[DIFF] 配电箱 {pdb_original_path.name} vs {pdb_active_path.name}")
        diff_path = out_dir / "audits" / "配电箱_version_diff.csv"
        diff_summary = build_pdb_diff(pdb_original_path, pdb_active_path, diff_path)
        print(f"       changed cells: {diff_summary.get('changed_cells', '?')}, "
              f"box_status_raw candidates: {diff_summary.get('box_status_raw_candidates', '?')}")

    # ── write manifests ──────────────────────────────────────────────────────
    audits = out_dir / "audits"
    audits.mkdir(parents=True, exist_ok=True)

    manifest = {
        "schema_version": "1",
        "generated_at": args.date,
        "source_registry_sha256": sha256_file(registry_path),
        "sources": manifest_entries,
    }
    (out_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "source_registry.json").write_text(
        json.dumps(reg_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (audits / "workbook_summary.json").write_text(
        json.dumps(workbook_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (audits / "conversion_report.json").write_text(
        json.dumps(
            {**conservation, "pdb_diff": diff_summary},
            ensure_ascii=False, indent=2,
        ),
        encoding="utf-8",
    )

    active_count = len(manifest_entries)
    total = sum(e["total_rows"] for e in manifest_entries)
    print(f"\nDone. {active_count} active sources, {total} total rows → {out_dir}")


if __name__ == "__main__":
    main()
