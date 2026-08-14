"""rebuild_golden_from_excel.py — 从最新人工核对 Excel 重建三份 golden fixture JSON。

非生产能力。Golden 重建后必须经人工审阅并通过 E2E fresh 验收才能作为基准。

用法：
    python scripts/rebuild_golden_from_excel.py
"""
from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

DOCS = REPO / "tests" / "fixtures" / "documents" / "bid_list"
GOLDEN_DIR = REPO / "data" / "golden"

_ABS_TOL = 0.05
_REL_TOL = 0.005


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _coerce_num(v):
    if v is None or str(v).strip() in ("", "-"):
        return None
    s = str(v).strip().replace(",", "").replace("，", "")
    pct = s.endswith("%")
    if pct:
        s = s[:-1].strip()
    try:
        f = float(s)
        return f / 100.0 if pct else f
    except (ValueError, TypeError):
        return None


# ── 凯硕新正 ───────────────────────────────────────────────────────────────────

def _build_kaishuo() -> dict:
    """凯硕新正投标清单.xlsx → quote_kaishuo golden.

    列：序号/专业/项目名称/规格/型号/工作压力/阀体/阀芯/阀板/阀杆/密封圈/
        单位/数量/单价(不含税)/合计(不含税)/税率/税额/单价(含税)/合价(含税)/品牌/备注
    所有字段均为 raw；不含税合计/税额标 ambiguous（分角级舍入差异）。
    """
    import openpyxl
    xlsx = DOCS / "凯硕新正投标清单.xlsx"
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active

    raw_headers = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    col = {h: i for i, h in enumerate(raw_headers)}

    def _get(row_vals, key):
        idx = col.get(key)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    rows = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        seq = str(_get(row_vals, "序号") or "").strip()
        if not seq:
            continue
        if not seq.isdigit():
            continue  # skip subtotal/header rows

        materials = {}
        for mat_col in ("阀体", "阀芯", "阀板", "阀杆", "密封圈"):
            v = str(_get(row_vals, mat_col) or "").strip()
            if v and v != "-":
                materials[mat_col] = v

        rows.append({
            "seq": seq,
            "profession": str(_get(row_vals, "专业") or "").strip(),
            "name": str(_get(row_vals, "项目名称") or "").strip(),
            "spec": str(_get(row_vals, "规格") or "").strip(),
            "model": str(_get(row_vals, "型号") or "").strip(),
            "pressure": str(_get(row_vals, "工作压力") or "").strip(),
            "materials": materials,
            "unit": str(_get(row_vals, "单位") or "").strip(),
            "qty": _coerce_num(_get(row_vals, "数量")),
            "unit_price_excl_tax": _coerce_num(_get(row_vals, "单价(不含税)")),
            "total_price_excl_tax": _coerce_num(_get(row_vals, "合计(不含税)")),
            "tax_rate": _coerce_num(_get(row_vals, "税率")),
            "tax_amount": _coerce_num(_get(row_vals, "税额")),
            "unit_price_incl_tax": _coerce_num(_get(row_vals, "单价(含税)")),
            "total_price_incl_tax": _coerce_num(_get(row_vals, "合价(含税)")),
            "brand": str(_get(row_vals, "品牌") or "").strip(),
            "remark": str(_get(row_vals, "备注") or "").strip(),
        })

    incl_sum = round(sum((r["total_price_incl_tax"] or 0) for r in rows), 2)
    declared = 932_154.0

    return {
        "doc_type": "quote",
        "source_file": "凯硕新正投标清单.xlsx",
        "source_sha256": _sha256(xlsx),
        "declared_total": declared,
        "row_count": len(rows),
        "field_sources": {
            "unit_price_incl_tax": "raw",
            "total_price_incl_tax": "raw",
            "unit_price_excl_tax": "raw",
            "total_price_excl_tax": "ambiguous",  # PDF 分角级舍入差异
            "tax_amount": "ambiguous",
            "tax_rate": "raw",
            "qty": "raw",
            "name": "raw",
            "spec": "raw",
        },
        "audit_status": "excel_verified",
        "audit_notes": "人工核对89行；单价含税/不含税均来自Excel原始列；seq2/5/73存在分角级舍入差异(A类)",
        "rows": rows,
        "validation": {
            "row_count": len(rows),
            "seq_range": [1, max(int(r["seq"]) for r in rows)],
            "incl_sum": incl_sum,
            "declared_total_diff": round(abs(incl_sum - declared), 2),
        },
    }


# ── 泰科龙 ─────────────────────────────────────────────────────────────────────

def _build_taikelong() -> dict:
    """泰科龙投标清单.xlsx → quote_taikelong golden.

    列：序号/专业/项目名称/规格/型号/工作压力/阀体/阀芯/阀板/阀杆/密封圈/
        单位/数量/单价(不含税)/合计(不含税)/税率/税额/价税合计/品牌/备注/系统
    注：Excel 无含税单价列，unit_price_incl_tax=null。
    """
    import openpyxl
    xlsx = DOCS / "泰科龙投标清单.xlsx"
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active

    raw_headers = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(raw_headers)}

    def _get(row_vals, key):
        idx = col.get(key)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    rows = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        seq = str(_get(row_vals, "序号") or "").strip()
        if not seq or not seq.isdigit():
            continue

        materials = {}
        for mat_col in ("阀体", "阀芯", "阀板", "阀杆", "密封圈"):
            v = str(_get(row_vals, mat_col) or "").strip()
            if v and v != "-":
                materials[mat_col] = v

        rows.append({
            "seq": seq,
            "profession": str(_get(row_vals, "专业") or "").strip(),
            "name": str(_get(row_vals, "项目名称") or "").strip(),
            "spec": str(_get(row_vals, "规格") or "").strip(),
            "model": str(_get(row_vals, "型号") or "").strip(),
            "pressure": str(_get(row_vals, "工作压力") or "").strip(),
            "materials": materials,
            "unit": str(_get(row_vals, "单位") or "").strip(),
            "qty": _coerce_num(_get(row_vals, "数量")),
            "unit_price_excl_tax": _coerce_num(_get(row_vals, "单价(不含税)")),
            "total_price_excl_tax": _coerce_num(_get(row_vals, "合计(不含税)")),
            "tax_rate": _coerce_num(_get(row_vals, "税率")),
            "tax_amount": _coerce_num(_get(row_vals, "税额")),
            "unit_price_incl_tax": None,  # Excel 无此列；PDF 转置表不直接提供含税单价
            "total_price_incl_tax": _coerce_num(_get(row_vals, "价税合计")),
            "brand": str(_get(row_vals, "品牌") or "").strip(),
            "remark": str(
                _get(row_vals, "备注") or _get(row_vals, "系统") or
                _get(row_vals, "备注/系统") or ""
            ).strip(),
        })

    incl_sum = round(sum((r["total_price_incl_tax"] or 0) for r in rows), 2)
    declared = 1_067_616.41

    return {
        "doc_type": "quote",
        "source_file": "泰科龙投标文件.pdf",
        "source_sha256": _sha256(xlsx),
        "declared_total": declared,
        "row_count": len(rows),
        "field_sources": {
            "unit_price_incl_tax": "derived",       # Excel 无此列
            "total_price_incl_tax": "raw",           # "价税合计" 列
            "unit_price_excl_tax": "raw",
            "total_price_excl_tax": "raw",
            "tax_amount": "raw",
            "tax_rate": "raw",
            "qty": "raw",
            "name": "raw",
            "spec": "raw",
        },
        "audit_status": "excel_verified",
        "audit_notes": "人工核对89行；PDF为转置表；unit_price_incl_tax需从PDF含税合价÷数量推算，Excel未存",
        "rows": rows,
        "validation": {
            "row_count": len(rows),
            "seq_range": [1, max(int(r["seq"]) for r in rows)],
            "incl_sum": incl_sum,
            "declared_total_diff": round(abs(incl_sum - declared), 2),
        },
    }


# ── 上海绵存 ───────────────────────────────────────────────────────────────────

def _build_miancun() -> dict:
    """上海绵存投标清单.xlsx → quote_miancun golden.

    列：序号/品名/规格/型号/单位/数量/单价/合价
    PDF 审计结论：PDF "单价"/"合价" 实为含税价，无不含税列。
    故：
      unit_price_incl_tax = Excel "单价" 列（含税单价原文）
      total_price_incl_tax = Excel "合价" 列（含税合价原文）
      所有不含税/税额字段 = null（不推导）
    """
    import openpyxl
    xlsx = DOCS / "上海绵存投标清单.xlsx"
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active

    raw_headers = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(raw_headers)}

    def _get(row_vals, key):
        idx = col.get(key)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    rows = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        seq = str(_get(row_vals, "序号") or "").strip()
        if not seq or not seq.isdigit():
            continue

        rows.append({
            "seq": seq,
            "name": str(_get(row_vals, "品名") or "").strip(),
            "spec": str(_get(row_vals, "规格") or "").strip(),
            "model": str(_get(row_vals, "型号") or "").strip(),
            "unit": str(_get(row_vals, "单位") or "").strip(),
            "qty": _coerce_num(_get(row_vals, "数量")),
            "unit_price_excl_tax": None,         # Excel 无不含税列，不推导
            "total_price_excl_tax": None,
            "tax_rate": None,
            "tax_amount": None,
            "unit_price_incl_tax": _coerce_num(_get(row_vals, "单价")),   # PDF 含税单价
            "total_price_incl_tax": _coerce_num(_get(row_vals, "合价")),  # PDF 含税合价
            "brand": "",
            "remark": "",
        })

    incl_sum = round(sum((r["total_price_incl_tax"] or 0) for r in rows), 2)
    declared = 1_667_051.0

    return {
        "doc_type": "quote",
        "source_file": "上海绵存投标文件.pdf",
        "source_sha256": _sha256(xlsx),
        "declared_total": declared,
        "row_count": len(rows),
        "field_sources": {
            "unit_price_incl_tax": "raw",        # Excel "单价" = PDF 含税单价
            "total_price_incl_tax": "raw",        # Excel "合价" = PDF 含税合价
            "unit_price_excl_tax": "derived",     # 不在 Excel 中，评价时降级
            "total_price_excl_tax": "derived",
            "tax_amount": "derived",
            "tax_rate": "ambiguous",
            "qty": "raw",
            "name": "raw",
            "spec": "raw",
        },
        "audit_status": "excel_verified",
        "audit_notes": (
            "人工核对89行（原86行+3行补录）；PDF仅有单价/合价列实为含税价；"
            "不含税字段均为derived，不参与E2E字段准确率评估。"
            "raw_name(ocr_candidate)未录入本版golden，评价name字段需对照PDF原文。"
        ),
        "rows": rows,
        "validation": {
            "row_count": len(rows),
            "seq_range": [1, max(int(r["seq"]) for r in rows)],
            "incl_sum": incl_sum,
            "declared_total_diff": round(abs(incl_sum - declared), 2),
        },
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    GOLDEN_DIR.mkdir(parents=True, exist_ok=True)
    configs = [
        ("quote_kaishuo", _build_kaishuo),
        ("quote_taikelong", _build_taikelong),
        ("quote_miancun", _build_miancun),
    ]
    for name, fn in configs:
        print(f"\n[{name}]")
        try:
            data = fn()
        except Exception as e:
            print(f"  ERROR: {e}")
            raise
        out = GOLDEN_DIR / f"{name}.json"
        out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        v = data["validation"]
        print(f"  rows={data['row_count']} seq={v['seq_range']}"
              f" incl_sum={v['incl_sum']:,} declared={data['declared_total']:,}"
              f" diff={v['declared_total_diff']}")
        print(f"  sha256={data['source_sha256'][:16]}...")
        print(f"  → {out.relative_to(REPO)}")


if __name__ == "__main__":
    main()
