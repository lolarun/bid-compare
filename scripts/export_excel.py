"""Export data analysis results to a formatted Excel workbook."""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from analyze_data import (
    ALGO_ASSESSMENT,
    LAYER3_TECH_PARAMS,
    PROFESSION_MAP,
    run_analysis,
)

BASE = Path(__file__).resolve().parent.parent
OUT_PATH = BASE / "docs" / "data" / "analysis" / "数据分析报告.xlsx"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="Arial", bold=True, size=16)
SECTION_FONT = Font(name="Arial", bold=True, size=12, color="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
BODY_BOLD = Font(name="Arial", size=10, bold=True)
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="666666")

BLUE_FILL = PatternFill("solid", fgColor="4472C4")
MED_BLUE_FILL = PatternFill("solid", fgColor="8DB4E2")
LIGHT_GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
ORANGE_FILL = PatternFill("solid", fgColor="ED7D31")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4EC")

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

ALGO_NAMES = {
    "(2)+(1)": "规格回归+历史比对",
    "(1)+(2)": "历史比对+规格回归",
    "(1)仅限": "历史比对(仅限)",
    "暂不可": "暂不可",
}

ALGO_DESC = {
    "历史比对": "基于IQR统计建立历史价格基线，超出[Q1-1.5IQR, Q3+1.5IQR]区间报警",
    "规格回归": "用规格参数拟合价格模型，修正Z分数>3.0判定异常",
}

ALGO_EXAMPLE = {
    "历史比对": "桥架200×100热镀锌，历史价45-55元/m，新报38元→低于Q1-1.5IQR→异常偏低",
    "规格回归": "防火阀800×400，模型预测850元，实际报1200元→Z=3.8>3.0→异常偏高",
}


def _write_header(ws, row, headers, start_col=1, fill=BLUE_FILL):
    for ci, h in enumerate(headers, start_col):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = fill
        c.border = BORDER
        c.alignment = CENTER_WRAP


def _write_row(ws, row, values, start_col=1, striped=False, bold_cols=None):
    fill = LIGHT_GRAY_FILL if striped else None
    for ci, v in enumerate(values, start_col):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = BODY_BOLD if (bold_cols and ci in bold_cols) else BODY_FONT
        if fill:
            c.fill = fill
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)


def _section_header(ws, row, text, ncols, start_col=1, fill=BLUE_FILL):
    end_col = start_col + ncols - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = SECTION_FONT
    c.fill = fill
    c.alignment = Alignment(vertical="center")
    for ci in range(start_col, end_col + 1):
        ws.cell(row=row, column=ci).border = BORDER
        ws.cell(row=row, column=ci).fill = fill


def _brand_text(r):
    if r["n_brands"] > 0 and r["brands"]:
        top3 = r["brands"][:3]
        if r["n_brands"] > 3:
            return ", ".join(top3) + f"等{r['n_brands']}家"
        return ", ".join(top3)
    return "无，需补充"


def _data_col(r):
    s = f"{r['n_valid_price']}/{r['n_projects']}/{r['n_brands']}"
    pct = float(r["price_missing_pct"])
    if pct > 1:
        s += f" (缺失{r['price_missing_pct']}%)"
    return s


def _set_col_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths, start_col):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 1: 数据概览 (data source + price distribution combined)
# ---------------------------------------------------------------------------
def create_overview(wb, results):
    ws = wb.create_sheet("数据概览")

    ws.merge_cells("A1:L1")
    ws.cell(row=1, column=1, value="数据分析报告").font = TITLE_FONT

    # --- Data source table (reordered: 数据来源 first, 表(Sheet) second) ---
    row = 3
    _section_header(ws, row, "数据来源汇总", 6)
    row += 1
    _write_header(ws, row, ["数据来源", "表（Sheet）", "专业", "品类", "汇总行数", "项目数"])
    row += 1

    src_data = [
        ("0桥架报价单格式模板.xls", 14, "电气", "桥架", "2,047", 13),
        ("0母线报价单格式模板.xls", 8, "电气", "母线槽", 140, 7),
        ("0配电箱.xlsx", 9, "电气", "配电箱", "—*", 9),
        ("0阀门询价格式.xls", 16, "给排水", "阀门", 917, 15),
        ("0不锈钢管清单.xlsx", 7, "给排水", "不锈钢管", 581, 5),
        ("0水箱报价清单.xlsx", 7, "给排水", "水箱", 15, 4),
        ("0潜水泵询价格式.xlsx", 8, "给排水", "潜水泵", 618, 7),
        ("0风口风阀报价单格式.xls", 13, "暖通", "风口风阀", "3,101", 11),
        ("0风盘报价单格式.xls", 10, "暖通", "风机盘管", 93, 8),
        ("0空调泵询价格式 .xlsx", 6, "暖通", "空调泵", 48, 4),
    ]
    for i, vals in enumerate(src_data):
        _write_row(ws, row, vals, striped=(i % 2 == 1))
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1,
            value="合计：3个专业、10个品类、10个Excel文件、98个Sheet、约7,560条汇总数据行").font = BODY_BOLD
    row += 2

    # --- Price distribution (all categories in one table) ---
    _section_header(ws, row, "价格分布分析", 12)
    row += 1

    note = ws.cell(row=row, column=1,
                   value="CV（变异系数）= 标准差÷均价，反映离散程度。以下按子类汇总，未控制规格差异，实际比价时按[匹配]条件分组后重新计算。")
    note.font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    row += 1

    price_headers = ["品类", "子类", "数据量", "均价", "中位价", "标准差",
                     "变异系数", "最低", "最高", "P10", "P90", "建议容差"]
    _write_header(ws, row, price_headers)
    row += 1

    stripe_idx = 0
    for r in results:
        cat = r["category"]
        if not r["price_stats"]:
            continue
        for s in r["price_stats"]:
            threshold_str = f"±{s['suggested_threshold']:.0%}"
            vals = [
                cat, s["子类"], s["count"],
                round(s["mean"]), round(s["median"]),
                round(s["std"]), round(s["cv"], 2),
                round(s["min"]), round(s["max"]),
                round(s["p10"]), round(s["p90"]),
                threshold_str,
            ]
            _write_row(ws, row, vals, striped=(stripe_idx % 2 == 1))

            ws.cell(row=row, column=7).number_format = "0.00"

            cv_val = s["cv"]
            cv_cell = ws.cell(row=row, column=7)
            if cv_val < 0.5:
                cv_cell.fill = GREEN_FILL
            elif cv_val <= 1.0:
                cv_cell.fill = YELLOW_FILL
            else:
                cv_cell.fill = RED_FILL

            stripe_idx += 1
            row += 1

    note = ws.cell(row=row, column=1,
                   value="注：配电箱为BOM组合定价，不同配置箱子价格可相差数百倍，CV极高属结构性原因。")
    note.font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)

    _set_col_widths(ws, [10, 18, 8, 8, 8, 8, 7, 8, 8, 8, 8, 10])
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Sheet 2: 物料分类 (3 profession tables side by side)
# ---------------------------------------------------------------------------
def create_classification(wb, results):
    ws = wb.create_sheet("物料分类")

    professions = ["电气", "给排水", "暖通"]
    prof_cats = {}
    for r in results:
        prof = PROFESSION_MAP.get(r["category"], "未分类")
        prof_cats.setdefault(prof, []).append(r)

    col_offset = 0
    for prof in professions:
        cats = prof_cats.get(prof, [])
        if not cats:
            continue

        start_col = col_offset + 1
        _section_header(ws, 1, prof, 3, start_col=start_col, fill=BLUE_FILL)
        _write_header(ws, 2, ["品类", "子类", "数据条数"], start_col=start_col)

        row = 3
        stripe_idx = 0
        for r in cats:
            cat = r["category"]
            sorted_subs = sorted(r["subcat_counts"].items(), key=lambda x: -x[1])
            cat_start = row
            for subcat, count in sorted_subs:
                _write_row(ws, row, [cat, subcat, count],
                           start_col=start_col, striped=(stripe_idx % 2 == 1))
                stripe_idx += 1
                row += 1
            if row - cat_start > 1:
                ws.merge_cells(start_row=cat_start, start_column=start_col,
                               end_row=row - 1, end_column=start_col)

        for i, w in enumerate([10, 18, 10], start_col):
            ws.column_dimensions[get_column_letter(i)].width = w

        # gap column
        gap_col = start_col + 3
        ws.column_dimensions[get_column_letter(gap_col)].width = 3
        col_offset += 4

    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Sheet 3: 物料属性 (Layer 1 + unified Layer 2 + Layer 3)
# ---------------------------------------------------------------------------
def create_attributes(wb):
    ws = wb.create_sheet("物料属性")
    row = 1

    # --- Layer 1 ---
    _section_header(ws, row, "Layer 1 — 基础属性（所有品类共享）", 4)
    row += 1
    _write_header(ws, row, ["属性", "说明", "", ""])
    row += 1
    layer1 = [
        ("物料编码", "系统内唯一标识"),
        ("物料名称", "标准化品名"),
        ("规格型号", "关键区分描述"),
        ("专业(大类)", "电气/给排水/暖通"),
        ("品类(小类)", "桥架/阀门/风口风阀等"),
        ("子类", "闸阀/蝶阀/止回阀等"),
        ("基本计量单位", "m/台/套/座等"),
        ("品牌/制造商", "供应商品牌"),
        ("执行标准", "国标/行标/企标编号"),
    ]
    for i, (attr, desc) in enumerate(layer1):
        _write_row(ws, row, [attr, desc, "", ""], striped=(i % 2 == 1))
        row += 1
    row += 1

    # --- Layer 2 (unified single table) ---
    _section_header(ws, row, "Layer 2 — 扩展属性（按品类动态）", 4)
    row += 1
    note = ws.cell(row=row, column=1,
                   value='[匹配] = 必须相同才可比，[差异] = 允许不同但需说明')
    note.font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    _write_header(ws, row, ["品类", "扩展属性", "数据来源列", "比价规则"])
    row += 1
    stripe_idx = 0
    for cat, params in LAYER3_TECH_PARAMS.items():
        for name, source, rule in params:
            tag = "[匹配]" if rule == "匹配" else "[差异]"
            _write_row(ws, row, [cat, name, source, tag], striped=(stripe_idx % 2 == 1))
            tag_cell = ws.cell(row=row, column=4)
            if rule == "匹配":
                tag_cell.font = Font(name="Arial", size=10, bold=True, color="CC0000")
            else:
                tag_cell.font = Font(name="Arial", size=10, color="666666")
            stripe_idx += 1
            row += 1
    row += 1

    # --- Layer 3 ---
    _section_header(ws, row, "Layer 3 — 商务属性（所有品类共享）", 4)
    row += 1
    _write_header(ws, row, ["属性", "说明", "比价作用", "数据来源"])
    row += 1
    layer3 = [
        ("不含税单价", "比价核心指标", "价格对比基准", "各品类独立列"),
        ("含税单价", "含增值税价格", "合同总价计算", "各品类独立列"),
        ("税率", "通常13%", "含税/不含税换算", "各品类独立列"),
        ("数量", "工程量", "总价计算", "各品类独立列"),
        ("备注", "原文保留，不拆分", "LLM比价分析时读取", "各品类独立列"),
    ]
    for i, vals in enumerate(layer3):
        _write_row(ws, row, vals, striped=(i % 2 == 1))
        row += 1
    row += 1

    note = ws.cell(row=row, column=1,
                   value="备注列原文完整保留，比价时由LLM提取付款条件、质保等商务信息进行横向对比。")
    note.font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    _set_col_widths(ws, [18, 22, 30, 15])
    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Sheet 4: 比价建议 (with full algorithm names + description + examples)
# ---------------------------------------------------------------------------
def create_recommendations(wb, results):
    ws = wb.create_sheet("比价建议")

    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value="各品类比价建议").font = Font(name="Arial", bold=True, size=14)

    # --- Algorithm description section ---
    row = 3
    _section_header(ws, row, "比价算法说明", 3)
    row += 1
    _write_header(ws, row, ["算法名称", "方法描述", "适用范例"])
    row += 1
    for algo_name in ["历史比对", "规格回归"]:
        _write_row(ws, row, [algo_name, ALGO_DESC[algo_name], ALGO_EXAMPLE[algo_name]])
        ws.cell(row=row, column=1).font = BODY_BOLD
        row += 1

    row += 1

    # --- Recommendation table ---
    _section_header(ws, row, "各品类比价建议", 10)
    row += 1
    rec_headers = ["品类", "总评", "数据(条/项目/供应商)", "建议算法",
                   "回归模型", "[匹配]条件", "[差异]解释", "主要品牌",
                   "算法描述", "适用范例"]
    _write_header(ws, row, rec_headers)
    row += 1

    for i, r in enumerate(results):
        cat = r["category"]
        a = ALGO_ASSESSMENT.get(cat, {})
        primary_code = a.get("primary", "—")
        algo_display = ALGO_NAMES.get(primary_code, primary_code)

        # Build description and example from component algorithms
        algo_parts = []
        example_parts = []
        if "规格回归" in algo_display:
            algo_parts.append(ALGO_DESC["规格回归"])
            example_parts.append(ALGO_EXAMPLE["规格回归"])
        if "历史比对" in algo_display:
            algo_parts.append(ALGO_DESC["历史比对"])
            example_parts.append(ALGO_EXAMPLE["历史比对"])

        vals = [
            cat,
            a.get("summary", "—"),
            _data_col(r),
            algo_display,
            a.get("regression", "—"),
            a.get("match_cond", "—"),
            a.get("diff_explain", "—"),
            _brand_text(r),
            "；".join(algo_parts) if algo_parts else "—",
            "；".join(example_parts) if example_parts else "—",
        ]
        _write_row(ws, row, vals, striped=(i % 2 == 1))

        verdict = a.get("summary", "")
        vc = ws.cell(row=row, column=2)
        if "可支撑" in verdict:
            vc.fill = GREEN_FILL
        elif "⚠" in verdict:
            vc.fill = YELLOW_FILL
        row += 1

    _set_col_widths(ws, [10, 12, 22, 18, 22, 22, 18, 35, 40, 45])
    ws.sheet_properties.tabColor = "4472C4"


# ---------------------------------------------------------------------------
# Sheet 5: 待确认事项 (reordered by category)
# ---------------------------------------------------------------------------
def create_pending(wb):
    ws = wb.create_sheet("待确认事项")

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="待确认事项").font = Font(name="Arial", bold=True, size=14)

    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1,
            value="以下事项需与客户确认后方可进入系统设计阶段，每项均给出建议方案。").font = NOTE_FONT

    row = 4

    groups = [
        ("物料主数据", [
            ("物料分类体系",
             "当前大类/品类/子类沿用数据文件结构，不一定与一建内部编码一致",
             "以一建内部材料编码为准，国标GB/T50531作为兜底参考"),
            ("Layer 2 扩展属性确认",
             '各品类的[匹配]/[差异]划分直接决定"哪些报价可以放在一起比"',
             "请各专业负责人审核各品类的扩展属性表，确认匹配/差异标注是否合理"),
            ("品牌作为比价维度的权重",
             "客户希望品牌参与权重比价，但结构化数据中仅品牌可用作权重维度",
             "方案A：品牌仅作为筛选条件；方案B：品牌参与加权评分（需定义品牌档次映射表）"),
            ("备注信息提取范围",
             "备注中含付款条件、质保、技术补充等，LLM可提取但需明确优先级",
             "首轮仅提取付款条件和质保信息用于比价报告展示，技术补充归类到Layer 2辅助判断"),
        ]),
        ("比价算法", [
            ("比价算法选择确认",
             "各品类建议算法已给出，需确认是否采纳",
             "按建议执行：有回归条件的品类用规格回归+历史比对，BOM定价类用历史比对"),
        ]),
        ("数据补充", [
            ("配电箱比价粒度",
             "配电箱为BOM组合定价，当前无多供应商数据",
             '按"相同回路配置+相同元器件规格"逐箱比价；需补充≥3家供应商报价'),
            ("配电箱供应商数据补充",
             "9个项目均无品牌/供应商信息，无法横向比价",
             "优先补充配电箱的多供应商报价数据"),
            ("水箱数据补充",
             "仅14条有效价格，统计意义不足",
             "补充更多项目水箱报价（目标≥50条），否则暂不启用统计报警"),
            ("风机盘管供应商补充",
             "仅3家（新晃、特灵、格力），回归数据偏少",
             "补充1-2家供应商；按管制+风量分组后单组数据不足以建立高置信度回归"),
            ("空调泵历史数据补充",
             "46条有效价格、4个项目，规格分散后每组数据有限",
             "补充更多项目数据以稳定历史比对基线"),
        ]),
    ]

    num = 1
    for group_name, items in groups:
        _section_header(ws, row, group_name, 4, fill=MED_BLUE_FILL)
        ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=10, color="000000")
        row += 1
        _write_header(ws, row, ["#", "确认事项", "背景", "建议方案"], fill=ORANGE_FILL)
        row += 1
        for i, (title, bg, plan) in enumerate(items):
            _write_row(ws, row, [num, title, bg, plan], striped=(i % 2 == 1), bold_cols={2})
            num += 1
            row += 1
        row += 1

    _set_col_widths(ws, [5, 22, 45, 45])
    ws.sheet_properties.tabColor = "ED7D31"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def export_excel(results, output_path=None):
    if output_path is None:
        output_path = OUT_PATH
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    create_overview(wb, results)
    create_classification(wb, results)
    create_attributes(wb)
    create_recommendations(wb, results)
    create_pending(wb)

    wb.save(output_path)
    print(f"Excel报告已写入: {output_path}")


if __name__ == "__main__":
    print("正在运行数据分析...")
    results = run_analysis()
    export_excel(results)
