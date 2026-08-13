"""create_golden_fixtures.py — 把 Excel 标准答案转为 golden fixture JSON。

仅生产数据，非生产能力。

用法：
    python scripts/create_golden_fixtures.py
"""
import json
import sys
from pathlib import Path

# ── 路径 ────────────────────────────────────────────────────────────────────
REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

DOCS = REPO / "docs" / "test"
GOLDEN_DIR = REPO / "data" / "golden"

FIXTURES = [
    {
        "name": "quote_taikelong",
        "xlsx": DOCS / "泰科龙投标清单.xlsx",
        "doc_type": "quote",
        "declared_total": 1_067_616.41,
    },
    # 金桥招标: xlsx 存为 data/golden，用 parse_tender_xlsx 解析
    # 凯硕新正、绵存: xlsx 制作完成后补充
]


def _coerce_num(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("，", ""))
    except (ValueError, TypeError):
        return None


def create_taikelong_fixture():
    """泰科龙投标清单.xlsx → data/golden/quote_taikelong.json"""
    import openpyxl

    src_path = DOCS / "泰科龙投标清单.xlsx"
    if not src_path.exists():
        print(f"[skip] {src_path} 不存在")
        return

    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb.active

    # 读表头（第1行）
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"  headers: {headers}")

    col = {h: i for i, h in enumerate(headers)}

    def _get(row, key):
        idx = col.get(key)
        return row[idx].value if idx is not None else None

    rows = []
    for r in ws.iter_rows(min_row=2):
        seq = _get(r, "序号")
        name = _get(r, "项目名称")
        if seq is None and name is None:
            continue
        if str(seq or "").strip() == "" and str(name or "").strip() == "":
            continue

        rows.append({
            "seq": str(seq or "").strip(),
            "profession": str(_get(r, "专业") or "").strip(),
            "name": str(name or "").strip(),
            "spec": str(_get(r, "规格") or _get(r, "型号") or "").strip(),
            "materials": {
                "阀体": str(_get(r, "阀体") or "").strip(),
                "阀芯": str(_get(r, "阀芯") or "").strip(),
                "阀板": str(_get(r, "阀板") or "").strip(),
                "阀杆": str(_get(r, "阀杆") or "").strip(),
                "密封圈": str(_get(r, "密封圈") or "").strip(),
            },
            "unit": str(_get(r, "单位") or "").strip(),
            "qty": _coerce_num(_get(r, "数量")),
            "unit_price_excl_tax": _coerce_num(_get(r, "单价(不含税)")),
            "total_price_excl_tax": _coerce_num(_get(r, "合计(不含税)")),
            "tax_rate": _coerce_num(_get(r, "税率")),
            "tax_amount": _coerce_num(_get(r, "税额")),
            "total_price_incl_tax": _coerce_num(_get(r, "价税合计")),
            "brand": str(_get(r, "品牌") or "").strip(),
            "remark": str(_get(r, "备注") or _get(r, "系统") or "").strip(),
        })

    quote_lines = [r for r in rows if r["seq"].isdigit()]
    total_sum = sum((r["total_price_incl_tax"] or 0) for r in quote_lines)

    fixture = {
        "doc_type": "quote",
        "source_file": "泰科龙投标文件.pdf",
        "declared_total": 1_067_616.41,
        "row_count": len(quote_lines),
        "rows": quote_lines,
        "validation": {
            "row_count": len(quote_lines),
            "seq_range": [min(int(r["seq"]) for r in quote_lines),
                          max(int(r["seq"]) for r in quote_lines)],
            "total_price_incl_tax_sum": round(total_sum, 2),
            "declared_total_diff": round(abs(total_sum - 1_067_616.41), 2),
            "tax_rate_unique": list({r["tax_rate"] for r in quote_lines}),
            "brand_unique": list({r["brand"] for r in quote_lines}),
        },
    }

    GOLDEN_DIR.mkdir(parents=True, exist_ok=True)
    out = GOLDEN_DIR / "quote_taikelong.json"
    out.write_text(json.dumps(fixture, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[ok] {out} — {len(quote_lines)} rows, sum={total_sum:.2f}")
    return fixture


def create_tender_jingqiao_fixture():
    """金桥招标清单 xlsx → data/golden/tender_jingqiao.json"""
    try:
        from apps.api.services.tender.tender_list import parse_tender_xlsx
    except ImportError:
        print("[skip] parse_tender_xlsx import failed — add repo to sys.path")
        return

    src_path = DOCS / "金桥地体上盖招标文件.xlsx"
    if not src_path.exists():
        print(f"[skip] {src_path} 不存在")
        return

    anchors = parse_tender_xlsx(str(src_path))
    rows = [
        {
            "seq": str(a.seq),
            "name": a.name,
            "spec": a.spec,
            "unit": a.unit,
            "qty": a.qty,
            "brand": a.brand,
            "profession": a.profession,
            "materials": a.materials,
        }
        for a in anchors
    ]

    fixture = {
        "doc_type": "tender",
        "source_file": "金桥招标PDF.pdf",
        "row_count": len(rows),
        "rows": rows,
        "validation": {
            "row_count": len(rows),
            "seq_range": [
                min(int(r["seq"]) for r in rows if str(r["seq"]).isdigit()),
                max(int(r["seq"]) for r in rows if str(r["seq"]).isdigit()),
            ],
        },
    }

    GOLDEN_DIR.mkdir(parents=True, exist_ok=True)
    out = GOLDEN_DIR / "tender_jingqiao.json"
    out.write_text(json.dumps(fixture, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[ok] {out} — {len(rows)} rows")
    return fixture


if __name__ == "__main__":
    print("=== create_golden_fixtures ===")
    create_taikelong_fixture()
    create_tender_jingqiao_fixture()
    print("done.")
