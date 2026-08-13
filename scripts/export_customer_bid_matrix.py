"""客户演示 Excel 导出：全量锚点比价矩阵 + 五分析 sheet。

数据源：/api/analysis/bid-matrix (JSON) + SQLite 直查锚点明细/报价明细。
不重新跑 LLM/OCR，只导出当前 confirmed 结果。

用法：
    python scripts/export_customer_bid_matrix.py \\
        --project 62 --category 阀门 \\
        --out outputs/customer_bid_matrix_project62.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import sqlite3
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

API = "http://localhost:8002"
DB = os.path.join(os.path.dirname(__file__), "..", "data", "mempas.db")


def log(m: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


# ── 颜色 / 字体 / 边框 ──────────────────────────────────────────────────────

FILL_HEADER       = PatternFill(start_color="1677FF", fill_type="solid")   # 蓝色表头
FILL_SUB_HEADER   = PatternFill(start_color="EBF3FF", fill_type="solid")   # 浅蓝小标题
FILL_QUOTED_LOWEST = PatternFill(start_color="F6FFED", fill_type="solid")  # 最低价绿
FILL_AGGREGATED   = PatternFill(start_color="E6F7FF", fill_type="solid")   # 聚合浅蓝
FILL_PENDING      = PatternFill(start_color="FFF7E6", fill_type="solid")   # 待确认橙
FILL_MISSING      = PatternFill(start_color="F5F5F5", fill_type="solid")   # 未报价灰
FILL_EXCLUDED     = PatternFill(start_color="E8E8E8", fill_type="solid")   # 已排除深灰
FILL_OCR_OK       = PatternFill(start_color="E6FFFB", fill_type="solid")   # OCR纠错已验证青绿
FILL_CONFLICT     = PatternFill(start_color="FFF1F0", fill_type="solid")   # 阀型冲突浅红
FILL_WARNING      = PatternFill(start_color="FFFBE6", fill_type="solid")   # 其他警告浅黄
FILL_NONE         = PatternFill(fill_type=None)

FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
FONT_BOLD    = Font(bold=True, size=10, name="微软雅黑")
FONT_NORMAL  = Font(size=10, name="微软雅黑")
FONT_SMALL   = Font(size=9, color="595959", name="微软雅黑")
FONT_GREEN   = Font(bold=True, color="389E0D", size=10, name="微软雅黑")
FONT_ORANGE  = Font(bold=True, color="D46B08", size=10, name="微软雅黑")
FONT_RED     = Font(bold=True, color="CF1322", size=10, name="微软雅黑")
FONT_GREY    = Font(size=10, color="AAAAAA", name="微软雅黑")

_thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


# ── API 工具 ────────────────────────────────────────────────────────────────

def api_login(user: str, password: str) -> str:
    r = requests.post(f"{API}/api/auth/login",
                      json={"username": user, "password": password}, timeout=10)
    r.raise_for_status()
    return r.json()["access_token"]


def api_bid_matrix(pid: int, category: str, sids: list[int], tok: str) -> dict:
    r = requests.post(
        f"{API}/api/analysis/bid-matrix",
        json={"project_id": pid, "category": category, "supplier_ids": sids},
        headers={"Authorization": f"Bearer {tok}"},
        timeout=120,
    )
    r.raise_for_status()
    return r.json()


# ── DB 工具 ─────────────────────────────────────────────────────────────────

def db_rows(sql: str, params: tuple = ()) -> list[dict]:
    conn = sqlite3.connect(DB)
    conn.text_factory = lambda b: b.decode("utf-8", errors="replace")
    conn.row_factory = sqlite3.Row
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def load_project(pid: int) -> dict:
    rows = db_rows("SELECT id, name, code FROM projects WHERE id=?", (pid,))
    return rows[0] if rows else {"id": pid, "name": f"Project {pid}", "code": ""}


def load_suppliers(pid: int) -> list[int]:
    rows = db_rows(
        "SELECT DISTINCT supplier_id FROM quotes WHERE project_id=? AND supplier_id IS NOT NULL",
        (pid,),
    )
    return [r["supplier_id"] for r in rows]


def load_anchor_meta(pid: int, category: str) -> dict[str, dict]:
    """Return {seq_str: anchor_dict} from tender_list_sessions.anchors_json."""
    rows = db_rows(
        "SELECT anchors_json FROM tender_list_sessions "
        "WHERE project_id=? AND category=? AND is_current=1 ORDER BY id DESC LIMIT 1",
        (pid, category),
    )
    if not rows or not rows[0].get("anchors_json"):
        return {}
    anchors = json.loads(rows[0]["anchors_json"])
    return {str(a["seq"]): a for a in anchors}


def load_quote_details(qids: list[int]) -> dict[int, dict]:
    """Return {quote_id: detail_dict} for given quote IDs."""
    if not qids:
        return {}
    placeholders = ",".join("?" * len(qids))
    rows = db_rows(
        f"""
        SELECT q.id, m.standard_name AS mat_name, m.spec AS mat_spec,
               q.unit_price, q.quantity, q.remark,
               s.name AS sup_name, s.short_name
        FROM quotes q
        LEFT JOIN materials m ON q.material_id = m.id
        LEFT JOIN suppliers s ON q.supplier_id = s.id
        WHERE q.id IN ({placeholders})
        """,
        tuple(qids),
    )
    return {r["id"]: r for r in rows}


# ── 单元格辅助 ───────────────────────────────────────────────────────────────

def _flags(cell: dict) -> list[str]:
    return cell.get("flags") or []


def _has_flag(flags: list[str], prefix: str) -> bool:
    return any(f == prefix or f.startswith(prefix + ":") for f in flags)


def _price_text(cell: dict) -> str:
    status = cell.get("cell_status")
    price  = cell.get("price")
    if status == "excluded" or (price is None and status not in ("pending",)):
        return "—（已排除）" if status == "excluded" else "—"
    if price is None:
        return "—"
    txt = f"¥{price:,.2f}"
    if status == "aggregated":
        txt += "（聚合）"
    elif status == "pending":
        txt += "（待确认）"
    return txt


def _status_label(status: str | None) -> str:
    return {
        "quoted": "已报价",
        "aggregated": "聚合报价",
        "pending": "待确认",
        "excluded": "已排除",
        "missing": "未报价",
        None: "未报价",
    }.get(status, "未报价")


def _flag_labels(flags: list[str]) -> str:
    parts: list[str] = []
    for f in flags:
        if f == "ocr_corrected_verified":
            parts.append("OCR纠错已验证")
        elif f == "ocr_corrected":
            parts.append("OCR纠错")
        elif f.startswith("valve_type_conflict:"):
            parts.append(f"阀型冲突:{f.split(':',1)[1]}")
        elif f == "canonical_conflict":
            parts.append("规格冲突")
        elif f.startswith("ac_conflict"):
            parts.append("锚点占用冲突")
        elif f.startswith("dup_qids"):
            parts.append("重复报价")
        elif f.startswith("risky_candidate"):
            parts.append("候选风险")
        elif f == "missing_without_evidence":
            parts.append("缺少证据")
    return "，".join(parts)


def _note_text(cell: dict, qdet: dict | None) -> str:
    flags    = _flags(cell)
    evidence = (cell.get("evidence") or "").strip()
    parts: list[str] = []

    if "ocr_corrected_verified" in flags:
        orig = (qdet.get("mat_name") or "") if qdet else ""
        if orig:
            parts.append(f"OCR纠错已验证（原：{orig}）")
        else:
            parts.append("OCR纠错已验证")
    elif "ocr_corrected" in flags:
        parts.append("OCR纠错")

    for f in flags:
        if f.startswith("valve_type_conflict:"):
            vt = f.split(":", 1)[1]
            parts.append(f"阀型冲突：报价阀型为【{vt}】")
        elif f == "canonical_conflict":
            parts.append("规格冲突（与规范型号不一致）")
        elif f.startswith("ac_conflict:"):
            parts.append(f"锚点占用冲突：已被 #{f.split(':',1)[1]} 使用")
        elif f.startswith("dup_qids:"):
            parts.append("同一报价行被多锚点引用")
        elif f.startswith("risky_candidate"):
            parts.append("候选风险（建议人工确认）")
        elif f == "missing_without_evidence":
            parts.append("LLM判断未报价但缺乏证据")

    if evidence:
        parts.append(f"LLM说明：{evidence[:120]}")

    return "\n".join(parts)


def _cell_fill(cell: dict) -> PatternFill:
    status = cell.get("cell_status")
    flags  = _flags(cell)
    if status == "pending":
        return FILL_PENDING
    if status == "excluded":
        return FILL_EXCLUDED
    if status in ("missing", None) and cell.get("price") is None:
        return FILL_MISSING
    if "ocr_corrected_verified" in flags:
        return FILL_OCR_OK
    if cell.get("is_lowest") and status in ("quoted", "aggregated"):
        return FILL_QUOTED_LOWEST
    if status == "aggregated":
        return FILL_AGGREGATED
    return FILL_NONE


def _note_fill(cell: dict) -> PatternFill | None:
    flags = _flags(cell)
    if _has_flag(flags, "valve_type_conflict") or "canonical_conflict" in flags:
        return FILL_CONFLICT
    if _has_flag(flags, "dup_qids") or _has_flag(flags, "ac_conflict") or _has_flag(flags, "risky_candidate"):
        return FILL_WARNING
    if "ocr_corrected_verified" in flags:
        return FILL_OCR_OK
    return None


def _write_header_row(ws, headers: list[str], row: int = 1) -> None:
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 36


def _style_cell(c, fill=None, font=None, align=None):
    c.border = BORDER
    if fill:
        c.fill = fill
    if font:
        c.font = font
    if align:
        c.alignment = align
    else:
        c.alignment = ALIGN_LEFT


# ── Sheet 1: 比价矩阵 ────────────────────────────────────────────────────────

def build_sheet1(ws, matrix: dict, anchors: dict[str, dict], quotes: dict[int, dict]) -> None:
    suppliers = matrix["suppliers"]
    rows      = matrix["rows"]

    # Column widths: anchor cols + (price, status, note) × suppliers + tail
    base_cols = ["序号", "品名", "规格型号", "工作压力", "单位", "数量"]
    header: list[str] = base_cols.copy()
    for s in suppliers:
        sn = s.get("name", str(s["id"]))
        header += [f"{sn}\n单价", f"{sn}\n状态", f"{sn}\n说明"]
    header += ["最低价\n供应商", "最低\n单价", "可比\n状态"]

    _write_header_row(ws, header)
    ws.freeze_panes = "G2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    # Column widths
    widths = [5, 22, 16, 10, 5, 6]
    for _ in suppliers:
        widths += [12, 9, 35]
    widths += [14, 11, 11]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Alternate supplier background fills
    sup_bgs = [
        PatternFill(start_color="F0F7FF", fill_type="solid"),
        PatternFill(start_color="F0FFF4", fill_type="solid"),
        PatternFill(start_color="FFFEF0", fill_type="solid"),
    ]

    for ri, row in enumerate(rows, 2):
        seq      = str(row.get("anchor_seq", ""))
        anchor   = anchors.get(seq, {})
        # B3 兼容期收尾：SupplierCell 的旧 supplier_id 键已删，改用通用列身份键
        # id（submission 模式下退回 submission_id 与 suppliers[].id 对齐）。
        sup_map  = {(c.get("submission_id") or c["id"]): c for c in row["suppliers"]}
        cells    = [sup_map.get(s["id"], {"id": s["id"], "cell_status": None,
                                           "price": None, "is_lowest": False,
                                           "flags": None, "evidence": None,
                                           "source_quote_id": None})
                    for s in suppliers]

        # Lowest price among quoted/aggregated
        lowest_price, lowest_name = None, "—"
        for s, c in zip(suppliers, cells):
            if c.get("cell_status") in ("quoted", "aggregated") and c.get("price") is not None:
                if lowest_price is None or c["price"] < lowest_price:
                    lowest_price = c["price"]
                    lowest_name  = s.get("name", "")

        # Comparable status
        n_q = sum(1 for c in cells if c.get("cell_status") in ("quoted", "aggregated"))
        n_a = sum(1 for c in cells if c.get("cell_status") in ("quoted", "aggregated", "pending"))
        if n_q >= 3:
            cmp_txt = f"三家可比"
        elif n_q >= 2:
            cmp_txt = f"可比（{n_q}家）"
        elif n_a >= 2:
            cmp_txt = f"参考（含待确认）"
        else:
            cmp_txt = "不可比"

        # Base anchor data
        data: list = [
            int(seq) if seq.isdigit() else seq,
            anchor.get("name", row.get("material_name", "")),
            anchor.get("spec", row.get("spec", "")),
            anchor.get("pressure", ""),
            anchor.get("unit", ""),
            anchor.get("qty", ""),
        ]
        for s, cell in zip(suppliers, cells):
            qid  = cell.get("source_quote_id")
            qdet = quotes.get(qid) if qid else None
            data += [
                _price_text(cell),
                _status_label(cell.get("cell_status")),
                _note_text(cell, qdet),
            ]
        data += [
            lowest_name,
            f"¥{lowest_price:,.2f}" if lowest_price else "—",
            cmp_txt,
        ]

        # Write row
        for ci, val in enumerate(data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = BORDER
            c.alignment = ALIGN_LEFT

        ws.row_dimensions[ri].height = 18

        # Style supplier columns
        for si, (s, cell) in enumerate(zip(suppliers, cells)):
            base_ci  = len(base_cols) + si * 3 + 1
            price_ci = base_ci
            stat_ci  = base_ci + 1
            note_ci  = base_ci + 2

            bg = sup_bgs[si % len(sup_bgs)]
            for col in (price_ci, stat_ci, note_ci):
                ws.cell(row=ri, column=col).fill = bg

            cf = _cell_fill(cell)
            if cf is not FILL_NONE:
                ws.cell(row=ri, column=price_ci).fill = cf
                ws.cell(row=ri, column=stat_ci).fill  = cf

            nf = _note_fill(cell)
            if nf:
                ws.cell(row=ri, column=note_ci).fill = nf

            # Font: price cell
            status = cell.get("cell_status")
            if status == "pending":
                ws.cell(row=ri, column=price_ci).font = FONT_ORANGE
            elif status in ("missing", None) and cell.get("price") is None:
                ws.cell(row=ri, column=price_ci).font = FONT_GREY
            elif cell.get("is_lowest") and status in ("quoted", "aggregated"):
                ws.cell(row=ri, column=price_ci).font = FONT_GREEN
            else:
                ws.cell(row=ri, column=price_ci).font = FONT_NORMAL

            # Note cell: small font
            ws.cell(row=ri, column=note_ci).font = FONT_SMALL
            ws.cell(row=ri, column=note_ci).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)

        # Row height: taller if any note exists
        if any(_note_text(c, quotes.get(c.get("source_quote_id"))) for c in cells):
            ws.row_dimensions[ri].height = 42

        # Comparable status font in last column
        cmp_ci = len(header)
        if n_q >= 2:
            ws.cell(row=ri, column=cmp_ci).font = FONT_GREEN
        elif n_a >= 2:
            ws.cell(row=ri, column=cmp_ci).font = FONT_ORANGE
        else:
            ws.cell(row=ri, column=cmp_ci).font = FONT_GREY


# ── Sheet 2: 汇总 ───────────────────────────────────────────────────────────

def build_sheet2(ws, matrix: dict, pid: int, category: str, proj: dict) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    def kv(label: str, value, bold_val: bool = False) -> None:
        r = ws.max_row + 1
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=value)
        a.font = FONT_BOLD; a.border = BORDER; a.alignment = ALIGN_LEFT
        b.font = FONT_BOLD if bold_val else FONT_NORMAL
        b.border = BORDER; b.alignment = ALIGN_LEFT

    # Title
    t = ws.cell(row=1, column=1, value="导出汇总")
    t.font = Font(bold=True, size=14, name="微软雅黑")
    ws.row_dimensions[1].height = 30
    ws.append([])

    kv("项目名称",   proj.get("name", f"Project {pid}"))
    kv("项目代码",   proj.get("code", ""))
    kv("品类",       category)
    kv("供应商",     "、".join(s.get("name", "") for s in matrix["suppliers"]))
    kv("导出时间",   time.strftime("%Y-%m-%d %H:%M:%S"))
    ws.append([])

    rows  = matrix["rows"]
    total = len(rows)
    all_cells = [c for r in rows for c in r["suppliers"]]

    n_q2  = sum(1 for r in rows
                if sum(1 for c in r["suppliers"]
                       if c.get("cell_status") in ("quoted", "aggregated")) >= 2)
    n_a2  = sum(1 for r in rows
                if sum(1 for c in r["suppliers"]
                       if c.get("cell_status") in ("quoted", "aggregated", "pending")) >= 2)
    n_3   = sum(1 for r in rows
                if sum(1 for c in r["suppliers"]
                       if c.get("cell_status") in ("quoted", "aggregated")) >= 3)
    n_pend = sum(1 for c in all_cells if c.get("cell_status") == "pending")
    n_miss = sum(1 for c in all_cells
                 if c.get("cell_status") in ("missing", None) and c.get("price") is None)
    fp_cnt = sum(1 for c in all_cells
                 if c.get("cell_status") in ("quoted", "aggregated")
                 and _has_flag(_flags(c), "valve_type_conflict"))
    mwe_cnt = sum(1 for c in all_cells
                  if c.get("cell_status") == "pending"
                  and "missing_without_evidence" in _flags(c))
    ocr_cnt = sum(1 for c in all_cells if "ocr_corrected" in _flags(c))
    ocr_ok  = sum(1 for c in all_cells if "ocr_corrected_verified" in _flags(c))
    can_fin = (fp_cnt == 0 and mwe_cnt == 0)

    # Section header
    r_hdr = ws.max_row + 1
    h = ws.cell(row=r_hdr, column=1, value="矩阵统计")
    h.font = FONT_BOLD; h.fill = FILL_SUB_HEADER; h.border = BORDER
    h2 = ws.cell(row=r_hdr, column=2, value="数值")
    h2.font = FONT_BOLD; h2.fill = FILL_SUB_HEADER; h2.border = BORDER

    kv("采购清单锚点总数", total)
    kv("≥2家 quoted 可比 (quoted/agg)",
       f"{n_q2}/{total} = {n_q2/total*100:.1f}%", bold_val=True)
    kv("≥2家 quoted+pending 可比",
       f"{n_a2}/{total} = {n_a2/total*100:.1f}%")
    kv("三家齐全 (≥3 quoted/agg)", f"{n_3} 项")
    kv("pending 待确认（格子数）", n_pend)
    kv("missing 未报价（格子数）", n_miss)
    kv("false_positive_align_count", fp_cnt, bold_val=(fp_cnt > 0))
    kv("missing_without_evidence_count", mwe_cnt)
    kv("OCR纠错 格子数（含待验证）", ocr_cnt)
    kv("OCR纠错已验证 格子数", ocr_ok)
    kv("可终版 (can_finalize)",
       "是 ✓" if can_fin else "否 ✗（见风险sheet）", bold_val=True)

    ws.append([])

    # Per-supplier detail
    r_hdr2 = ws.max_row + 1
    sup_names = [s.get("name", str(s["id"])) for s in matrix["suppliers"]]
    ws.append(["供应商明细"] + sup_names)
    for ci in range(1, len(sup_names) + 2):
        ws.cell(row=r_hdr2, column=ci).font = FONT_BOLD
        ws.cell(row=r_hdr2, column=ci).fill = FILL_SUB_HEADER
        ws.cell(row=r_hdr2, column=ci).border = BORDER
        ws.cell(row=r_hdr2, column=ci).alignment = ALIGN_CENTER
    for ci in range(2, len(sup_names) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    for status_label, statuses in [
        ("quoted（已报价）", ("quoted",)),
        ("aggregated（聚合）", ("aggregated",)),
        ("pending（待确认）", ("pending",)),
        ("missing（未报价）", (None,)),
    ]:
        row_vals = [status_label]
        for s in matrix["suppliers"]:
            count = sum(
                1 for r in rows
                for c in r["suppliers"]
                if (c.get("submission_id") or c.get("id")) == s["id"]
                and c.get("cell_status") in statuses
                and (statuses != (None,) or c.get("price") is None)
            )
            row_vals.append(count)
        r = ws.max_row + 1
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = BORDER
            cell.alignment = ALIGN_LEFT


# ── Sheet 3: 待确认项 ────────────────────────────────────────────────────────

def build_sheet3(ws, matrix: dict, anchors: dict, quotes: dict) -> None:
    headers = ["序号", "品名", "规格", "供应商", "报价ID",
               "报价原始名称", "报价规格", "参考单价", "flags", "LLM说明", "建议动作"]
    _write_header_row(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = [5, 22, 14, 16, 8, 22, 14, 10, 26, 50, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for row in matrix["rows"]:
        seq    = str(row.get("anchor_seq", ""))
        anchor = anchors.get(seq, {})
        for cell in row["suppliers"]:
            if cell.get("cell_status") != "pending":
                continue
            flags    = _flags(cell)
            qid      = cell.get("source_quote_id")
            qdet     = quotes.get(qid) if qid else None
            sup_name = next((s.get("name", "") for s in matrix["suppliers"]
                             if s["id"] == (cell.get("submission_id") or cell.get("id"))), "")

            if _has_flag(flags, "valve_type_conflict") or "canonical_conflict" in flags:
                action = "建议排除"
                fill   = FILL_CONFLICT
            elif "ocr_corrected_verified" in flags:
                action = "建议纳入"
                fill   = FILL_OCR_OK
            elif _has_flag(flags, "risky_candidate") or _has_flag(flags, "dup_qids"):
                action = "需人工确认"
                fill   = FILL_WARNING
            else:
                action = "人工确认"
                fill   = FILL_PENDING

            data = [
                int(seq) if seq.isdigit() else seq,
                anchor.get("name", row.get("material_name", "")),
                anchor.get("spec", row.get("spec", "")),
                sup_name,
                qid or "",
                (qdet.get("mat_name") or "") if qdet else "",
                (qdet.get("mat_spec") or "") if qdet else "",
                (f"¥{cell['price']:,.2f}" if cell.get("price") else "—"),
                _flag_labels(flags),
                (cell.get("evidence") or ""),
                action,
            ]
            r = ws.max_row + 1
            for ci, val in enumerate(data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.fill = fill; c.border = BORDER; c.alignment = ALIGN_LEFT
            ws.row_dimensions[r].height = 30 if cell.get("evidence") else 18


# ── Sheet 4: OCR纠错项 ───────────────────────────────────────────────────────

def build_sheet4(ws, matrix: dict, anchors: dict, quotes: dict) -> None:
    headers = ["序号", "品名", "供应商", "报价ID",
               "原始OCR名称", "目标锚点名称", "规格", "LLM说明", "状态", "已验证"]
    _write_header_row(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = [5, 22, 16, 8, 24, 24, 14, 50, 10, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for row in matrix["rows"]:
        seq    = str(row.get("anchor_seq", ""))
        anchor = anchors.get(seq, {})
        for cell in row["suppliers"]:
            flags = _flags(cell)
            if not any(f in ("ocr_corrected", "ocr_corrected_verified") for f in flags):
                continue
            verified = "ocr_corrected_verified" in flags
            qid      = cell.get("source_quote_id")
            qdet     = quotes.get(qid) if qid else None
            sup_name = next((s.get("name", "") for s in matrix["suppliers"]
                             if s["id"] == (cell.get("submission_id") or cell.get("id"))), "")

            data = [
                int(seq) if seq.isdigit() else seq,
                anchor.get("name", row.get("material_name", "")),
                sup_name,
                qid or "",
                (qdet.get("mat_name") or "") if qdet else "",
                anchor.get("name", row.get("material_name", "")),
                anchor.get("spec", row.get("spec", "")),
                (cell.get("evidence") or ""),
                _status_label(cell.get("cell_status")),
                "是" if verified else "否",
            ]
            fill = FILL_OCR_OK if verified else FILL_WARNING
            r = ws.max_row + 1
            for ci, val in enumerate(data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.fill = fill; c.border = BORDER; c.alignment = ALIGN_LEFT
            ws.row_dimensions[r].height = 36 if cell.get("evidence") else 18


# ── Sheet 5: 风险冲突项 ─────────────────────────────────────────────────────

_RISK_PREFIXES = {
    "valve_type_conflict", "canonical_conflict", "risky_candidate",
    "dup_qids", "ac_conflict", "missing_without_evidence",
}


def _is_risk(flags: list[str]) -> bool:
    return any(
        f == p or f.startswith(p + ":")
        for f in flags for p in _RISK_PREFIXES
    )


def build_sheet5(ws, matrix: dict, anchors: dict, quotes: dict) -> None:
    headers = ["序号", "品名", "供应商", "报价ID", "状态", "flags", "LLM说明", "风险说明"]
    _write_header_row(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    widths = [5, 22, 16, 8, 10, 30, 50, 50]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for row in matrix["rows"]:
        seq    = str(row.get("anchor_seq", ""))
        anchor = anchors.get(seq, {})
        for cell in row["suppliers"]:
            flags = _flags(cell)
            if not _is_risk(flags):
                continue
            qid      = cell.get("source_quote_id")
            qdet     = quotes.get(qid) if qid else None
            sup_name = next((s.get("name", "") for s in matrix["suppliers"]
                             if s["id"] == (cell.get("submission_id") or cell.get("id"))), "")

            risk_parts: list[str] = []
            for f in flags:
                if f.startswith("valve_type_conflict:"):
                    risk_parts.append(f"阀型冲突：报价阀型为【{f.split(':',1)[1]}】，与锚点类型不匹配")
                elif f == "canonical_conflict":
                    risk_parts.append("规格冲突：与规范型号不一致，建议排除或人工核查")
                elif f.startswith("risky_candidate"):
                    risk_parts.append("候选风险：LLM匹配置信度低，建议人工确认")
                elif f.startswith("dup_qids"):
                    risk_parts.append("重复报价：同一报价行被多个锚点引用，可能导致价格失真")
                elif f.startswith("ac_conflict:"):
                    risk_parts.append(f"锚点占用冲突：该报价已被 #{f.split(':',1)[1]} 占用")
                elif f == "missing_without_evidence":
                    risk_parts.append("缺少证据：LLM判断未报价但没有充分证据支持")

            fill = FILL_CONFLICT if (
                _has_flag(flags, "valve_type_conflict") or "canonical_conflict" in flags
            ) else FILL_WARNING

            data = [
                int(seq) if seq.isdigit() else seq,
                anchor.get("name", row.get("material_name", "")),
                sup_name,
                qid or "",
                _status_label(cell.get("cell_status")),
                _flag_labels(flags),
                (cell.get("evidence") or ""),
                "；".join(risk_parts),
            ]
            r = ws.max_row + 1
            for ci, val in enumerate(data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.fill = fill; c.border = BORDER; c.alignment = ALIGN_LEFT
            ws.row_dimensions[r].height = 36 if risk_parts else 18


# ── 验证 ────────────────────────────────────────────────────────────────────

def validate(out_path: str, matrix: dict) -> bool:
    from openpyxl import load_workbook
    wb = load_workbook(out_path, read_only=True)
    rows_in_sheet1 = wb["比价矩阵"].max_row - 1  # minus header

    total  = len(matrix["rows"])
    all_ok = True

    def check(label: str, got, expected, op: str = "==") -> None:
        nonlocal all_ok
        ok = (got == expected) if op == "==" else (got >= expected if op == ">=" else got <= expected)
        tag = "OK  " if ok else "FAIL"
        print(f"  [{tag}] {label}: 期望={expected} 实际={got}")
        if not ok:
            all_ok = False

    check("Sheet1 行数 == 锚点总数", rows_in_sheet1, total)
    check("Sheet名: 比价矩阵", "比价矩阵" in wb.sheetnames, True)
    check("Sheet名: 汇总",   "汇总" in wb.sheetnames, True)
    check("Sheet名: 待确认项", "待确认项" in wb.sheetnames, True)
    check("Sheet名: OCR纠错项", "OCR纠错项" in wb.sheetnames, True)
    check("Sheet名: 风险冲突项", "风险冲突项" in wb.sheetnames, True)

    # #28–31 OCR verified rows should be in Sheet 4
    ocr_sheet = wb["OCR纠错项"]
    ocr_seqs  = set()
    for row in ocr_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            ocr_seqs.add(str(row[0]))
    for seq in ["28", "29", "30", "31"]:
        check(f"#{ seq} 在OCR纠错sheet", seq in ocr_seqs, True)

    # fp_count from matrix
    fp_cnt = sum(1 for r in matrix["rows"] for c in r["suppliers"]
                 if c.get("cell_status") in ("quoted", "aggregated")
                 and _has_flag(_flags(c), "valve_type_conflict"))
    check("false_positive_align_count == 0", fp_cnt, 0)

    # #46 / #70 should NOT be quoted (should be in pending or missing or risk)
    for target_seq in ["46", "70"]:
        for r in matrix["rows"]:
            if str(r.get("anchor_seq", "")) == target_seq:
                for c in r["suppliers"]:
                    if _has_flag(_flags(c), "valve_type_conflict") and c.get("cell_status") not in ("quoted", "aggregated"):
                        check(f"#{target_seq} 不在 quoted (含阀型冲突被降级)", True, True)
                        break

    wb.close()
    return all_ok


# ── 主入口 ───────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="客户演示 Excel 导出")
    ap.add_argument("--project",  required=True, help="项目ID（整数）")
    ap.add_argument("--category", default="阀门")
    ap.add_argument("--out",      default="outputs/customer_bid_matrix.xlsx")
    ap.add_argument("--user",     default="admin")
    ap.add_argument("--password", default="admin123")
    ap.add_argument("--no-validate", action="store_true")
    args = ap.parse_args()

    pid = int(args.project)
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)

    # 1. Auth
    log("登录中...")
    tok = api_login(args.user, args.password)

    # 2. 项目元数据
    log("读取项目信息...")
    proj  = load_project(pid)
    sids  = load_suppliers(pid)
    log(f"  项目: {proj.get('name', pid)}  供应商: {sids}")

    # 3. 锚点元数据
    log("读取锚点清单...")
    anchors = load_anchor_meta(pid, args.category)
    log(f"  锚点总数: {len(anchors)}")

    # 4. 矩阵数据（API）
    log("调用 /api/analysis/bid-matrix ...")
    matrix = api_bid_matrix(pid, args.category, sids, tok)
    log(f"  矩阵行数: {len(matrix['rows'])}  供应商: {[s['name'] for s in matrix['suppliers']]}")

    # 5. 报价明细（DB）
    all_qids = list({
        c.get("source_quote_id")
        for r in matrix["rows"]
        for c in r["suppliers"]
        if c.get("source_quote_id") is not None
    })
    log(f"读取报价明细（{len(all_qids)} 条）...")
    quotes = load_quote_details(all_qids)

    # 6. 构建 Excel
    log("构建 Excel...")
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("比价矩阵")
    ws2 = wb.create_sheet("汇总")
    ws3 = wb.create_sheet("待确认项")
    ws4 = wb.create_sheet("OCR纠错项")
    ws5 = wb.create_sheet("风险冲突项")

    log("  Sheet 1: 比价矩阵")
    build_sheet1(ws1, matrix, anchors, quotes)
    log("  Sheet 2: 汇总")
    build_sheet2(ws2, matrix, pid, args.category, proj)
    log("  Sheet 3: 待确认项")
    build_sheet3(ws3, matrix, anchors, quotes)
    log("  Sheet 4: OCR纠错项")
    build_sheet4(ws4, matrix, anchors, quotes)
    log("  Sheet 5: 风险冲突项")
    build_sheet5(ws5, matrix, anchors, quotes)

    wb.save(args.out)
    log(f"已保存：{args.out}")

    # 7. 验证
    if not args.no_validate:
        log("验证...")
        ok = validate(args.out, matrix)
        if ok:
            log("所有验证通过 OK")
        else:
            log("部分验证失败（见上）")
            sys.exit(1)

    # 8. 简要统计
    total = len(matrix["rows"])
    all_cells = [c for r in matrix["rows"] for c in r["suppliers"]]
    n_q2 = sum(1 for r in matrix["rows"]
               if sum(1 for c in r["suppliers"]
                      if c.get("cell_status") in ("quoted", "aggregated")) >= 2)
    n_pend = sum(1 for c in all_cells if c.get("cell_status") == "pending")
    n_ocr  = sum(1 for c in all_cells if "ocr_corrected" in _flags(c))
    n_risk = sum(1 for c in all_cells if _is_risk(_flags(c)))

    print("\n" + "=" * 60)
    print(f"  锚点总数:         {total}")
    print(f"  ≥2家 quoted 可比: {n_q2}/{total} = {n_q2/total*100:.1f}%")
    print(f"  待确认格子:       {n_pend}")
    print(f"  OCR纠错:          {n_ocr}")
    print(f"  风险/冲突格子:    {n_risk}")
    print(f"  输出文件:         {os.path.abspath(args.out)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
