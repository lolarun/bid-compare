"""招标清单解析 — 结构化 Excel → 锚点行(TenderAnchor)。

这是 `docs/design/05-bid-comparison-intelligence-layers.md` §9 第 1 步:把招标清单(工程量
清单)解析成一组"锚点行",作为比价矩阵的纵轴骨架与供应商报价匹配的目标。

第一版只支持**规范表头**(列名可识别;材质可拆多子列)。表头怪异(无标题行、
多级合并到无法按列名识别)的情况退 LLM 读表头,留作后续迭代——见设计文档 §8。

解析是纯代码、确定性的(智能化程度极低,见设计文档 §2),不调用任何 LLM。

通用性:不写死列位置,靠**列名同义识别**列。任何品类的清单只要是带可识别表头
的表格即可解析;品类差异体现在"材质/扩展属性"这些 freeform 列,匹配阶段再用。
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

import openpyxl

from apps.api.services.ingestion.canonical import extract_valve_canonical


# ── 列名同义映射:把表头单元格文本归一到标准字段 ──────────────────────────
# 第一版规范表头;识别不到的列进 raw(freeform),匹配阶段可用。
_COL_SYNONYMS: dict[str, tuple[str, ...]] = {
    "seq":        ("序号", "序", "编号"),
    "profession": ("专业",),
    "name":       ("项目名称", "材料名称", "材料(设备)名称", "材料（设备）名称", "名称", "品名", "货物名称"),
    "spec":       ("规格", "规格型号"),
    "model":      ("型号",),
    "pressure":   ("公称压力", "压力", "PN"),
    "material":   ("材质",),
    "unit":       ("单位", "计量单位"),
    "qty":        ("数量", "工程量", "招标数量"),
    "brand":      ("品牌",),
    "remark":     ("备注", "说明"),
}

# 材质常见子列(多列归一到一个 materials 字典)。识别不到子标签则按位置编号。
_MATERIAL_SUBCOLS = ("阀体", "阀芯", "阀板", "阀座", "密封圈", "材料", "本体", "衬里")

# 数据区终止信号:序号列出现这些文本视为表尾(合计/说明行)。
_FOOTER_MARKERS = ("含税", "合价", "合计", "说明", "备注：", "总计", "小计")


@dataclass
class TenderAnchor:
    """一条招标清单锚点行。"""

    seq: Any                              # 序号(原值)
    name: str                            # 项目/材料名称
    spec: str = ""                       # 规格(常含 DN)
    model: str = ""                      # 型号
    pressure: str = ""                   # 公称压力(常含 PN)
    materials: dict[str, str] = field(default_factory=dict)  # {子列名: 材质文本}
    unit: str = ""
    qty: float | None = None
    brand: str = ""
    profession: str = ""
    remark: str = ""
    row_index: int = -1                  # 源行号(0-based),便于回溯
    raw: dict[str, str] = field(default_factory=dict)        # 未识别列的 freeform
    source_ref: dict = field(default_factory=dict)           # PDF 来源 {page, row} 便于追溯

    canonical: dict = field(default_factory=dict)  # valve canonical key for hard-filter matching

    def material_text(self) -> str:
        """材质拼成单串(供匹配/展示)。"""
        return "/".join(v for v in self.materials.values() if v)


def rebuild_anchors(session) -> list[TenderAnchor]:
    """从 TenderListSession.anchors_json 重建 TenderAnchor 列表。

    取代 routes/analysis.py 中两处重复的重建逻辑。优先用存储的 canonical(含
    valve_type)，否则用完整字段重算。
    """
    from apps.api.services.ingestion.canonical import extract_valve_canonical

    anchors: list[TenderAnchor] = []
    for a in (session.anchors_json or []):
        ta = TenderAnchor(
            seq=int(a.get("seq") or 0),
            name=str(a.get("name") or ""),
            spec=str(a.get("spec") or ""),
            model=str(a.get("model") or ""),
            pressure=str(a.get("pressure") or ""),
            materials=dict(a.get("materials") or {}),
            unit=str(a.get("unit") or ""),
            qty=float(a.get("qty") or 0) or None,
            brand=str(a.get("brand") or ""),
            profession=str(a.get("profession") or ""),
            remark=str(a.get("remark") or ""),
            source_ref=dict(a.get("source_ref") or {}),
        )
        stored_canon = a.get("canonical")
        if stored_canon and isinstance(stored_canon, dict) and stored_canon.get("valve_type"):
            ta.canonical = stored_canon
        else:
            ta.canonical = extract_valve_canonical(
                ta.name, ta.spec, ta.pressure, ta.material_text()
            )
        anchors.append(ta)
    return anchors


def anchor_to_json(anchor: "TenderAnchor", category: str | None = None) -> dict:
    """序列化 TenderAnchor → anchors_json 字典(rebuild_anchors 可还原)。

    category: 显式品类(classify 结果)；None 时即时识别。
    """
    if category is None:
        from apps.api.services.ingestion.category_classify import classify_category
        category = classify_category(
            anchor.name, anchor.spec, anchor.pressure, anchor.material_text()
        ).category
    return {
        "seq": str(anchor.seq),
        "name": anchor.name,
        "spec": anchor.spec,
        "model": anchor.model,
        "pressure": anchor.pressure,
        "materials": anchor.materials,
        "unit": anchor.unit,
        "qty": anchor.qty,
        "brand": anchor.brand,
        "profession": anchor.profession,
        "remark": anchor.remark,
        "category": category,
        "canonical": anchor.canonical,
        "source_ref": anchor.source_ref,
    }


def group_anchors_by_category(
    anchors: list["TenderAnchor"], default_category: str = "",
) -> dict[str, list[dict]]:
    """按品类分组锚点并序列化。unknown(空品类)回退到 default_category。

    Returns {category: [anchor_json, ...]}。default_category 为空且存在 unknown
    时，unknown 锚点被丢弃(调用方应先保证有默认品类)。
    """
    from apps.api.services.ingestion.category_classify import classify_category

    groups: dict[str, list[dict]] = {}
    for a in anchors:
        cat = classify_category(a.name, a.spec, a.pressure, a.material_text()).category
        if not cat:
            cat = default_category
        if not cat:
            continue
        groups.setdefault(cat, []).append(anchor_to_json(a, cat))
    return groups


@dataclass
class SheetInfo:
    """design/24 B1：一个 Sheet 的采购清单候选摘要，供预览的 Sheet 切换器用。"""

    name: str
    looks_like_list: bool   # 能否找到规范表头(_find_header_row 命中)
    row_count: int          # looks_like_list=True 时是数据行数，否则是非空单元格粗计数


def _load_sheet_rows(
    source: str | bytes | io.BytesIO, sheet: str | None = None,
) -> tuple[list[list], str]:
    """打开 workbook，取指定 Sheet(或默认 active)的原始行。

    Returns:
        (rows, resolved_sheet_name)
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        return rows, ws.title
    finally:
        wb.close()


def list_tender_sheets(source: str | bytes | io.BytesIO) -> list[SheetInfo]:
    """design/24 B1：列出 Excel 全部 Sheet，标注哪些像采购清单。

    "像清单" = 能找到规范表头(_find_header_row 命中)——与 parse_tender_xlsx
    判定表头的规则完全同一份，不另起一套"猜测"逻辑。真正需要解析出锚点时仍
    只应调用 parse_tender_xlsx，这里只给预览用的候选摘要。
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        out: list[SheetInfo] = []
        for name in wb.sheetnames:
            rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
            header_idx = _find_header_row(rows) if rows else None
            if header_idx is None:
                non_empty = sum(
                    1 for r in rows if any(c is not None and str(c).strip() for c in r)
                )
                out.append(SheetInfo(name=name, looks_like_list=False, row_count=non_empty))
                continue
            colmap, mat_cols, has_subheader = _map_columns(rows, header_idx)
            data_start = header_idx + (2 if has_subheader else 1)
            n = 0
            for ri in range(data_start, len(rows)):
                row = rows[ri]
                seq = _cell(row, colmap.get("seq"))
                if seq is None or str(seq).strip() == "":
                    continue
                if any(m in str(seq) for m in _FOOTER_MARKERS):
                    break
                name_v = _cell(row, colmap.get("name"))
                if name_v is None or str(name_v).strip() == "":
                    continue
                n += 1
            out.append(SheetInfo(name=name, looks_like_list=True, row_count=n))
        return out
    finally:
        wb.close()


def pick_default_sheet(sheets: list[SheetInfo]) -> str | None:
    """design/24 B1 auto-detect 规则：候选 Sheet 里数据行数最多的那个，不是第一个
    header 匹配的。真实附件常有"汇总表"排在前面且表头形似——选行数最多的更稳。

    没有任何 Sheet 像清单时返回 None，调用方据此决定是否报错。
    """
    candidates = [s for s in sheets if s.looks_like_list]
    if not candidates:
        return None
    return max(candidates, key=lambda s: s.row_count).name


def parse_tender_xlsx(
    source: str | bytes | io.BytesIO, sheet: str | None = None,
) -> list[TenderAnchor]:
    """解析招标清单 xlsx,返回锚点行列表。

    Args:
        source: 文件路径、字节内容或 BytesIO。
        sheet: 指定 Sheet 名；None 时用 workbook 的 active sheet(单 Sheet 文件的
            默认行为不变)。多 Sheet 场景下调用方应先用 list_tender_sheets +
            pick_default_sheet 选定，再传进来——本函数本身不做多 Sheet 探测。

    Returns:
        TenderAnchor 列表(已剔除标题行、表头行、表尾说明行)。

    Raises:
        ValueError: 找不到可识别的表头(规范表头缺失)。
    """
    rows, _ = _load_sheet_rows(source, sheet)
    if not rows:
        raise ValueError("空工作表")

    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError("找不到可识别的表头(第一版仅支持规范表头;序号/名称/数量缺失)")

    header = rows[header_idx]
    colmap, mat_cols, has_subheader = _map_columns(rows, header_idx)

    data_start = header_idx + (2 if has_subheader else 1)
    anchors: list[TenderAnchor] = []
    for ri in range(data_start, len(rows)):
        row = rows[ri]
        seq = _cell(row, colmap.get("seq"))
        # 数据区终止:序号空 或 命中表尾标记
        if seq is None or str(seq).strip() == "":
            continue
        if any(m in str(seq) for m in _FOOTER_MARKERS):
            break
        name = _cell(row, colmap.get("name"))
        if name is None or str(name).strip() == "":
            continue  # 无名称行跳过(分隔/空行)

        materials: dict[str, str] = {}
        for label, ci in mat_cols:
            v = _cell(row, ci)
            if v is not None and str(v).strip():
                materials[label] = str(v).strip()

        # 未识别列 → raw
        raw: dict[str, str] = {}
        known_cols = set(colmap.values()) | {ci for _, ci in mat_cols}
        for ci, hv in enumerate(header):
            if ci in known_cols or hv is None or str(hv).strip() == "":
                continue
            v = _cell(row, ci)
            if v is not None and str(v).strip():
                raw[str(hv).strip()] = str(v).strip()

        anchor = TenderAnchor(
            seq=seq,
            name=str(name).strip(),
            spec=_str(row, colmap.get("spec")),
            model=_str(row, colmap.get("model")),
            pressure=_str(row, colmap.get("pressure")),
            materials=materials,
            unit=_str(row, colmap.get("unit")),
            qty=_num(_cell(row, colmap.get("qty"))),
            brand=_str(row, colmap.get("brand")),
            profession=_str(row, colmap.get("profession")),
            remark=_str(row, colmap.get("remark")),
            row_index=ri,
            raw=raw,
        )
        anchor.canonical = extract_valve_canonical(
            anchor.name, anchor.spec, anchor.pressure, anchor.material_text()
        )
        anchors.append(anchor)
    return anchors


# ── helpers ────────────────────────────────────────────────────────────────
def _find_header_row(rows: list[list], scan_limit: int = 10) -> int | None:
    """表头行 = 前若干行中,同时出现"序号(类)"和"名称(类)"的那一行。"""
    seq_kw = _COL_SYNONYMS["seq"]
    name_kw = _COL_SYNONYMS["name"]
    for i, row in enumerate(rows[:scan_limit]):
        texts = [str(c).strip() for c in row if c is not None]
        has_seq = any(any(k == t for k in seq_kw) for t in texts)
        has_name = any(any(k in t for k in name_kw) for t in texts)
        if has_seq and has_name:
            return i
    return None


def _map_columns(
    rows: list[list], header_idx: int,
) -> tuple[dict[str, int], list[tuple[str, int]], bool]:
    """把表头列映射到标准字段;识别材质多子列(可能用到下一行子表头)。

    Returns:
        (colmap, material_cols, has_subheader)
        colmap: {标准字段: 列号}
        material_cols: [(子列标签, 列号), ...]
        has_subheader: 材质是否使用了下一行作为子表头
    """
    header = rows[header_idx]
    sub = rows[header_idx + 1] if header_idx + 1 < len(rows) else []

    colmap: dict[str, int] = {}
    for ci, cell in enumerate(header):
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        for fld, syns in _COL_SYNONYMS.items():
            if fld == "material":
                continue  # 材质单独处理
            if fld in colmap:
                continue
            if any(s == text for s in syns) or any(s in text for s in syns):
                colmap[fld] = ci
                break

    # 材质列:找到主表头"材质"所在列;若其右侧主表头为空而子表头有标签 → 多子列
    mat_cols: list[tuple[str, int]] = []
    has_subheader = False
    mat_start = None
    for ci, cell in enumerate(header):
        if cell is not None and str(cell).strip() in _COL_SYNONYMS["material"]:
            mat_start = ci
            break
    if mat_start is not None:
        # 向右扩展:主表头为空(被合并)的列归入材质组
        ci = mat_start
        while ci < len(header):
            main = header[ci]
            if ci > mat_start and main is not None and str(main).strip():
                break  # 遇到下一个主表头,材质组结束
            sub_label = ""
            if ci < len(sub) and sub[ci] is not None:
                sub_label = str(sub[ci]).strip()
            if sub_label and sub_label in _MATERIAL_SUBCOLS:
                has_subheader = True
                mat_cols.append((sub_label, ci))
            elif sub_label:
                has_subheader = True
                mat_cols.append((sub_label, ci))
            else:
                # 无子标签:单列材质
                mat_cols.append(("材质", ci))
            ci += 1
        # 去重保序
        seen = set()
        mat_cols = [(l, c) for l, c in mat_cols if not (c in seen or seen.add(c))]

    return colmap, mat_cols, has_subheader


def _cell(row: list, ci: int | None):
    if ci is None or ci >= len(row):
        return None
    return row[ci]


def _str(row: list, ci: int | None) -> str:
    v = _cell(row, ci)
    return "" if v is None else str(v).strip()


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("，", "")
    try:
        return float(s)
    except ValueError:
        return None
