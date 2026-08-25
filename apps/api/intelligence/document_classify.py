"""document_classify.py — Tier 0 自动分类：拖进来的文件是招标/投标/清单哪一种。

design/28 §3。三级判据阶梯的第一级：**瞬时、零模型调用**，只做确定性能判定
的部分，判不了的显式交给 Tier 1（识别后信号，cut 3 落地，本文件尚不含）/
Tier 2（LLM 兜底，cut 6）。这是架构约束，不是偷懒——§2 实测过金桥招标 xlsx
的价格列填充率，真实处于"清单"和"报价单"之间；本层遇到这种情况必须回答
"不确定"而不是猜，猜对了也是错的判据。

只处理 Excel（xlsx/xls）。PDF 在扩展名层面只能确定"是文档"，判不了招标还是
投标——那需要封面标量/表格产物这类识别后才有的信号（design/28 §3 Tier 1），
Tier 0 对 PDF 不作分类判定，调用方应直接进入 Tier 1。
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal

import openpyxl

from apps.api.core.utils import parse_num
from apps.api.intelligence.tender_text_layer import has_usable_text_layer
from apps.api.intelligence.vl_quote import _PRICE_SLOTS, map_columns

# 表头候选行：扫前几行取"槽位命中数最多"的一行当表头，不假定表头必在第 1 行
# ——真实交付文档常见标题行/空行前导（design/28 语料本身没有这个问题，但生
# 产上传不能假设跟语料一样干净）。
_HEADER_SCAN_ROWS = 5

# design/28 §2 实测两个参照点：凯硕新正 60/60=100%（strong）、金桥招标 xlsx
# 32/60≈53%（ambiguous，且是刻意保留、期望分类器答"不确定"的验收样本）。
# 阈值取二者之间、明显偏向 strong 一侧，不卡在某个实测出来的具体百分比上
# ——那是两点拟合，不是分布，卡精确值会把测量误差当成判据。
FILL_RATE_STRONG = 0.90

# 判据的**另一半**（2026-08-23 补）。原来只有「几乎全满 → 报价清单」和「压根没有
# 价格列 → 采购清单」两支，中间缺了最典型的一种形态：**表头有价格列、格子全是空的
# 空白表**。那正是采购清单的定义性特征（design/28 §2：投标方待填），却因为落不进
# 任何一支而被判成"不确定"，然后弹窗、再被误route 成报价清单——金桥采购清单实测
# 就是这样变成比价矩阵里的一列供应商的。
#
# 阈值给 2% 而不是 0：真实空白表里混进一两个示例值（"如：128.00"）是常见的，
# 不该让整份表因此失去"空白"的身份。
FILL_RATE_BLANK = 0.02


ExcelVerdict = Literal["tender_list", "bid_list", "uncertain"]
ExcelConfidence = Literal["definitive", "strong", "ambiguous"]
PdfTextLayer = Literal["native", "scanned"]


@dataclass
class ExcelClassification:
    """一份 Excel 文件的 Tier 0 判定结果。"""

    file_path: str
    verdict: ExcelVerdict
    confidence: ExcelConfidence
    price_columns: list[str] = field(default_factory=list)  # 命中的价格列表头原文
    fill_rate: float | None = None                          # 价格列整体非空率；无价格列时为 None
    row_count: int = 0
    reason: str = ""                                        # 人类可读，供确认屏/审计日志


@dataclass
class PdfClassification:
    """一份 PDF 文件的 Tier 0 判定结果。

    只是"文档"，不判定招标还是投标——那需要封面标量/表格产物这类识别后才
    有的信号（Tier 1）。这里只给出下游要用的一个结构性信号：原生文字层还
    是纯扫描件，决定走文字直抽（design/25 轨A）还是必须过 OCR/VL。
    """

    file_path: str
    text_layer: PdfTextLayer
    reason: str = ""


def classify_pdf(file_path: str) -> PdfClassification:
    native = has_usable_text_layer(file_path)
    return PdfClassification(
        file_path=file_path,
        text_layer="native" if native else "scanned",
        reason=("检测到可用文字层，原生 PDF。" if native
                else "未检测到可用文字层（或纯扫描件），需要走 OCR/VL 识别。"),
    )


def _read_all_rows(file_path: str, sheet: str | None = None) -> list[list[str]]:
    """把表格文件读成 `list[list[str]]`，xlsx/xls 与 csv 走同一个出口。

    CSV 之所以在这里（而不是"CSV 不归 Tier 0 管"）：判据本身跟载体无关——
    看表头有没有价格列、填充率多少。此前 `classify_tier0` 对 `.csv` 返回
    None，路由把它当成"不支持的文件类型"挡在门外，而 `extract_quote_tabular`
    其实读得好好的（实测徐汇四份报价 CSV 各 136 项全部解析成功）。能力在，
    门口拦了，用户看到的是"不支持"。

    编码顺序跟 `tabular_ingestion._load_dataframe` 保持一致：国内供应商导出的
    CSV 有相当比例是 GBK。
    """
    if Path(file_path).suffix.lower() == ".csv":
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                with open(file_path, encoding=enc, newline="") as fh:
                    return [[c.strip() for c in row] for row in csv.reader(fh)]
            except UnicodeDecodeError:
                continue
        return []
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = [[("" if c is None else str(c).strip()) for c in row]
            for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _load_header_and_rows(file_path: str, sheet: str | None = None,
                           ) -> tuple[list[str], list[list[str]]]:
    """读表格文件，在前 `_HEADER_SCAN_ROWS` 行里选槽位命中数最多的一行当表头。

    命中数并列时取更靠前的行——真实文档里越往后越可能是数据行本身凑巧撞上
    价格类关键词（比如备注写了"单价另议"）。
    """
    rows = _read_all_rows(file_path, sheet)
    if not rows:
        return [], []

    best_idx, best_hits = 0, -1
    for idx, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        hits = len(map_columns(row))
        if hits > best_hits:
            best_idx, best_hits = idx, hits
    return rows[best_idx], rows[best_idx + 1:]


def _column_is_all_zero(rows: list[list[str]], idx: int) -> bool:
    """这一列的非空取值是不是**无一例外都是零**？整列空返回 False。

    整列空由 `_looks_filled` 那条路自然算成 0% 填充，不需要在这里重复判定；
    这里只管"填了，但填的全是 0"这一种。
    """
    seen = 0
    for row in rows:
        if idx >= len(row):
            continue
        v = (row[idx] or "").strip()
        if not v:
            continue
        n = parse_num(v)
        if n is None or n != 0:
            return False
        seen += 1
    return seen > 0


def _looks_filled(cell: str) -> bool:
    """价格列的一格是不是"有值"——只判"这一格有没有人填过东西"，不判"填的
    数值是不是合理的价格"。

    **"0" 算填了，纯空白不算**——这是实测踩过的坑（design/28 §2 的金桥招标
    xlsx 复核）：这份表价格列里"单价"整列是空白，"合计"/"含税合计"整列是
    字面 "0"，是两种不同的信号（一列压根没人碰过，另一列有人显式填了"暂
    定零价"），会被"0 当未填"这种偷懒判据错误地合并成同一种"空"，结果整
    份表算出 0% 填充率，看起来像"确定性空白清单"，但其实是"部分列有值、
    部分列没有"的真实模糊状态——本该判"不确定"交给下一级，不该在这一层
    被这个简化判据误判成"definitive"。数值合法性（"面议"这类非数字文本、
    是否真的是合理价格）是识别阶段的事，不在 Tier 0 重复判断。
    """
    s = (cell or "").strip()
    if not s or s in ("-", "—", "/", "\\"):
        return False
    return True


def classify_excel(file_path: str, sheet: str | None = None) -> ExcelClassification:
    """design/28 §3 Tier 0 Excel 分支：

        无价格列               → 采购清单（definitive）—— 空白表单本身就是判据，
                                  投标方要填的清单不可能自带价格。
        价格列几乎全部填满      → 报价清单（strong）
        价格列部分填 / 界线不清 → 不确定，交给 Tier 2
    """
    header, rows = _load_header_and_rows(file_path, sheet)
    if not header:
        return ExcelClassification(
            file_path=file_path, verdict="uncertain", confidence="ambiguous",
            reason="工作表为空或无法识别表头，无法判定。",
        )

    cmap = map_columns(header)
    price_headers = sorted({cmap[slot] for slot in _PRICE_SLOTS if slot in cmap})
    idx_of = {h: i for i, h in enumerate(header)}
    price_idx = [idx_of[h] for h in price_headers if h in idx_of]

    # **整列全零的价格列不算价格列**（2026-08-23）。
    #
    # `_looks_filled` 逐格把 "0" 算成"填了"，那条判据是对的：一格写 0 和一格
    # 空白是两种不同的信号。但**整列**看就不成立——89 项全报 0 元不是报价，是
    # 占位。金桥采购清单实测：`单价(不含税)` 整列空，`合计(不含税)`/`税额`/
    # `价税合计` 三列的非空取值**无一例外全是 "0"**，一个真实价格都没有；旧判据
    # 却算出 63% 填充率 → uncertain → 弹窗 → 一路落进报价清单，于是**采购清单
    # 变成了比价矩阵里的一列供应商**（0/89、合计 ¥0）。那是类别错误，不是
    # 置信度问题。
    #
    # 这跟 design/39 里「锚点数量为 0 = 没标数量，不是数量是零」是同一条推理：
    # 单点的零可以是事实，整列的零是"这一列没有信息"。
    price_idx = [i for i in price_idx if not _column_is_all_zero(rows, i)]

    if not price_idx:
        return ExcelClassification(
            file_path=file_path, verdict="tender_list", confidence="definitive",
            price_columns=[], fill_rate=None, row_count=len(rows),
            reason="表头未识别出任何价格列——空白清单表，投标方待填，"
                   "这是采购清单的定义性特征（design/28 §2）。",
        )

    filled = total = 0
    for row in rows:
        for i in price_idx:
            total += 1
            if i < len(row) and _looks_filled(row[i]):
                filled += 1
    fill_rate = (filled / total) if total else 0.0

    if fill_rate >= FILL_RATE_STRONG:
        verdict, confidence = "bid_list", "strong"
        reason = (f"识别到价格列 {price_headers}，填充率 {fill_rate:.0%} ≥ "
                  f"{FILL_RATE_STRONG:.0%} 门槛——已填好价格的报价单。")
    elif fill_rate <= FILL_RATE_BLANK:
        verdict, confidence = "tender_list", "definitive"
        reason = (f"识别到价格列 {price_headers}，但填充率只有 {fill_rate:.0%}"
                  f"——有价格表头、格子是空的，正是投标方待填的空白清单表"
                  f"（design/28 §2）。")
    else:
        verdict, confidence = "uncertain", "ambiguous"
        reason = (f"识别到价格列 {price_headers}，填充率 {fill_rate:.0%}，"
                  f"介于清单和报价单之间，本层不猜，交给下一级判据。")

    return ExcelClassification(
        file_path=file_path, verdict=verdict, confidence=confidence,
        price_columns=price_headers, fill_rate=fill_rate, row_count=len(rows),
        reason=reason,
    )


def classify_tier0(
    file_path: str, sheet: str | None = None,
) -> ExcelClassification | PdfClassification | None:
    """按扩展名分派。

    xlsx/xls/csv → `ExcelClassification`（可能是 definitive/strong 的确定判
    定，也可能是 verdict="uncertain" 的"判不了"——两者都是正常返回值）。
    pdf      → `PdfClassification`（永远判不了招标/投标，只给文字层信号；
                调用方看到这个类型就该知道"还没分类完，进 Tier 1"）。
    其他扩展名 → None，调用方视为本层不支持，不是分类失败。

    **`.csv` 于 2026-08-23 补入。** 此前它落进"其他扩展名"返回 None，路由据此
    答 `kind="unsupported"`，界面显示「不支持的文件类型」——而报价 CSV 的解析
    链路一直是通的（`extract_quote_tabular` 实测四份各 136 项全过）。能力在，
    门口拦了。判据与载体无关（看表头有没有价格列、填充率多少），CSV 就是单表
    的 Excel，所以复用同一支判据、也仍然返回 `ExcelClassification`：前端按
    `kind === 'excel'` 分流，不用改。
    """
    suffix = Path(file_path).suffix.lower()
    if suffix in (".xlsx", ".xls", ".csv"):
        return classify_excel(file_path, sheet)
    if suffix == ".pdf":
        return classify_pdf(file_path)
    return None
