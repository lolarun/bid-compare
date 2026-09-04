"""Integration tests for the compare main flow.

Validates Phase 3 end-to-end pipeline:
1. Upload quote PDFs/images for two suppliers via /api/intake/upload
2. Poll until DONE
3. POST /api/quotes/batch-confirm to convert results → Quote rows
4. POST /api/analysis/bid-matrix returns rows + totals + recommended supplier

Uses MockProvider with canned per-supplier responses so no LLM is called.
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from PIL import Image
from sqlalchemy import func, select, update

from apps.api.intelligence.pipeline import ExtractionPipeline
from apps.api.intelligence.providers.mock import MockProvider

FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"


def _png() -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", (16, 16), "white").save(buf, format="PNG")
    return buf.getvalue()


# Two distinct canned quote responses (different prices to make matrix interesting)
SUPPLIER_A_QUOTE = {
    "supplier_name": "供应商A",
    "quote_date": "2026-05-20",
    "items": [
        {
            "material": "DN100 闸阀",
            "spec": "Z45X-16Q",
            "brand": "良工",
            "unit": "个",
            "qty": 10,
            "unit_price": 720,
            "total_price": 7200,        # 正常报价单原文就有合价列
        },
        {
            "material": "DN50 闸阀",
            "spec": "Z45X-16Q",
            "brand": "良工",
            "unit": "个",
            "qty": 20,
            "unit_price": 380,
            "total_price": 7600,
        },
    ],
}

SUPPLIER_B_QUOTE = {
    "supplier_name": "供应商B",
    "quote_date": "2026-05-20",
    "items": [
        {
            "material": "DN100 闸阀",
            "spec": "Z45X-16Q",
            "brand": "正丰",
            "unit": "个",
            "qty": 10,
            "unit_price": 690,
            "total_price": 6900,
        },
        {
            "material": "DN50 闸阀",
            "spec": "Z45X-16Q",
            "brand": "正丰",
            "unit": "个",
            "qty": 20,
            "unit_price": 400,
            "total_price": 8000,
        },
    ],
}


class _CycleProvider(MockProvider):
    """MockProvider that returns canned responses in round-robin order.

    Lets a single TestClient serve multiple distinct "supplier uploads".
    """

    def __init__(self, responses: list[dict]):
        super().__init__()
        self._responses = list(responses)
        self._idx = 0

    def extract(self, images, schema, prompt, timeout=90, **kwargs):
        canned = self._responses[self._idx % len(self._responses)]
        self._idx += 1
        from apps.api.intelligence.base import ExtractionResponse
        return ExtractionResponse(
            data=canned,
            raw_text=json.dumps(canned, ensure_ascii=False),
            confidence=1.0,
            tokens_used=0,
            provider="mock-cycle",
            duration_ms=1,
        )


@pytest.fixture
def compare_client(temp_db, monkeypatch, tmp_path, auth_override):
    monkeypatch.setenv("LLM_PROVIDER", "mock")
    monkeypatch.setattr(
        "apps.api.services.ingestion.document_ingestion.UPLOAD_DIR", tmp_path / "uploads"
    )

    cycle_provider = _CycleProvider([SUPPLIER_A_QUOTE, SUPPLIER_B_QUOTE])

    monkeypatch.setattr(
        "apps.api.main._build_pipeline",
        lambda: ExtractionPipeline(cycle_provider),
    )
    from apps.api.main import app

    with TestClient(app) as c:
        yield c


class TestPhase3CompareFlow:
    def test_full_pipeline_to_bid_matrix(self, compare_client):
        """End-to-end: upload×2 → batch-confirm×2 → bid-matrix."""
        # ── 1. Upload quote for supplier A ──
        r = compare_client.post(
            "/api/intake/upload",
            data={"type": "quote", "category": "阀门", "project_id": ""},
            files={"file": ("A.png", _png(), "image/png")},
        )
        assert r.status_code == 200, r.text
        job_a = r.json()["id"]

        # ── 2. Upload quote for supplier B (different image bytes → different hash) ──
        # Vary bytes by saving a slightly different image
        buf_b = io.BytesIO()
        Image.new("RGB", (16, 16), (250, 250, 250)).save(buf_b, format="PNG")
        r = compare_client.post(
            "/api/intake/upload",
            data={"type": "quote", "category": "阀门"},
            files={"file": ("B.png", buf_b.getvalue(), "image/png")},
        )
        assert r.status_code == 200
        job_b = r.json()["id"]

        # ── 3. Both jobs should be DONE ──
        for jid in (job_a, job_b):
            r = compare_client.get(f"/api/intake/jobs/{jid}")
            body = r.json()
            assert body["status"] == "done", body

        # ── 4. Pre-create suppliers (P0: supplier_id must be explicit) ──
        rs_a = compare_client.post(
            "/api/suppliers",
            json={"name": "供应商A", "categories": ["阀门"]},
        )
        assert rs_a.status_code == 201, rs_a.text
        supplier_a_id = rs_a.json()["id"]

        rs_b = compare_client.post(
            "/api/suppliers",
            json={"name": "供应商B", "categories": ["阀门"]},
        )
        assert rs_b.status_code == 201, rs_b.text
        supplier_b_id = rs_b.json()["id"]

        # ── 5. batch-confirm for both suppliers — new P0 contract ──
        r = compare_client.post(
            "/api/quotes/batch-confirm",
            json={
                "job_id": job_a,
                "supplier_id": supplier_a_id,
                "supplier_name": "供应商A",
                "project_name": "Phase3 测试比价项目",
                "category": "阀门",
            },
        )
        assert r.status_code == 200, r.text
        a = r.json()
        assert a["line_count"] == 2, a
        assert a["supplier_id"] == supplier_a_id
        assert a["project_id"]
        sub_a_id = a["submission_id"]

        r = compare_client.post(
            "/api/quotes/batch-confirm",
            json={
                "job_id": job_b,
                "supplier_id": supplier_b_id,
                "supplier_name": "供应商B",
                "project_id": a["project_id"],  # link to same project
                "category": "阀门",
            },
        )
        assert r.status_code == 200, r.text
        b = r.json()
        assert b["line_count"] == 2, b
        sub_b_id = b["submission_id"]

        project_id = a["project_id"]

        # ── 5.5. Create TenderListSession (required before match/bid-matrix with category) ──
        anchor_items = [
            {"seq": "1", "name": "DN100 闸阀", "spec": "Z45X-16Q", "unit": "个", "qty": 10, "category": "阀门"},
            {"seq": "2", "name": "DN50 闸阀", "spec": "Z45X-16Q", "unit": "个", "qty": 20, "category": "阀门"},
        ]
        r = compare_client.post(
            "/api/analysis/tender-list/confirm",
            json={
                "project_id": project_id,
                "category": "阀门",
                "file_name": "test.xlsx",
                "anchors_json": anchor_items,
                "anchors_total": len(anchor_items),
                "source_type": "excel",
            },
        )
        assert r.status_code == 200, r.text

        # ── 5.6. tender-list/match — persists used_submission_ids on the session ──
        r = compare_client.post(
            "/api/analysis/tender-list/match",
            data={
                "project_id": str(project_id),
                "category": "阀门",
                "supplier_ids": f"{supplier_a_id},{supplier_b_id}",
                "submission_ids": f"{sub_a_id},{sub_b_id}",
            },
        )
        assert r.status_code == 200, r.text

        # Verify used_submission_ids is now persisted on the session
        r = compare_client.get(
            "/api/analysis/tender-list/current",
            params={"project_id": project_id, "category": "阀门"},
        )
        assert r.status_code == 200, r.text
        session_data = r.json()
        used_sids = set(session_data.get("used_submission_ids") or [])
        assert used_sids == {sub_a_id, sub_b_id}, f"used_submission_ids={used_sids}"

        # ── 6. bid-matrix returns non-empty rows + totals + recommended ──
        r = compare_client.post(
            "/api/analysis/bid-matrix",
            json={
                "project_id": project_id,
                "supplier_ids": [supplier_a_id, supplier_b_id],
                "category": "阀门",
            },
        )
        assert r.status_code == 200, r.text
        matrix = r.json()
        assert matrix["project_id"] == project_id
        assert len(matrix["suppliers"]) == 2
        assert len(matrix["rows"]) >= 1
        assert len(matrix["totals"]) == 2

        # Each row should have a recommended supplier letter (A or B)
        for row in matrix["rows"]:
            assert "suppliers" in row
            assert len(row["suppliers"]) == 2  # one cell per supplier

        # Totals carry total + avg_deviation per supplier
        for t in matrix["totals"]:
            assert "id" in t
            assert "total" in t
            assert "avg_deviation" in t

        # Schema: recommendation_blocked field must be present
        assert "recommendation_blocked" in matrix

    def test_batch_confirm_rejects_non_quote_job(self, compare_client):
        # Upload a tender (wrong type) and try to confirm as quote
        r = compare_client.post(
            "/api/intake/upload",
            data={"type": "tender"},
            files={"file": ("t.png", _png(), "image/png")},
        )
        tender_job = r.json()["id"]

        # supplier_id=1 is fake but job-type check fires before supplier lookup
        r2 = compare_client.post(
            "/api/quotes/batch-confirm",
            json={"job_id": tender_job, "supplier_id": 1, "category": "阀门", "supplier_name": "X"},
        )
        assert r2.status_code == 400
        assert "must be 'quote'" in r2.json()["detail"]

    def test_batch_confirm_unknown_brands_reported(self, compare_client):
        # Create a supplier so supplier_id can be provided (P0 requirement)
        rs = compare_client.post(
            "/api/suppliers",
            json={"name": "TestSupplierBrands", "categories": ["阀门"]},
        )
        assert rs.status_code == 201, rs.text
        supplier_id = rs.json()["id"]

        r = compare_client.post(
            "/api/intake/upload",
            data={"type": "quote", "category": "阀门"},
            files={"file": ("A.png", _png(), "image/png")},
        )
        job_id = r.json()["id"]
        r2 = compare_client.post(
            "/api/quotes/batch-confirm",
            json={
                "job_id": job_id,
                "supplier_id": supplier_id,
                "supplier_name": "TestSupplierBrands",
                "category": "阀门",
            },
        )
        assert r2.status_code == 200, r2.text
        body = r2.json()
        # Canned response has brands 良工/正丰; neither is in seeded brand_tiers
        assert len(body["unknown_brands"]) > 0


# ── Shared setup helper ──────────────────────────────────────────────────────

def _make_full_project(client) -> dict:
    """Upload×2 → batch-confirm×2 → tender-confirm → tender-match.

    Returns state dict used by gate and export consistency tests.
    """
    # Upload two supplier quote images
    r = client.post(
        "/api/intake/upload",
        data={"type": "quote", "category": "阀门", "project_id": ""},
        files={"file": ("A.png", _png(), "image/png")},
    )
    assert r.status_code == 200
    job_a = r.json()["id"]

    buf_b = io.BytesIO()
    Image.new("RGB", (16, 16), (250, 250, 250)).save(buf_b, format="PNG")
    r = client.post(
        "/api/intake/upload",
        data={"type": "quote", "category": "阀门"},
        files={"file": ("B.png", buf_b.getvalue(), "image/png")},
    )
    assert r.status_code == 200
    job_b = r.json()["id"]

    rs_a = client.post("/api/suppliers", json={"name": "供应商A", "categories": ["阀门"]})
    assert rs_a.status_code == 201
    supplier_a_id = rs_a.json()["id"]
    rs_b = client.post("/api/suppliers", json={"name": "供应商B", "categories": ["阀门"]})
    assert rs_b.status_code == 201
    supplier_b_id = rs_b.json()["id"]

    r = client.post("/api/quotes/batch-confirm", json={
        "job_id": job_a, "supplier_id": supplier_a_id,
        "supplier_name": "供应商A", "project_name": "Gate测试项目", "category": "阀门",
    })
    assert r.status_code == 200
    a = r.json()
    project_id = a["project_id"]
    sub_a_id = a["submission_id"]

    r = client.post("/api/quotes/batch-confirm", json={
        "job_id": job_b, "supplier_id": supplier_b_id,
        "supplier_name": "供应商B", "project_id": project_id, "category": "阀门",
    })
    assert r.status_code == 200
    sub_b_id = r.json()["submission_id"]

    anchor_items = [
        {"seq": "1", "name": "DN100 闸阀", "spec": "Z45X-16Q", "unit": "个", "qty": 10, "category": "阀门"},
        {"seq": "2", "name": "DN50 闸阀",  "spec": "Z45X-16Q", "unit": "个", "qty": 20, "category": "阀门"},
    ]
    r = client.post("/api/analysis/tender-list/confirm", json={
        "project_id": project_id, "category": "阀门",
        "file_name": "test.xlsx", "anchors_json": anchor_items,
        "anchors_total": len(anchor_items), "source_type": "excel",
    })
    assert r.status_code == 200

    r = client.post("/api/analysis/tender-list/match", data={
        "project_id": str(project_id), "category": "阀门",
        "supplier_ids": f"{supplier_a_id},{supplier_b_id}",
        "submission_ids": f"{sub_a_id},{sub_b_id}",
    })
    assert r.status_code == 200, r.text

    return {
        "project_id": project_id,
        "supplier_a_id": supplier_a_id,
        "supplier_b_id": supplier_b_id,
        "sub_a_id": sub_a_id,
        "sub_b_id": sub_b_id,
    }


# ── Gate tests ───────────────────────────────────────────────────────────────

class TestBidMatrixGates:
    def test_alignment_never_run_returns_recoverable_error(self, compare_client):
        """项目 137 实测复现（2026-08-24）：两家报价真的入库了、招标清单也确认了，
        唯独 `tender-list/match` 从没跑过——`used_submission_ids` 因此是空的。

        这不是构造出来的边角情况：截至这条测试写下的时刻，**全前端没有任何界面
        会调用 `/analysis/tender-list/match`**（`AnchorReviewMatrix.vue` 只读现成
        对齐结果）。正常的 上传→入库→比价 流程走到底，任何项目都会撞上这道门。

        断言两件事：① error 码是 `alignment_not_run`，不是裸字符串（前端要靠它
        识别"能自动补救"）；② 补跑一次 match 之后重试 `/bid-matrix` 真的能成功
        ——这正是前端 `runAnalysis` 现在做的恢复动作，这里在后端把同一条路径
        走一遍，不只测错误码本身。
        """
        # 复刻 _make_full_project 到 tender-list/confirm 那一步，但**跳过 match**。
        r = compare_client.post(
            "/api/intake/upload",
            data={"type": "quote", "category": "阀门", "project_id": ""},
            files={"file": ("A.png", _png(), "image/png")},
        )
        job_a = r.json()["id"]
        buf_b = io.BytesIO()
        Image.new("RGB", (16, 16), (250, 250, 250)).save(buf_b, format="PNG")
        r = compare_client.post(
            "/api/intake/upload",
            data={"type": "quote", "category": "阀门"},
            files={"file": ("B.png", buf_b.getvalue(), "image/png")},
        )
        job_b = r.json()["id"]

        rs_a = compare_client.post("/api/suppliers", json={"name": "供应商A137", "categories": ["阀门"]})
        supplier_a_id = rs_a.json()["id"]
        rs_b = compare_client.post("/api/suppliers", json={"name": "供应商B137", "categories": ["阀门"]})
        supplier_b_id = rs_b.json()["id"]

        r = compare_client.post("/api/quotes/batch-confirm", json={
            "job_id": job_a, "supplier_id": supplier_a_id,
            "supplier_name": "供应商A137", "project_name": "137复现项目", "category": "阀门",
        })
        assert r.status_code == 200, r.text
        project_id = r.json()["project_id"]
        sub_a_id = r.json()["submission_id"]

        r = compare_client.post("/api/quotes/batch-confirm", json={
            "job_id": job_b, "supplier_id": supplier_b_id,
            "supplier_name": "供应商B137", "project_id": project_id, "category": "阀门",
        })
        assert r.status_code == 200, r.text
        sub_b_id = r.json()["submission_id"]

        anchor_items = [
            {"seq": "1", "name": "DN100 闸阀", "spec": "Z45X-16Q", "unit": "个", "qty": 10, "category": "阀门"},
            {"seq": "2", "name": "DN50 闸阀",  "spec": "Z45X-16Q", "unit": "个", "qty": 20, "category": "阀门"},
        ]
        r = compare_client.post("/api/analysis/tender-list/confirm", json={
            "project_id": project_id, "category": "阀门",
            "file_name": "test.xlsx", "anchors_json": anchor_items,
            "anchors_total": len(anchor_items), "source_type": "excel",
        })
        assert r.status_code == 200, r.text
        # 到这里为止：两家都真的入库了（有 BidQuoteLine），session 也确认了——
        # 唯独没调用过 tender-list/match。这正是项目 137 的状态。

        r = compare_client.post("/api/analysis/bid-matrix", json={
            "project_id": project_id, "category": "阀门",
            "supplier_ids": [supplier_a_id, supplier_b_id],
            "submission_ids": [sub_a_id, sub_b_id],
        })
        assert r.status_code == 409, r.text
        detail = r.json()["detail"]
        assert detail["error"] == "alignment_not_run", detail

        # 前端恢复动作：补跑一次 match，再重试 bid-matrix。
        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(project_id), "category": "阀门",
            "submission_ids": f"{sub_a_id},{sub_b_id}",
        })
        assert r.status_code == 200, r.text

        r = compare_client.post("/api/analysis/bid-matrix", json={
            "project_id": project_id, "category": "阀门",
            "supplier_ids": [supplier_a_id, supplier_b_id],
            "submission_ids": [sub_a_id, sub_b_id],
        })
        assert r.status_code == 200, r.text
        assert len(r.json()["rows"]) > 0

    def test_submission_ids_mismatch_returns_409(self, compare_client):
        """body.submission_ids ≠ session.used_submission_ids → 409."""
        state = _make_full_project(compare_client)
        r = compare_client.post(
            "/api/analysis/bid-matrix",
            json={
                "project_id": state["project_id"],
                "supplier_ids": [state["supplier_a_id"], state["supplier_b_id"]],
                "category": "阀门",
                "submission_ids": [state["sub_a_id"]],  # intentionally incomplete → mismatch
            },
        )
        assert r.status_code == 409, r.text
        # 2026-08-24 改：detail 从裸字符串改成 {error, message} 结构化——前端
        # 要按 error 码识别"能不能自动重新对齐再重试"，不能靠翻译过的中文文案
        # 做字符串匹配（那是给人看的话，会变；error 码不会）。
        detail = r.json()["detail"]
        assert detail["error"] == "alignment_stale", detail

    def test_stale_finalization_does_not_empty_matrix(self, compare_client, db_session):
        """B2 回归：过期 finalize 快照（group_ids 指向旧轮次的组）不得把矩阵清空。

        复现生产项目 59：重跑 match 产生新 confirmed 组后，旧 finalize 快照仍锁定
        被 superseded 的旧组 → 按它过滤 allowed_group_ids 会排除全部当前组 → 全「未报价」。
        修复后：检测到快照与当前 confirmed 组零交集 → 回退当前组 + 告警。
        """
        from apps.api.models.alignment_finalization import AlignmentFinalization

        state = _make_full_project(compare_client)
        pid = state["project_id"]
        body = {
            "project_id": pid,
            "supplier_ids": [state["supplier_a_id"], state["supplier_b_id"]],
            "category": "阀门",
        }

        def _priced_cells(matrix):
            return sum(
                1 for row in matrix["rows"] for c in row["suppliers"]
                if c.get("price") is not None
            )

        # 基线（无 finalize 快照）：矩阵有报价格
        base = compare_client.post("/api/analysis/bid-matrix", json=body).json()
        base_priced = _priced_cells(base)
        assert base_priced > 0, "基线矩阵应有报价格（mock 报价已对齐）"

        # 注入过期 finalize 快照：group_ids 指向不存在的旧组
        db_session.add(AlignmentFinalization(
            project_id=pid, category="阀门", status="finalized",
            group_ids_json=[999999, 999998],
        ))
        db_session.commit()

        # B2：矩阵不得被过期快照清空；报价格数与基线一致；并给出过期告警
        r = compare_client.post("/api/analysis/bid-matrix", json=body)
        assert r.status_code == 200, r.text
        m = r.json()
        assert _priced_cells(m) == base_priced, "过期 finalize 快照不应导致矩阵全空"
        assert m.get("not_finalized_warning"), "应提示快照已过期并回退"


# ── Export consistency tests ─────────────────────────────────────────────────

class TestBidMatrixExportConsistency:
    def test_export_matches_bid_matrix(self, compare_client):
        """Export Excel and /bid-matrix JSON must have identical supplier columns and row count."""
        import openpyxl

        state = _make_full_project(compare_client)
        project_id = state["project_id"]
        supplier_a_id = state["supplier_a_id"]
        supplier_b_id = state["supplier_b_id"]

        # Get bid-matrix JSON
        r = compare_client.post(
            "/api/analysis/bid-matrix",
            json={
                "project_id": project_id,
                "supplier_ids": [supplier_a_id, supplier_b_id],
                "category": "阀门",
            },
        )
        assert r.status_code == 200, r.text
        matrix = r.json()

        # No legacy source_quote_id in any cell
        for row in matrix["rows"]:
            for cell in row["suppliers"]:
                assert cell.get("source_quote_id") is None, (
                    f"legacy source_quote_id={cell.get('source_quote_id')} in cell for row {row.get('material_name')}"
                )

        # Download export Excel
        r_export = compare_client.get(
            "/api/export/bid-matrix",
            params={
                "supplier_ids": f"{supplier_a_id},{supplier_b_id}",
                "project_id": project_id,
                "category": "阀门",
            },
        )
        assert r_export.status_code == 200, r_export.text

        wb = openpyxl.load_workbook(io.BytesIO(r_export.content))
        assert "比价矩阵" in wb.sheetnames
        ws = wb["比价矩阵"]

        # Row count: header(1) + data rows + totals(1)
        excel_data_rows = ws.max_row - 2
        assert excel_data_rows == len(matrix["rows"]), (
            f"Excel data rows={excel_data_rows} != matrix rows={len(matrix['rows'])}"
        )

        # Supplier names appear in the header row
        header_vals = [cell.value or "" for cell in ws[1]]
        for sup in matrix["suppliers"]:
            assert any(sup["name"] in h for h in header_vals), (
                f"Supplier {sup['name']!r} not found in Excel header: {header_vals}"
            )

        # matrix_distribution present in both or neither
        if matrix.get("matrix_distribution"):
            assert "供应商覆盖摘要" in wb.sheetnames
            assert matrix["matrix_distribution"]["supplier_count"] == len(matrix["suppliers"])


# ── Category persistence tests ───────────────────────────────────────────────

def _make_project_with_confirm(client) -> dict:
    """创建项目 + 确认招标清单（单品类），不执行 match。

    Returns dict with project_id, category.
    """
    r = client.post("/api/projects", json={"name": "品类恢复测试", "remark": ""})
    assert r.status_code == 201
    project_id = r.json()["id"]

    anchor_items = [
        {"seq": "1", "name": "DN100 闸阀", "spec": "Z45X-16Q", "unit": "个", "qty": 10, "category": "阀门"},
    ]
    r = client.post("/api/analysis/tender-list/confirm", json={
        "project_id": project_id,
        "category": "阀门",
        "file_name": "test.xlsx",
        "anchors_json": anchor_items,
        "anchors_total": len(anchor_items),
        "source_type": "excel",
    })
    assert r.status_code == 200, r.text
    return {"project_id": project_id, "category": "阀门", "confirm_resp": r.json()}


class TestCategoryPersistence:
    """品类持久化恢复三层防线测试。"""

    def test_confirm_returns_primary_category(self, compare_client):
        """tender-list/confirm 响应必须包含 primary_category 且与单品类 session 一致。"""
        state = _make_project_with_confirm(compare_client)
        resp = state["confirm_resp"]
        assert "primary_category" in resp, "confirm 响应缺少 primary_category 字段"
        assert resp["primary_category"] == "阀门", (
            f"primary_category={resp['primary_category']!r} 应为 '阀门'"
        )
        assert resp["primary_category"] != "", "primary_category 不得为空字符串"

    def test_current_sessions_returns_sessions_and_primary(self, compare_client):
        """confirm 后调 /current-sessions 能还原 sessions 和 primary_category。"""
        state = _make_project_with_confirm(compare_client)
        project_id = state["project_id"]

        r = compare_client.get(
            "/api/analysis/tender-list/current-sessions",
            params={"project_id": project_id},
        )
        assert r.status_code == 200, r.text
        data = r.json()

        assert "sessions" in data
        assert "primary_category" in data
        assert len(data["sessions"]) == 1
        assert data["sessions"][0]["category"] == "阀门"
        assert data["primary_category"] == "阀门"
        assert data["primary_category"] != ""
        # 接口字段完整（前端恢复所需）
        s = data["sessions"][0]
        assert "id" in s and "category" in s and "anchors_total" in s

    def test_current_sessions_404_for_new_project(self, compare_client):
        """无 current session 的新项目返回 404，不得 500。"""
        r = compare_client.post("/api/projects", json={"name": "空项目", "remark": ""})
        assert r.status_code == 201
        new_pid = r.json()["id"]

        r = compare_client.get(
            "/api/analysis/tender-list/current-sessions",
            params={"project_id": new_pid},
        )
        assert r.status_code == 404, f"新项目应返回 404，实际 {r.status_code}: {r.text}"

    def test_batch_confirm_empty_category_rejected(self, compare_client):
        """batch-confirm 提交 category='' 后端应返回 4xx，防止空品类数据入库。"""
        r = compare_client.post(
            "/api/intake/upload",
            data={"type": "quote"},
            files={"file": ("A.png", _png(), "image/png")},
        )
        assert r.status_code == 200
        job_id = r.json()["id"]

        r = compare_client.post("/api/quotes/batch-confirm", json={
            "job_id": job_id,
            "supplier_name": "测试供应商",
            "category": "",
        })
        assert r.status_code in (400, 422), (
            f"空 category 应被拒绝(400/422)，实际 {r.status_code}: {r.text}"
        )


# ── Submission resolver tests ─────────────────────────────────────────────────

def _setup_two_subs_same_supplier(client) -> dict:
    """为同一 supplier 创建两份 submission（sub_a 和 sub_b），项目+清单均就绪。

    模拟 repair 注入旧 submission + 用户新上传场景。
    返回 {project_id, supplier_id, sub_a_id, sub_b_id, category}
    """
    # 项目 + 供应商
    rp = client.post("/api/projects", json={"name": "Resolver测试", "remark": ""})
    assert rp.status_code == 201
    project_id = rp.json()["id"]

    rs = client.post("/api/suppliers", json={"name": "凯硕新正", "categories": ["阀门"]})
    assert rs.status_code == 201
    supplier_id = rs.json()["id"]

    def _upload_and_confirm(suffix: str) -> int:
        r = client.post(
            "/api/intake/upload",
            data={"type": "quote"},
            files={"file": (f"{suffix}.png", _png(), "image/png")},
        )
        assert r.status_code == 200
        job_id = r.json()["id"]
        r2 = client.post("/api/quotes/batch-confirm", json={
            "job_id": job_id,
            "supplier_id": supplier_id,
            "supplier_name": "凯硕新正",
            "project_id": project_id,
            "category": "阀门",
        })
        assert r2.status_code == 200, r2.text
        return r2.json()["submission_id"]

    sub_a_id = _upload_and_confirm("quote_a")
    sub_b_id = _upload_and_confirm("quote_b")

    # 招标清单
    anchors = [{"seq": "1", "name": "DN100 闸阀", "spec": "Z45X", "unit": "个", "qty": 5, "category": "阀门"}]
    r = client.post("/api/analysis/tender-list/confirm", json={
        "project_id": project_id, "category": "阀门",
        "file_name": "t.xlsx", "anchors_json": anchors,
        "anchors_total": len(anchors), "source_type": "excel",
    })
    assert r.status_code == 200, r.text

    return {
        "project_id": project_id,
        "supplier_id": supplier_id,
        "sub_a_id": sub_a_id,
        "sub_b_id": sub_b_id,
        "category": "阀门",
    }


class TestSubmissionResolver:
    """resolve_active_submissions 语义正确性测试。"""

    def test_submission_ids_excludes_supplier_union(self, compare_client):
        """supplier_ids=[sid] 且 submission_ids=[sub_b] 时，不得返回 sub_a（同供应商）。"""
        from apps.api.core.database import SessionLocal
        from apps.api.services.submission.bid_submission_resolve import (
            resolve_active_submissions,
        )

        state = _setup_two_subs_same_supplier(compare_client)

        db = SessionLocal()
        try:
            result = resolve_active_submissions(
                db,
                project_id=state["project_id"],
                category=state["category"],
                supplier_ids=[state["supplier_id"]],
                submission_ids=[state["sub_b_id"]],  # 明确只要 sub_b
            )
        finally:
            db.close()

        assert state["sub_b_id"] in result, "sub_b 必须在结果中"
        assert state["sub_a_id"] not in result, (
            f"sub_a={state['sub_a_id']} 不应出现：submission_ids 已指定唯一权威集合"
        )

    def test_superseded_submission_excluded(self, compare_client):
        """superseded 状态的 submission 永不参与 resolve，无论 supplier_ids 是否匹配。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidSubmission
        from apps.api.services.submission.bid_submission_resolve import (
            resolve_active_submissions,
        )

        state = _setup_two_subs_same_supplier(compare_client)

        # 将 sub_a 标记为 superseded
        db = SessionLocal()
        try:
            sub_a = db.get(BidSubmission, state["sub_a_id"])
            sub_a.status = "superseded"
            db.commit()

            result = resolve_active_submissions(
                db,
                project_id=state["project_id"],
                category=state["category"],
                supplier_ids=[state["supplier_id"]],
            )
        finally:
            db.close()

        assert state["sub_a_id"] not in result, "superseded submission 不得参与 resolve"
        assert state["sub_b_id"] in result, "非 superseded 的 sub_b 必须保留"

    def test_explicit_sub_b_only_consumes_sub_b(self, compare_client):
        """同一 supplier 两份 submission，显式选择 sub_b 时结果集只含 sub_b。"""
        from apps.api.core.database import SessionLocal
        from apps.api.services.submission.bid_submission_resolve import (
            resolve_active_submissions,
        )

        state = _setup_two_subs_same_supplier(compare_client)

        db = SessionLocal()
        try:
            result = resolve_active_submissions(
                db,
                project_id=state["project_id"],
                category=state["category"],
                submission_ids=[state["sub_b_id"]],
            )
        finally:
            db.close()

        assert list(result.keys()) == [state["sub_b_id"]], (
            f"期望结果只含 sub_b={state['sub_b_id']}，实际={list(result.keys())}"
        )

    def test_no_submission_ids_legacy_path(self, compare_client):
        """不传 submission_ids 时，supplier_ids 仍能正常查到全部 active submissions。"""
        from apps.api.core.database import SessionLocal
        from apps.api.services.submission.bid_submission_resolve import (
            resolve_active_submissions,
        )

        state = _setup_two_subs_same_supplier(compare_client)

        db = SessionLocal()
        try:
            result = resolve_active_submissions(
                db,
                project_id=state["project_id"],
                category=state["category"],
                supplier_ids=[state["supplier_id"]],
                # submission_ids 不传 → 旧路径，返回该 supplier 全部 active
            )
        finally:
            db.close()

        assert state["sub_a_id"] in result
        assert state["sub_b_id"] in result

    def test_match_used_submission_ids_exact(self, compare_client):
        """match 后 used_submission_ids 必须精确等于传入的 submission_ids，不多不少。"""
        state = _setup_two_subs_same_supplier(compare_client)
        project_id = state["project_id"]
        sub_b_id = state["sub_b_id"]

        # 只用 sub_b 做 match
        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(project_id),
            "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(sub_b_id),
        })
        assert r.status_code == 200, r.text

        # 拉 used_submission_ids
        r2 = compare_client.get(
            "/api/analysis/tender-list/current",
            params={"project_id": project_id, "category": "阀门"},
        )
        assert r2.status_code == 200
        used = set(r2.json().get("used_submission_ids") or [])
        assert used == {sub_b_id}, (
            f"used_submission_ids={sorted(used)} 应精确等于 {{{sub_b_id}}}"
        )


# ── Quality gate tests ────────────────────────────────────────────────────────

class TestQualityGate:
    """price coverage 质量门：eligible 行中 unit_price>0 比率 < 80% → 422。"""

    def _make_sub_project(self, client, supplier_name: str, project_name: str) -> dict:
        """上传 + batch-confirm + 确认招标清单。返回 {project_id, supplier_id, sub_id}。"""
        r = client.post("/api/intake/upload",
            data={"type": "quote", "category": "阀门"},
            files={"file": ("q.png", _png(), "image/png")})
        assert r.status_code == 200
        job_id = r.json()["id"]

        rs = client.post("/api/suppliers", json={"name": supplier_name, "categories": ["阀门"]})
        assert rs.status_code == 201
        supplier_id = rs.json()["id"]

        r = client.post("/api/quotes/batch-confirm", json={
            "job_id": job_id, "supplier_id": supplier_id,
            "supplier_name": supplier_name, "project_name": project_name, "category": "阀门",
        })
        assert r.status_code == 200, r.text
        a = r.json()
        project_id = a["project_id"]
        sub_id = a["submission_id"]

        anchors = [{"seq": "1", "name": "DN100 闸阀", "spec": "Z45X", "unit": "个", "qty": 5, "category": "阀门"}]
        r = client.post("/api/analysis/tender-list/confirm", json={
            "project_id": project_id, "category": "阀门", "file_name": "t.xlsx",
            "anchors_json": anchors, "anchors_total": 1, "source_type": "excel",
        })
        assert r.status_code == 200, r.text

        return {"project_id": project_id, "supplier_id": supplier_id, "sub_id": sub_id}

    def test_low_price_coverage_blocks_match(self, compare_client):
        """全部 BQL 清零 unit_price → 覆盖率 0% < 80%，match 返回 422。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine

        state = self._make_sub_project(compare_client, "低覆盖供应商", "质量门测试A")
        sub_id = state["sub_id"]

        db = SessionLocal()
        try:
            db.execute(update(BidQuoteLine).where(
                BidQuoteLine.submission_id == sub_id
            ).values(unit_price=None))
            db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(sub_id),
        })
        assert r.status_code == 422, r.text
        detail = r.json()["detail"]
        assert isinstance(detail, dict), f"detail 应为 dict，实际：{detail!r}"
        assert detail.get("error") == "submission_quality_gate_failed"
        assert "failures" in detail
        assert len(detail["failures"]) == 1
        assert detail["failures"][0]["submission_id"] == sub_id

    def test_summary_rows_excluded_from_coverage(self, compare_client):
        """合计行 raw_name='合计' 排除在 eligible 外；其余行有价格 → 通过质量门。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine

        state = self._make_sub_project(compare_client, "含合计行供应商", "合计行排除测试")
        sub_id = state["sub_id"]

        # 注入一行合计摘要行（无价格）
        db = SessionLocal()
        try:
            db.add(BidQuoteLine(
                submission_id=sub_id,
                raw_name="合计",
                category="阀门",
                unit_price=None,
                qty=None,
                total_price=None,
            ))
            db.commit()
        finally:
            db.close()

        # 原有 2 条明细行全有价格 + 1 合计行无价格；合计行被排除，coverage=100% → 通过
        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(sub_id),
        })
        assert r.status_code == 200, f"含合计行时不应被质量门拦截: {r.text}"

    def test_quality_gate_failure_carries_structured_metrics(self, compare_client):
        """422 详情结构含 submission_id / eligible / price_ok / coverage / threshold。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine

        state = self._make_sub_project(compare_client, "指标验证供应商", "指标验证项目")
        sub_id = state["sub_id"]

        # 2 eligible 行中清零第一行 → coverage = 1/2 = 50% < 80%
        db = SessionLocal()
        try:
            first = db.scalar(select(BidQuoteLine).where(
                BidQuoteLine.submission_id == sub_id
            ))
            first.unit_price = None
            db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(sub_id),
        })
        assert r.status_code == 422
        failures = r.json()["detail"]["failures"]
        f = failures[0]
        assert f["submission_id"] == sub_id
        # 新结构: issues 列表，找到 price_coverage 项
        pc = next((i for i in f["issues"] if i["check"] == "price_coverage"), None)
        assert pc is not None, f"应有 price_coverage issue，实际 issues={f['issues']}"
        assert pc["eligible"] >= 1
        assert pc["price_ok"] < pc["eligible"]
        assert pc["coverage"] < 0.8
        assert pc["threshold"] == 0.8

    def test_superseded_submission_blocked_before_quality_gate(self, compare_client):
        """submission_ids 含 superseded 时，状态校验先于质量门触发（均返回 409）。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidSubmission

        state = _setup_two_subs_same_supplier(compare_client)
        sub_a_id = state["sub_a_id"]

        db = SessionLocal()
        try:
            sub_a = db.get(BidSubmission, sub_a_id)
            sub_a.status = "superseded"
            db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]),
            "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(sub_a_id),
        })
        # 409（不是 422）：superseded 触发的是归属/状态校验（D3 的 resolve_active_
        # submissions gate），不是本文件测的 _quality_failures 数据质量门——
        # 两者恰好都在 match 路由里，前者先判、409 未受评审 E1 的 422 统一影响。
        assert r.status_code == 409, r.text

    def test_full_coverage_passes_quality_gate(self, compare_client):
        """100% price coverage（MockProvider 默认有价格）顺利通过质量门。"""
        state = _setup_two_subs_same_supplier(compare_client)

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(state["sub_b_id"]),
        })
        assert r.status_code == 200, f"100% coverage 应通过质量门: {r.text}"


# ── Extended quality gate tests ───────────────────────────────────────────────

class TestExtendedQualityGate:
    """扩展质量门：算术错误率、系统性VAT混用、单行集中度、声明总价核对。"""

    def _base_setup(self, client, supplier_name: str, project_name: str) -> dict:
        r = client.post("/api/intake/upload",
            data={"type": "quote", "category": "阀门"},
            files={"file": ("q.png", _png(), "image/png")})
        assert r.status_code == 200
        job_id = r.json()["id"]

        rs = client.post("/api/suppliers", json={"name": supplier_name, "categories": ["阀门"]})
        supplier_id = rs.json()["id"]

        r = client.post("/api/quotes/batch-confirm", json={
            "job_id": job_id, "supplier_id": supplier_id,
            "supplier_name": supplier_name, "project_name": project_name, "category": "阀门",
        })
        assert r.status_code == 200
        a = r.json()
        project_id, sub_id = a["project_id"], a["submission_id"]

        anchors = [{"seq": "1", "name": "DN100 闸阀", "spec": "Z45X", "unit": "个", "qty": 5, "category": "阀门"}]
        client.post("/api/analysis/tender-list/confirm", json={
            "project_id": project_id, "category": "阀门", "file_name": "t.xlsx",
            "anchors_json": anchors, "anchors_total": 1, "source_type": "excel",
        })
        return {"project_id": project_id, "supplier_id": supplier_id, "sub_id": sub_id}

    def test_100pct_coverage_but_arithmetic_error_blocked(self, compare_client):
        """price_coverage=100% 但存在硬算术错误（偏差>12.5%）仍被拦截。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine

        state = self._base_setup(compare_client, "算术错误供应商", "算术错误测试")
        sub_id = state["sub_id"]

        # 将所有行的 total_price 乘以 2.6（偏差 = 160%，远超 VAT tolerance）。
        # 不用整数倍（×2/×3/×4等）——那些是 check_row_arithmetic 认得的"报价口径
        # 倍率"（按束/按根报价），会被归类为 multiplier 而非 mismatch，不参与
        # arithmetic_error_rate（评审 C2 修复：共享实现比这条测试原先假设的更精确，
        # 用非整数倍保持"真实算术错误"这个测试意图不变）。
        db = SessionLocal()
        try:
            rows = db.scalars(select(BidQuoteLine).where(BidQuoteLine.submission_id == sub_id)).all()
            for row in rows:
                if row.total_price:
                    row.total_price = row.total_price * 2.6
            db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(state["sub_id"]),
        })
        assert r.status_code == 422, r.text
        detail = r.json()["detail"]
        checks = [i["check"] for i in detail["failures"][0]["issues"]]
        assert "arithmetic_error_rate" in checks or "line_concentration" in checks, (
            f"应有算术错误或集中度检查，实际 checks={checks}"
        )

    def test_systematic_vat_mismatch_blocked(self, compare_client):
        """超过20%的行 total_price = qty×unit_price×1.13（系统性VAT混用）→ 422。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine

        state = self._base_setup(compare_client, "VAT混用供应商", "VAT混用测试")
        sub_id = state["sub_id"]

        # 将所有行的 total_price 改为 qty*unit_price*1.13（模拟不含税单价+含税合价列混用）
        db = SessionLocal()
        try:
            rows = db.scalars(select(BidQuoteLine).where(BidQuoteLine.submission_id == sub_id)).all()
            for row in rows:
                if row.qty and row.unit_price:
                    row.total_price = round(row.qty * row.unit_price * 1.13, 2)
            db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(state["sub_id"]),
        })
        assert r.status_code == 422, r.text
        detail = r.json()["detail"]
        checks = [i["check"] for i in detail["failures"][0]["issues"]]
        assert "systematic_vat_mismatch" in checks, (
            f"应触发 systematic_vat_mismatch，实际 checks={checks}"
        )

    def test_single_line_exceeds_60pct_of_total_blocked(self, compare_client):
        """单行金额占总金额 >60% → 疑似章节小计被误识别为单行产品 → 422。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine

        state = self._base_setup(compare_client, "高集中度供应商", "集中度测试")
        sub_id = state["sub_id"]

        # 将第一行 total_price 设为其他所有行总和的 10 倍（集中度 ≈ 91%）
        db = SessionLocal()
        try:
            rows = db.scalars(select(BidQuoteLine).where(BidQuoteLine.submission_id == sub_id)).all()
            other_sum = sum(r.total_price or 0 for r in rows[1:])
            rows[0].total_price = other_sum * 10  # ≈ 91% concentration
            db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(state["sub_id"]),
        })
        assert r.status_code == 422, r.text
        detail = r.json()["detail"]
        checks = [i["check"] for i in detail["failures"][0]["issues"]]
        assert "line_concentration" in checks, f"应触发 line_concentration，实际 checks={checks}"

    def test_declared_total_mismatch_blocked(self, compare_client):
        """明细合计与声明总价偏差>3%（通过 _doc_meta.bid_total）→ 422。

        必须用 flag_modified 显式标脏：JSON 列做"浅拷贝顶层 dict → 就地改嵌套
        dict → 整体重赋值"这套操作时，若嵌套 key（_doc_meta）已经存在（VL 路径
        接入声明总价抽取后就是如此），浅拷贝出的嵌套 dict 与 job.result 里的是
        **同一个对象**——原地改它，"新值"和 SQLAlchemy 认为的"历史值"其实指向
        同一份已被改过的数据，== 比较判不出差异，UPDATE 不会发出，改动看似成功
        实则从未落盘。这不是猜测：直接复现过（mutation 后立即读到新值，换一个
        全新 session 读，读到的还是旧值）。
        """
        from sqlalchemy.orm.attributes import flag_modified

        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine
        from apps.api.models.extraction_job import ExtractionJob

        state = self._base_setup(compare_client, "声明总价供应商", "声明总价测试")
        sub_id = state["sub_id"]

        db = SessionLocal()
        try:
            from apps.api.models.bid_submission import BidSubmission
            sub = db.get(BidSubmission, sub_id)
            rows = db.scalars(select(BidQuoteLine).where(BidQuoteLine.submission_id == sub_id)).all()
            actual_sum = sum(r.total_price or 0 for r in rows)

            # 将 ExtractionJob result._doc_meta.bid_total 设为 actual_sum 的 50%（巨大偏差）
            if sub.job_id:
                job = db.get(ExtractionJob, sub.job_id)
                if job and isinstance(job.result, dict):
                    result = dict(job.result)
                    result.setdefault("_doc_meta", {})
                    result["_doc_meta"]["bid_total"] = actual_sum * 0.5  # 100% deviation
                    job.result = result
                    flag_modified(job, "result")
                    db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(state["sub_id"]),
        })
        assert r.status_code == 422, r.text
        detail = r.json()["detail"]
        checks = [i["check"] for i in detail["failures"][0]["issues"]]
        assert "declared_total_mismatch" in checks, (
            f"应触发 declared_total_mismatch，实际 checks={checks}"
        )

    def test_single_line_exceeds_declared_total_blocked(self, compare_client):
        """单行金额超过声明总价 → 数学不可能，必须被拦截。

        flag_modified 的必要性同 test_declared_total_mismatch_blocked 上方注释。
        """
        from sqlalchemy.orm.attributes import flag_modified

        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine, BidSubmission
        from apps.api.models.extraction_job import ExtractionJob

        state = self._base_setup(compare_client, "超限行供应商", "超限行测试")
        sub_id = state["sub_id"]

        db = SessionLocal()
        try:
            sub = db.get(BidSubmission, sub_id)
            rows = db.scalars(select(BidQuoteLine).where(BidQuoteLine.submission_id == sub_id)).all()

            # 声明总价 = 100 yuan；将第一行 total_price 设为 50000 yuan（远超声明总价）
            declared = 100.0
            rows[0].total_price = 50000.0

            if sub.job_id:
                job = db.get(ExtractionJob, sub.job_id)
                if job and isinstance(job.result, dict):
                    result = dict(job.result)
                    result.setdefault("_doc_meta", {})
                    result["_doc_meta"]["bid_total"] = declared
                    job.result = result
                    flag_modified(job, "result")
            db.commit()
        finally:
            db.close()

        r = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(state["project_id"]), "category": "阀门",
            "supplier_ids": str(state["supplier_id"]),
            "submission_ids": str(state["sub_id"]),
        })
        assert r.status_code == 422, r.text
        detail = r.json()["detail"]
        checks = [i["check"] for i in detail["failures"][0]["issues"]]
        assert "line_exceeds_declared_total" in checks or "declared_total_mismatch" in checks, (
            f"应触发 line_exceeds_declared_total 或 declared_total_mismatch，实际 checks={checks}"
        )


# ── compare-state 刷新可恢复：按项目列已入库 submission + 在途 job ─────────────────

class TestCompareStateRestore:
    """GET /api/analysis/compare-state：刷新后重建供应商报价进度（confirmed + inflight）。"""

    def test_lists_confirmed_submissions_and_inflight_jobs(self, compare_client):
        # 建项目
        rp = compare_client.post("/api/projects", json={"name": "刷新恢复项目", "remark": ""})
        assert rp.status_code == 201
        pid = rp.json()["id"]

        # 上传+入库一份（带 project_id）→ 形成 confirmed submission
        ru = compare_client.post("/api/intake/upload",
            data={"type": "quote", "category": "阀门", "project_id": str(pid)},
            files={"file": ("conf.png", _png(), "image/png")})
        assert ru.status_code == 200
        job_conf = ru.json()["id"]
        rs = compare_client.post("/api/suppliers", json={"name": "恢复供应商A", "categories": ["阀门"]})
        supplier_id = rs.json()["id"]
        rc = compare_client.post("/api/quotes/batch-confirm", json={
            "job_id": job_conf, "supplier_id": supplier_id,
            "supplier_name": "恢复供应商A", "project_id": pid, "category": "阀门",
        })
        assert rc.status_code == 200, rc.text
        sub_id = rc.json()["submission_id"]

        # 再上传一份但不入库（带 project_id）→ 在途 job
        buf = io.BytesIO(); Image.new("RGB", (16, 16), (240, 240, 240)).save(buf, format="PNG")
        ru2 = compare_client.post("/api/intake/upload",
            data={"type": "quote", "category": "阀门", "project_id": str(pid)},
            files={"file": ("inflight.png", buf.getvalue(), "image/png")})
        assert ru2.status_code == 200
        job_inflight = ru2.json()["id"]

        # compare-state
        r = compare_client.get("/api/analysis/compare-state", params={"project_id": pid})
        assert r.status_code == 200, r.text
        data = r.json()

        subs = data["submissions"]
        assert len(subs) == 1, subs
        assert subs[0]["submission_id"] == sub_id
        assert subs[0]["line_count"] == 2
        assert subs[0]["supplier_raw_name"] == "恢复供应商A"
        assert subs[0]["job_id"] == job_conf

        inflight_ids = {j["job_id"] for j in data["inflight_jobs"]}
        assert job_inflight in inflight_ids, data["inflight_jobs"]
        # 已入库的 job 不应再出现在 inflight
        assert job_conf not in inflight_ids

        # 品类必须跟着回来（2026-08-23）。**这条覆盖的是一个真实缺陷**：
        # 前端 `restoreBatchFiles` 重建卡片时，已入库的条目走
        # `if (entry.confirmed) continue`，永远触发不到那个回填品类的识别回调，
        # 于是刷新一次品类就变回空串、点预览被"还没有确定品类"挡住——而系统
        # 手里明明有品类（这里的已入库报价行上就带着）。刷新恢复这条路此前
        # 没有任何测试覆盖，缺陷因此一直没被发现。
        assert data["category"] == "阀门", (
            f"刷新恢复必须能拿回品类，实得 {data['category']!r}——"
            "拿不回来的话用户刷新一次就被挡在预览外面")

    def test_empty_for_new_project(self, compare_client):
        rp = compare_client.post("/api/projects", json={"name": "空恢复项目", "remark": ""})
        pid = rp.json()["id"]
        r = compare_client.get("/api/analysis/compare-state", params={"project_id": pid})
        assert r.status_code == 200, r.text
        # `category` 2026-08-23 新增：全新项目没有任何证据，必须是空串而不是
        # 猜一个——空串是"该由用户手选"的信号，猜出来的品类会一路带到入库。
        assert r.json() == {"submissions": [], "inflight_jobs": [], "category": ""}


# ── batch-confirm 复活废弃 submission（"老问题"：超时/旧轮次 superseded 后再上传同一文件）──

class TestBatchConfirmRevive:
    """同一文件重传 → 同 job → 同 batch_id；若历史 submission 已 superseded，
    再确认必须复活为 pending 并重建行，绝不把废弃 id 当幂等命中返回（否则 match 必 409）。"""

    def _upload_confirm(self, client, supplier_name="复活供应商", project_name="复活项目"):
        r = client.post("/api/intake/upload",
            data={"type": "quote", "category": "阀门"},
            files={"file": ("revive.png", _png(), "image/png")})
        assert r.status_code == 200
        job_id = r.json()["id"]
        rs = client.post("/api/suppliers", json={"name": supplier_name, "categories": ["阀门"]})
        assert rs.status_code == 201
        supplier_id = rs.json()["id"]
        r1 = client.post("/api/quotes/batch-confirm", json={
            "job_id": job_id, "supplier_id": supplier_id,
            "supplier_name": supplier_name, "project_name": project_name, "category": "阀门",
        })
        assert r1.status_code == 200, r1.text
        return job_id, supplier_id, r1.json()

    def test_superseded_prior_is_revived_not_returned(self, compare_client):
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidQuoteLine, BidSubmission

        job_id, supplier_id, first = self._upload_confirm(compare_client)
        sub_id = first["submission_id"]
        assert first["line_count"] == 2

        # 模拟旧轮次/修复脚本把它 superseded
        db = SessionLocal()
        try:
            db.get(BidSubmission, sub_id).status = "superseded"
            db.commit()
        finally:
            db.close()

        # 重新确认同一 job → 必须复活，而非返回废弃 id
        r2 = compare_client.post("/api/quotes/batch-confirm", json={
            "job_id": job_id, "supplier_id": supplier_id,
            "supplier_name": "复活供应商", "category": "阀门",
        })
        assert r2.status_code == 200, r2.text
        body = r2.json()
        assert body["submission_id"] == sub_id, "应复用同一行（batch_id 唯一）"
        assert not body.get("idempotent"), "废弃行不得当幂等命中"
        assert body["line_count"] == 2, "行必须被重建"

        db = SessionLocal()
        try:
            assert db.get(BidSubmission, sub_id).status == "pending", "复活后状态须回到 pending"
            n = db.scalar(select(func.count()).select_from(BidQuoteLine).where(BidQuoteLine.submission_id == sub_id))
            assert n == 2, "旧行清空后按本次结果重建"
        finally:
            db.close()

    def test_revived_submission_passes_match_gate(self, compare_client):
        """复活后的 submission 能正常通过 match 硬闸门（复现并验证修复'老问题'）。"""
        from apps.api.core.database import SessionLocal
        from apps.api.models.bid_submission import BidSubmission

        job_id, supplier_id, first = self._upload_confirm(
            compare_client, supplier_name="复活闸门供应商", project_name="复活闸门项目")
        sub_id = first["submission_id"]
        project_id = first["project_id"]

        db = SessionLocal()
        try:
            db.get(BidSubmission, sub_id).status = "superseded"
            db.commit()
        finally:
            db.close()

        # 重新确认 → 复活
        compare_client.post("/api/quotes/batch-confirm", json={
            "job_id": job_id, "supplier_id": supplier_id,
            "supplier_name": "复活闸门供应商", "category": "阀门",
        })

        anchors = [{"seq": "1", "name": "DN100 闸阀", "spec": "Z45X-16Q", "unit": "个", "qty": 10, "category": "阀门"}]
        rc = compare_client.post("/api/analysis/tender-list/confirm", json={
            "project_id": project_id, "category": "阀门", "file_name": "t.xlsx",
            "anchors_json": anchors, "anchors_total": 1, "source_type": "excel",
        })
        assert rc.status_code == 200, rc.text
        rm = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(project_id), "category": "阀门",
            "supplier_ids": str(supplier_id), "submission_ids": str(sub_id),
        })
        assert rm.status_code == 200, f"复活后的 submission 应通过 match 闸门: {rm.text}"


# ── 价格口径桥接集成测试（ExtractionDraft → ExtractionResponse → batch-confirm →
#    BidQuoteLine），并入比价 E2E 主链路 ───────────────────────────────────────────

import time as _time
import uuid as _uuid

from apps.api.intelligence.extraction_draft import (
    DraftRow,
    ExtractionDraft,
    QualityReport,
    SourceRef,
)
from apps.api.intelligence.pipeline import ExtractionPipeline


def _draft_from_rows(field_rows: list[dict], flags_per_row: dict | None = None) -> ExtractionDraft:
    """构造一个最小合成 ExtractionDraft（全部 quote_line），供桥接契约测试。"""
    rows = []
    for i, f in enumerate(field_rows):
        rows.append(DraftRow(
            row_index=i, row_type="quote_line", raw_cells={},
            fields=dict(f), source_ref=SourceRef(page=1, table=0, row=i + 1),
            validation_flags=list((flags_per_row or {}).get(i, [])),
        ))
    return ExtractionDraft(
        doc_type="quote", source_file="synthetic.pdf",
        page_count=1, processed_page_count=1, target_pages=[1],
        rows=rows, meta={"supplier_name": "桥接合成"}, quality=QualityReport(status="REVIEW"),
    )


def _insert_quote_job(result_data: dict) -> str:
    """把 ExtractionResponse.data 落成 done 状态的 ExtractionJob，返回 job_id。"""
    from apps.api.core.database import SessionLocal
    from apps.api.models.extraction_job import ExtractionJob
    job_id = _uuid.uuid4().hex
    db = SessionLocal()
    try:
        db.add(ExtractionJob(id=job_id, type="quote", status="done", result=result_data))
        db.commit()
    finally:
        db.close()
    return job_id


def _chain_draft_to_bql(client, draft: ExtractionDraft, supplier_name: str, category: str = "阀门"):
    """ExtractionDraft → _draft_to_quote_response → ExtractionJob → batch-confirm。

    返回 (batch_confirm_response, supplier_id, submission_id_or_None)。
    """
    pipe = ExtractionPipeline(MockProvider())
    resp = pipe._draft_to_quote_response(draft, {}, _time.time())
    job_id = _insert_quote_job(resp.data)

    rs = client.post("/api/suppliers", json={"name": supplier_name, "categories": [category]})
    assert rs.status_code == 201, rs.text
    supplier_id = rs.json()["id"]

    r = client.post("/api/quotes/batch-confirm", json={
        "job_id": job_id, "supplier_id": supplier_id,
        "supplier_name": supplier_name, "project_name": f"{supplier_name}-桥接项目",
        "category": category,
    })
    sub_id = r.json().get("submission_id") if r.status_code == 200 else None
    return r, supplier_id, sub_id


def _bql_dicts(submission_id: int) -> list[dict]:
    """读出 BidQuoteLine 关键字段为纯 dict（避免 detached ORM 访问）。"""
    from apps.api.core.database import SessionLocal
    from apps.api.models.bid_submission import BidQuoteLine
    db = SessionLocal()
    try:
        rows = db.scalars(select(BidQuoteLine).where(BidQuoteLine.submission_id == submission_id)).all()
        return [{
            "raw_name": r.raw_name, "qty": r.qty,
            "unit_price": r.unit_price, "unit_price_excl_tax": r.unit_price_excl_tax,
            "tax_rate": r.tax_rate, "total_price": r.total_price,
            "extraction_meta": dict(r.extraction_meta or {}),
        } for r in rows]
    finally:
        db.close()


class TestPriceBasisBridgeContract:
    """合成 4 种口径 + unknown 阻断，验证 effective 与 extraction_meta 写入正确。"""

    def test_unspecified_uses_generic_price(self, compare_client):
        """仅通用单价/合价 → price_basis=unspecified，effective 取通用字段。"""
        draft = _draft_from_rows([
            {"name": "DN50 闸阀", "spec": "Z45X", "qty": 2.0,
             "unit_price": 100.0, "total_price": 200.0},
        ])
        r, _sid, sub_id = _chain_draft_to_bql(compare_client, draft, "口径-未标注")
        assert r.status_code == 200, r.text
        rows = _bql_dicts(sub_id)
        assert len(rows) == 1
        row = rows[0]
        assert row["unit_price"] == 100.0
        assert row["total_price"] == 200.0
        assert row["extraction_meta"]["price_basis"] == "unspecified"
        assert row["extraction_meta"]["effective_unit_price"] == 100.0

    def test_incl_tax_uses_incl_fields(self, compare_client):
        """仅含税（带税率，避免被降级）→ price_basis=incl_tax，effective 取含税。"""
        draft = _draft_from_rows([
            {"name": "DN50 闸阀", "spec": "Z45X", "qty": 2.0,
             "unit_price_incl_tax": 113.0, "total_price_incl_tax": 226.0, "tax_rate": 0.13},
        ])
        r, _sid, sub_id = _chain_draft_to_bql(compare_client, draft, "口径-含税")
        assert r.status_code == 200, r.text
        row = _bql_dicts(sub_id)[0]
        assert row["unit_price"] == 113.0
        assert row["total_price"] == 226.0
        meta = row["extraction_meta"]
        assert meta["price_basis"] == "incl_tax"
        assert meta["raw_unit_price_incl_tax"] == 113.0
        assert meta["tax_rate"] == 0.13

    def test_dual_tax_prefers_incl_and_keeps_excl(self, compare_client):
        """含税与不含税并存 → dual_tax，effective 取含税，excl 原值保留。"""
        draft = _draft_from_rows([
            {"name": "DN50 闸阀", "spec": "Z45X", "qty": 2.0,
             "unit_price_incl_tax": 113.0, "unit_price_excl_tax": 100.0,
             "total_price_incl_tax": 226.0, "total_price_excl_tax": 200.0,
             "tax_rate": 0.13, "tax_amount": 26.0},
        ])
        r, _sid, sub_id = _chain_draft_to_bql(compare_client, draft, "口径-双口径")
        assert r.status_code == 200, r.text
        row = _bql_dicts(sub_id)[0]
        assert row["unit_price"] == 113.0          # effective = 含税
        assert row["total_price"] == 226.0
        assert row["unit_price_excl_tax"] == 100.0  # 不含税原值保留在列
        meta = row["extraction_meta"]
        assert meta["price_basis"] == "dual_tax"
        assert meta["raw_unit_price_excl_tax"] == 100.0
        assert meta["raw_total_price_excl_tax"] == 200.0
        assert meta["tax_amount"] == 26.0

    def test_excl_tax_uses_excl_fields(self, compare_client):
        """仅不含税 → excl_tax，effective 取不含税，basis 保留（禁止与含税静默混比）。"""
        draft = _draft_from_rows([
            {"name": "DN50 闸阀", "spec": "Z45X", "qty": 2.0,
             "unit_price_excl_tax": 100.0, "total_price_excl_tax": 200.0, "tax_rate": 0.13},
        ])
        r, _sid, sub_id = _chain_draft_to_bql(compare_client, draft, "口径-不含税")
        assert r.status_code == 200, r.text
        row = _bql_dicts(sub_id)[0]
        assert row["unit_price"] == 100.0
        assert row["total_price"] == 200.0
        assert row["unit_price_excl_tax"] == 100.0
        assert row["extraction_meta"]["price_basis"] == "excl_tax"

    def test_incl_unit_recovered_from_total_when_missing(self, compare_client):
        """含税合价+数量齐全但缺含税单价 → 同口径还原 effective_unit=合价÷数量（泰科龙形态）。"""
        draft = _draft_from_rows([
            # 有含税单价的行：直接用
            {"name": "DN50 闸阀", "spec": "Z45X", "qty": 2.0,
             "unit_price_incl_tax": 113.0, "unit_price_excl_tax": 100.0,
             "total_price_incl_tax": 226.0, "total_price_excl_tax": 200.0,
             "tax_rate": 0.13, "tax_amount": 26.0},
            # 缺含税单价、有含税合价+不含税单价（泰科龙缺列形态）→ 应还原 4408.11/3=1469.37
            {"name": "DN80 闸阀", "spec": "Z45X", "qty": 3.0,
             "unit_price_incl_tax": None, "unit_price_excl_tax": 975.25,
             "total_price_incl_tax": 4408.11, "total_price_excl_tax": 2926.0,
             "tax_rate": 0.13},
        ])
        r, _sid, sub_id = _chain_draft_to_bql(compare_client, draft, "口径-还原单价")
        assert r.status_code == 200, r.text
        rows = sorted(_bql_dicts(sub_id), key=lambda x: x["raw_name"])
        # 两行都应有比价单价（unit_price 非空），覆盖率 2/2
        assert all(x["unit_price"] is not None and x["unit_price"] > 0 for x in rows), rows
        recovered = [x for x in rows if x["extraction_meta"].get("effective_unit_recovered")]
        assert len(recovered) == 1, "恰一行通过合价÷数量还原"
        assert abs(recovered[0]["unit_price"] - round(4408.11 / 3.0, 4)) < 1e-6
        assert recovered[0]["total_price"] == 4408.11  # 合价不变

    def test_unknown_basis_blocks_comparison(self, compare_client):
        """**明确不报价**的行 → 口径 unknown、effective 为 None，覆盖率门拦截 match。

        夹具用「/」而不是"什么都不填"：两者语义不同——「/」是供应商明确不报此项
        （合法事实，可入库、不参与金额比较），空白是"该有金额却没读到"（缺陷，
        由派生金额门在入库前阻断）。这里要测的是前者之后的下游行为。
        """
        draft = _draft_from_rows([
            {"name": "DN50 闸阀", "spec": "Z45X", "qty": 2.0, "total_price": "/"},
        ])
        r, supplier_id, sub_id = _chain_draft_to_bql(compare_client, draft, "口径-未知")
        assert r.status_code == 200, r.text
        row = _bql_dicts(sub_id)[0]
        assert row["unit_price"] is None
        assert row["total_price"] is None
        assert row["extraction_meta"]["price_basis"] == "unknown"

        # 招标清单 + match：unknown 行价格覆盖率 0% → 质量门 422（证明不自动入比价）
        # 需要 project_id；从 batch-confirm 响应取
        project_id = r.json()["project_id"]
        anchors = [{"seq": "1", "name": "DN50 闸阀", "spec": "Z45X", "unit": "个", "qty": 2, "category": "阀门"}]
        rc = compare_client.post("/api/analysis/tender-list/confirm", json={
            "project_id": project_id, "category": "阀门", "file_name": "t.xlsx",
            "anchors_json": anchors, "anchors_total": 1, "source_type": "excel",
        })
        assert rc.status_code == 200, rc.text
        rm = compare_client.post("/api/analysis/tender-list/match", data={
            "project_id": str(project_id), "category": "阀门",
            "supplier_ids": str(supplier_id), "submission_ids": str(sub_id),
        })
        assert rm.status_code == 422, f"unknown 口径行应被质量门拦截: {rm.text}"


# TestPriceBasisBridgeFixtures（三份真实 fixture 重放）删除于 2026-08-11（最佳实践
# 评审 F1）：它靠已删除的 legacy 链路（SnapshotProvider + recognize_tables +
# _get_quote_adapter）重放 OCR HTML 快照，随 legacy 一起失效。TestPriceBasisBridgeContract
# 保留了口径本身的合成契约覆盖（含泰科龙"含税合价还原单价"的形态，见
# test_incl_unit_recovered_from_total_when_missing）；损失的是三份真实历史文档的
# 全链路 89 行/总额回归，以及凯硕 qty_arithmetic_mismatch 字段落库的真实数据验证。
# 要补回需要用 scripts/record_vl_snapshots.py 为 miancun/kaishuo/taikelong 录制
# VL 快照（真实 API 调用，需人工触发），当前 tests/fixtures/vl_snapshots/ 只有四份
# 电缆文档，未覆盖这三份阀门文档。
