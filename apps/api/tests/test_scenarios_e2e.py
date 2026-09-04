"""场景用例套件：按用户给定的文件组合 × 现有三阶段（上传 / 预览 / 比价）。

## 为什么单独开一份，而不是往既有 E2E 里塞

既有的 `test_paddle_quote_api_e2e.py` 测的是**一条链路能不能走通**（两家供应商、
一种组合）。这里测的是**组合覆盖**：有没有采购清单、输入是 PDF 还是 Excel/CSV、
阀门项目还是电缆项目——这几个维度会走进完全不同的分支（锚点轴 vs 报价派生轴、
识别链路 vs 零模型调用），而在此之前没有任何一份测试把它们摆在一起过。

用例表就是规格本身。它由用户给出，不是我从代码反推的：改判据要先改这张表。

## 覆盖矩阵

| 用例 | 项目 | 报价输入 | 有无采购清单 | 行轴 | 模型调用 |
|------|------|----------|--------------|------|----------|
| A1 | 金桥(阀门) | 3 份投标 PDF | 招标 PDF | tender_anchor | 报价侧回放 |
| A2 | 金桥(阀门) | 3 份投标 PDF | 无 | quote_derived | 报价侧回放 |
| A3 | 金桥(阀门) | 3 份报价 xlsx | 采购清单 xlsx | tender_anchor | **零** |
| A4 | 金桥(阀门) | 3 份报价 xlsx | 无 | quote_derived | **零** |
| B1 | 徐汇(电缆) | 4 份投标 PDF | 招标 PDF | tender_anchor | 报价+招标回放 |
| B2 | 徐汇(电缆) | 4 份投标 PDF | 无 | quote_derived | 报价侧回放 |
| B3 | 徐汇(电缆) | 4 份报价 CSV | 采购清单 xlsx | tender_anchor | **零** |

金桥招标 PDF 是原生文字层，走 `tender_text_layer` 直抽（实测 89 行、零模型调用）；
徐汇招标 PDF 直抽判定不可信、整份回落 VL，故需要 `tender_xuhui.json` 快照。
这个差异是**文档本身的性质**，不是配置，用例表如实记下来。

## 阶段划分与各自该回答的问题

- **阶段一 上传**：每份文件都被识别出来了吗？行数对不对？金额对不对？
  —— 这一阶段的断言拿**标准答案清单**（供应商自己的 Excel/CSV）逐值比对，
  是本文件里唯一有"准确率"含义的一层。
- **阶段二 预览**：每一家都进得了预览吗？—— **预览永远不拦**是产品级不变式
  （预览不落库、结果强制 `basis="preview"`，拦它没有安全收益只有"功能不好使"）。
- **阶段三 比价**：正式链路的行轴类型对不对、矩阵列数对不对。

## 口径：为什么按"金额值集合"比，不按行位置比

各家 PDF 里的行序是各家自己排的，而参照清单是采购清单的顺序——逐位比会从第一处
差异起全线崩（实测把一份 87.4% 的文档算成 53.7%）。按规格串比也不行：电缆长型号
`WDZA-YJY-3*150+2*70` 错一个字符就匹配不上，测的就变成了字符识别而不是金额识别。
故按**合价值集合**比，并在断言里如实标注这是上界（值出现在结果里，不保证挂在正确
的物料上）。挂载正确性由阶段三的行轴断言另行覆盖。
"""
from __future__ import annotations

import collections
import copy
import csv
import functools
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import pytest

from apps.api.core.utils import parse_num

REPO = Path(__file__).resolve().parents[3]
DOCS = REPO / "tests" / "fixtures" / "documents"
SNAPS = REPO / "tests" / "fixtures" / "paddle_snapshots"


@dataclass(frozen=True)
class Case:
    """一个场景用例。`bids` 与 `reference` 一一对应（同一家的投标件与标准答案）。"""
    cid: str
    project: str
    bids: tuple[str, ...]                 # 报价输入文件名
    references: tuple[str, ...]           # 各家标准答案清单（供应商自己的 Excel/CSV）
    tender: str | None                    # 采购清单/招标文件；None = 无
    axis_kind: str                        # 期望的行轴类型
    snapshots: tuple[str, ...] = ()       # 需要回放的快照 slug（空 = 零模型调用）
    anchor_rows: int | None = None        # 采购清单期望行数


_JQ_BIDS_PDF = ("金桥地体上盖项目-泰科龙投标文件.pdf",
                "金桥地体上盖项目-凯硕新正投标文件.pdf",
                "金桥地体上盖项目-上海绵存投标文件.pdf")
_JQ_REFS = ("金桥地体上盖项目-泰科龙报价清单.xlsx",
            "金桥地体上盖项目-凯硕新正报价清单.xlsx",
            "金桥地体上盖项目-上海绵存报价清单.xlsx")
_JQ_SNAPS = ("quote_taikelong", "quote_kaishuo", "quote_miancun")

_XH_BIDS_PDF = ("徐汇区华泾镇项目-上海浦东投标文件.pdf",
                "徐汇区华泾镇项目-亨通投标文件.pdf",
                "徐汇区华泾镇项目-宏胜投标文件.pdf",
                "徐汇区华泾镇项目-远东投标文件.pdf")
_XH_REFS = ("徐汇区华泾镇项目-上海浦东报价清单.csv",
            "徐汇区华泾镇项目-亨通报价清单.csv",
            "徐汇区华泾镇项目-宏胜报价清单.csv",
            "徐汇区华泾镇项目-远东报价清单.csv")
_XH_SNAPS = ("quote_pudong", "quote_hengtong", "quote_hongsheng", "quote_yuandong")

CASES: tuple[Case, ...] = (
    Case("A1", "金桥", _JQ_BIDS_PDF, _JQ_REFS, "金桥地体上盖项目-招标文件.pdf",
         "tender_anchor", _JQ_SNAPS, anchor_rows=89),
    Case("A2", "金桥", _JQ_BIDS_PDF, _JQ_REFS, None, "quote_derived", _JQ_SNAPS),
    Case("A3", "金桥", _JQ_REFS, _JQ_REFS, "金桥地体上盖项目-采购清单.xlsx",
         "tender_anchor", (), anchor_rows=89),
    Case("A4", "金桥", _JQ_REFS, _JQ_REFS, None, "quote_derived", ()),
    Case("B1", "徐汇", _XH_BIDS_PDF, _XH_REFS, "徐汇区华泾镇项目-招标文件.pdf",
         "tender_anchor", _XH_SNAPS + ("tender_xuhui",), anchor_rows=90),
    Case("B2", "徐汇", _XH_BIDS_PDF, _XH_REFS, None, "quote_derived", _XH_SNAPS),
    Case("B3", "徐汇", _XH_REFS, _XH_REFS, "徐汇区华泾镇项目-采购清单.xlsx",
         "tender_anchor", (), anchor_rows=90),
)

CASE_IDS = [c.cid for c in CASES]


def _missing(case: Case) -> list[str]:
    out = [f for f in case.bids + case.references if not (DOCS / f).exists()]
    if case.tender and not (DOCS / case.tender).exists():
        out.append(case.tender)
    out += [f"{s}.json" for s in case.snapshots if not (SNAPS / f"{s}.json").exists()]
    return out


@pytest.fixture(params=CASES, ids=CASE_IDS)
def case(request) -> Case:
    c = request.param
    gaps = _missing(c)
    if gaps:
        pytest.skip(f"用例 {c.cid} 缺文件：{gaps}")
    return c


# ── 标准答案读取 ─────────────────────────────────────────────────────────────
# 三种载体（xlsx / csv）列名不同，靠表头文字定位，不写死列下标——换一份新语料
# 不用改代码。

_ALIAS = {
    "name": ("项目名称", "品名", "材料/设备名称", "材料（设备）名称", "名称"),
    "spec": ("规格", "规格型号"),
    "qty": ("数量",),
    "unit_price": ("单价(不含税)", "单价（不含税）", "单价"),
    "total_price": ("合计(不含税)", "合计（不含税）", "合价", "合计"),
}


def _rows_of(path: Path) -> list[list[str]]:
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig") as fh:
            return [list(r) for r in csv.reader(fh)]
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [["" if c is None else str(c) for c in r] for r in ws.iter_rows(values_only=True)]


@functools.cache
def _read_reference_cached(path: Path) -> list[dict]:
    """标准答案清单 → [{name, spec, qty, unit_price, total_price, not_quoted}]。

    **条目行的判据是"有数量"，不是"有合价"**——两个真实原因，都是写这份测试时
    被语料纠正过来的：

    1. **采购清单本来就没有价格**。那是留给投标人填的空表（design/32 同一结论）。
       按"有合价"过滤会把整份采购清单滤成 0 行。
    2. **合法的"不报价"行有数量、没价格**。实测徐汇一家在某条目上单价合价都印着
       "/"，参照清单的核对说明原文记着「PDF原表单价、合价均标为/」。按"有合价"
       过滤会把它当成不存在，于是各家条目数对不上——而那恰恰是 CLAUDE.md 要求
       与"读不到"分开标记的那个合法状态。

    表尾合计行靠"没有数量"被自然挡在外面（实测踩过：不滤会让参照总额凭空翻倍，
    因为合计行的合价恰好等于其余行之和）。
    """
    rows = _rows_of(path)
    hi = next(i for i, r in enumerate(rows) if any(str(c).strip() == "数量" for c in r))
    header = [str(c).strip() for c in rows[hi]]
    idx: dict[str, int] = {}
    for slot, names in _ALIAS.items():
        for n in names:
            if n in header:
                idx[slot] = header.index(n)
                break
    out: list[dict] = []
    for r in rows[hi + 1:]:
        rec = {k: (str(r[i]).strip() if i < len(r) else "") for k, i in idx.items()}
        if parse_num(rec.get("qty", "")) is None:
            continue
        rec["not_quoted"] = parse_num(rec.get("total_price", "")) is None
        out.append(rec)
    return out



def read_reference(path: Path) -> list[dict]:
    """缓存解析结果，但每次交出**新的行 dict**——理由同 `recognize_snapshot`：
    调用方拿到的是可变对象，共享一份迟早会有人就地改，然后在别处炸。
    行本身是扁平的标量 dict，逐行浅拷贝即可，不需要 deepcopy。"""
    return [dict(r) for r in _read_reference_cached(path)]

class TestReferenceLists:
    """标准答案本身先自证——参照数据不可信的话，后面所有准确率都没有意义。"""

    def test_every_reference_list_parses(self, case: Case):
        for name in case.references:
            rows = read_reference(DOCS / name)
            assert rows, f"{name} 没解析出任何条目行"

    def test_all_suppliers_in_a_project_quote_the_same_item_count(self, case: Case):
        counts = {n: len(read_reference(DOCS / n)) for n in case.references}
        assert len(set(counts.values())) == 1, f"同一项目各家条目数不一致：{counts}"

    def test_not_quoted_rows_are_a_first_class_state(self, case: Case):
        """"原文明确不报价"是合法状态，不是缺陷——它有数量、没价格。

        徐汇语料里真实存在这样一行（某家在 `HYA-2*0.5` 上单价合价都印 "/"）。
        断言它**被当成条目行保留**、并且被标成 not_quoted，是在守 CLAUDE.md
        「"原文明确不报价"与"读不到"必须分开标记，不得合并成同一个空值语义」。
        """
        per_file = {n: [r for r in read_reference(DOCS / n) if r["not_quoted"]]
                    for n in case.references}
        total = sum(len(v) for v in per_file.values())
        if total == 0:
            pytest.skip("本项目语料里没有不报价行")
        for name, rows in per_file.items():
            for r in rows:
                assert parse_num(r["qty"]) is not None, f"{name}：不报价行必须仍带数量"

    def test_quantities_are_identical_across_suppliers(self, case: Case):
        """同一项目各家报的是同一份清单，数量必须逐行相同；单价必须不同。

        这条既是对语料的校验，也是 design/32 "数量序列可作对齐键" 的前提——
        前提不成立的话，报价派生轴（A2/A4/B2）就没有立足点。
        """
        refs = [read_reference(DOCS / n) for n in case.references]
        qty_seqs = [tuple(parse_num(r["qty"]) for r in rows) for rows in refs]
        assert len(set(qty_seqs)) == 1, "各家数量序列不一致，报价派生轴的前提不成立"
        price_seqs = {tuple(parse_num(r.get("unit_price") or "") for r in rows) for rows in refs}
        assert len(price_seqs) == len(refs), "各家单价完全相同——参照清单可能是同一份被复制"

    def test_anchor_list_matches_expected_size(self, case: Case):
        if case.tender is None or case.anchor_rows is None:
            pytest.skip("本用例没有采购清单")
        if Path(case.tender).suffix.lower() != ".xlsx":
            pytest.skip("招标 PDF 的行数由阶段一断言，这里只校验 Excel 采购清单")
        assert len(read_reference(DOCS / case.tender)) == case.anchor_rows

    def test_anchor_and_quotes_may_differ_in_length(self, case: Case):
        """采购清单与报价清单**不保证等长**，这是真实业务形态不是数据错误。

        金桥两侧都是 89 条、1:1；徐汇的采购清单只有 90 条而各家报价 136 条——
        采购清单是报价的子集。对齐链路必须能处理后者，用例表因此把两个项目都
        纳进来，而不是只测对得上的那个。
        """
        if case.tender is None or Path(case.tender).suffix.lower() != ".xlsx":
            pytest.skip("本用例没有 Excel 采购清单")
        anchor = len(read_reference(DOCS / case.tender))
        quote = len(read_reference(DOCS / case.references[0]))
        assert anchor > 0 and quote > 0
        assert (anchor == quote) is (case.project == "金桥"), (
            f"{case.project}：采购清单 {anchor} 条 vs 报价 {quote} 条，与用例表记录的形态不符")


# ═══════════════════════════════════════════════════════════════════════════
# 阶段一：上传（识别准确率基线）
# ═══════════════════════════════════════════════════════════════════════════
#
# 只对有 Paddle 快照的 PDF 用例断言——这是"识别引擎产出的行有多准"这件事唯一
# 有意义的地方；Excel/CSV 走零模型确定性解析（`tabular_ingestion.py`），阶段一
# 对它们没有"准确率"含义，跳过。
#
# 数字是**测量得到的基线，不是目标值**：先测出当前实际水平，再守住不退步。
# 改判据/改引擎让数字变好是预期之内的（改基线），变差要能解释为什么。
#
# 教训记在这（写测试的过程中真的踩到过一次）：算总额差值时，参照列取的是
# "合计(不含税)"（`_ALIAS` 里排在"合价"/"合计"之前），识别侧如果按
# `total_price → incl_tax → excl_tax` 的顺序取值，凯硕这份文档因为没填泛用的
# `total_price` 槽位会退到"价税合计"（含税），于是拿不含税的参照去比含税的
# 识别值，凭空比出 13% 的"税差"——不是数据问题，是比较口径不对齐。这里统一
# 用 `_total_of()`，优先级跟 `read_reference` 的列选择口径对齐（不含税优先）。

SUPPLIER_SNAPSHOT_PAGE_COUNTS = {
    "quote_taikelong": 53, "quote_kaishuo": 19, "quote_miancun": 31,
    "quote_hengtong": 11, "quote_hongsheng": 11, "quote_pudong": 15,
    "quote_yuandong": 19,
}

# 快照 slug → 标准答案文件名（一一对应，跟 CASES 里的 references 顺序无关，
# 这里是全量映射，供阶段一独立断言用）。
SNAPSHOT_REFERENCE = {
    "quote_taikelong": "金桥地体上盖项目-泰科龙报价清单.xlsx",
    "quote_kaishuo": "金桥地体上盖项目-凯硕新正报价清单.xlsx",
    "quote_miancun": "金桥地体上盖项目-上海绵存报价清单.xlsx",
    "quote_hengtong": "徐汇区华泾镇项目-亨通报价清单.csv",
    "quote_hongsheng": "徐汇区华泾镇项目-宏胜报价清单.csv",
    "quote_pudong": "徐汇区华泾镇项目-上海浦东报价清单.csv",
    "quote_yuandong": "徐汇区华泾镇项目-远东报价清单.csv",
}

# 2026-08-22 用真实快照测得（见本文件 git log 同日提交的量测脚本）。
# rows/ref_rows：识别产出的明细行数 / 标准答案条目数（不要求相等）。
# blank_total：三种合价字段都拿不到值的行数。column_shift：design/34 检测到
# 位置映射坏掉的行数。total_delta_pct：识别合价合计相对标答合价合计的偏差
# ——负值是少算（多半是 blank_total 造成），不是超收。
#
# **2026-08-23 更正**：本段原先写着"远东多出的 2 行是重复行，design/32 的副本
# 检测已处理"——错的。实测拆开是 `138 = 136 + 3 − 1`：招标正文的三个段落
# （"采购文件" / "简要内容" / "招标文件中未明确"）被当成物料行读了进来，同时
# 真丢了一行预分支电缆头（合价 551.89）。跟副本检测无关。净值 +2 把两侧的错
# 各自藏起来了——这正是下面要补逐行断言的原因。
#
# judged/exact/blank/flagged/silent 见 `value_audit()`。
UPLOAD_BASELINE = {
    "quote_taikelong": dict(rows=89, ref_rows=89, blank_total=9, column_shift=6, total_delta_pct=-26.22,
                            judged=49, exact=44, blank=5, flagged=0, silent=0),
    "quote_kaishuo": dict(rows=89, ref_rows=89, blank_total=0, column_shift=1, total_delta_pct=-2.02,
                          judged=57, exact=56, blank=0, flagged=0, silent=1),
    "quote_miancun": dict(rows=89, ref_rows=89, blank_total=1, column_shift=0, total_delta_pct=-0.01,
                          judged=51, exact=50, blank=1, flagged=0, silent=0),
    "quote_hengtong": dict(rows=132, ref_rows=136, blank_total=10, column_shift=1, total_delta_pct=-7.82,
                           judged=81, exact=63, blank=5, flagged=1, silent=12),
    "quote_hongsheng": dict(rows=136, ref_rows=136, blank_total=1, column_shift=0, total_delta_pct=-0.01,
                            judged=78, exact=78, blank=0, flagged=0, silent=0),
    "quote_pudong": dict(rows=130, ref_rows=136, blank_total=12, column_shift=0, total_delta_pct=-7.2,
                         judged=72, exact=57, blank=8, flagged=0, silent=7),
    "quote_yuandong": dict(rows=138, ref_rows=136, blank_total=16, column_shift=32, total_delta_pct=-2.59,
                           judged=113, exact=92, blank=10, flagged=5, silent=6),
}

# 基线允许的漂移——回放本应逐字节确定性，理论上应为 0；给一点容差防止浮点
# 累加误差之类的假阳性，不是给真实回归留口子。
_ROWS_TOLERANCE = 0
_BLANK_TOLERANCE = 1
_DELTA_PCT_TOLERANCE = 1.5   # 百分点
# 静默错值不给容差：**它是这组断言唯一要守的东西**。金额读错却不带 flag，
# 意味着系统不知道自己错了，下游没有任何机会拦住它。多一行都要红。
_SILENT_TOLERANCE = 0


def _num(fields: dict, key: str) -> float | None:
    v = fields.get(key)
    try:
        return float(v) if str(v).strip() else None
    except (TypeError, ValueError):
        return None


def _total_of(fields: dict) -> float | None:
    """跟 `read_reference` 的列选择口径对齐：不含税优先。"""
    return (_num(fields, "total_price_excl_tax") or _num(fields, "total_price")
            or _num(fields, "total_price_incl_tax"))


@functools.cache
def _recognize_snapshot_cached(slug: str):
    """回放一份 Paddle 快照，返回去重副本后的明细行（`DraftRow` 列表）。

    **按 slug 记忆化**：本文件有 6 个断言方法各自 parametrize 全部 7 个 slug，
    不缓存就是同一份文档被完整重解析 42 次（其中只有 7 次是不同输入）。这是
    纯函数——输入只有 slug、快照 JSON 在磁盘上是只读的、解析确定性；调用方
    只读返回的行，全文件没有一处写 `DraftRow`。所以缓存不改变任何断言语义，
    只是不再重复算。"""
    from apps.api.intelligence.paddle_vl import recognize_quote_paddle

    doc = json.loads((SNAPS / f"{slug}.json").read_text(encoding="utf-8"))
    draft = recognize_quote_paddle(
        "x.pdf", submit_and_parse=lambda *a, **k: doc,
        page_count=SUPPLIER_SNAPSHOT_PAGE_COUNTS[slug])
    rows = [r for r in draft.rows if r.row_type == "quote_line"]
    copy_nos = {str(r.fields.get("copy_no") or "") for r in rows}
    if len(copy_nos) > 1:
        first = sorted(copy_nos)[0]
        rows = [r for r in rows if str(r.fields.get("copy_no") or "") == first]
    return rows


def recognize_snapshot(slug: str):
    """缓存解析结果，但**每次返回深拷贝**。

    直接把缓存对象交出去是错的，实测踩过：`test_gap_fill.py::_real_rows()` 也调
    这个函数，而 `gap_fill.fill_gaps(rows, ...)` 会**就地写** `DraftRow`——补位
    本来就是往空格子里填值。共享同一份对象时，先跑的用例把格子填满，后跑的
    `find_gaps` 就什么都找不到了，断言在一个跟自己无关的地方失败。

    深拷贝相对整份文档重解析可以忽略不计，缓存的收益仍在。"""
    return copy.deepcopy(_recognize_snapshot_cached(slug))


def _audit_key(name, spec) -> str:
    """把品名+规格压成一个可比对的键。

    两边的写法差异是真实存在的、不是脏数据：徐汇的参照 CSV 把"预分支电缆头"
    写在**规格列**里，识别侧则拆成 name/spec 两格；全半角、`×`/`X`/`*`、空格
    和括号也各写各的。拼起来再归一化，两种写法就落到同一个键上。
    """
    s = unicodedata.normalize("NFKC", f"{name or ''}{spec or ''}").upper()
    return re.sub(r"[\s（）()]", "", s.replace("×", "*").replace("X", "*"))


@functools.cache
def value_audit(slug: str) -> dict:
    """逐行合价审计：识别出来的这一行，**值本身**对不对。

    存在的理由（2026-08-23 实测抓出来的）：亨通那份文档整块金额往上错了一行
    ——4*70+E35 拿到 4*50+E25 的合价、4*50+E25 拿到 4*35+E16 的，一条链 12 行。
    可 `rows` 仍是 132、`blank_total` 仍是 10、`column_shift` 仍是 1、
    `total_delta_pct` 仍是 −7.82%：上面四条断言**一条都不会红**。因为它们量的
    全是聚合量，而错位既不改行数也不改总额太多。金额都在、数量级都合理、
    `validation_flags` 是空的——只有逐行比对能看见。

    **只审"两边各只出现一次"的键。** 同名同规格在一份清单里可以重复出现
    （金桥的阀门规格就只有 DN20/DN25 这种，靠品名才分得开；徐汇也有重复行），
    重复键配错参照行会造出一堆假阳性——实测这个坑：早一版按规格单键匹配，宏胜
    136/136 行、合计只差 0.01%，却被算出"9 行静默错"，全是配错行。宁可少判
    一些行，不可判错。判不动的行数记在 `judged` 里，它自己也是个基线：它变小
    说明品名/规格被读坏了（泰科龙 89 行只有 49 行可判，就是名称列拖尾造成的）。

    四个出口互斥：
      exact   —— 合价与标答一致
      blank   —— 三种合价字段都取不到值（design/33 的识别空洞）
      flagged —— 值错，但 `validation_flags` 非空：**系统知道自己可能错了**
      silent  —— 值错且无任何 flag：系统不知道，下游也无从拦截
    """
    rows = recognize_snapshot(slug)
    ref = read_reference(DOCS / SNAPSHOT_REFERENCE[slug])

    ref_keys = collections.Counter(_audit_key(r.get("name"), r.get("spec")) for r in ref)
    got_keys = collections.Counter(_audit_key(r.fields.get("name"), r.fields.get("spec")) for r in rows)
    unique = {k for k in ref_keys if k and ref_keys[k] == 1 and got_keys.get(k) == 1}
    ref_total = {_audit_key(r.get("name"), r.get("spec")): parse_num(r.get("total_price") or "")
                 for r in ref}

    out = dict(judged=len(unique), exact=0, blank=0, flagged=0, silent=0, wrong_rows=[])
    for r in rows:
        key = _audit_key(r.fields.get("name"), r.fields.get("spec"))
        if key not in unique:
            continue
        got = _total_of(r.fields)
        want = ref_total[key]
        if got is not None and want is not None and round(got, 2) == round(float(want), 2):
            out["exact"] += 1
        elif got is None:
            out["blank"] += 1
        elif r.validation_flags:
            out["flagged"] += 1
        else:
            out["silent"] += 1
            out["wrong_rows"].append((r.fields.get("spec"), got, want))
    return out


@pytest.mark.slow  # 7 份快照 × 6 条断言，每条都要把整份文档解析一遍
@pytest.mark.parametrize("slug", sorted(UPLOAD_BASELINE), ids=sorted(UPLOAD_BASELINE))
class TestUploadStage:
    """阶段一：每份报价文件识别出来的行，跟标准答案比对。"""

    def test_row_count_matches_baseline(self, slug):
        if not (SNAPS / f"{slug}.json").exists():
            pytest.skip(f"快照缺失：{slug}.json")
        rows = recognize_snapshot(slug)
        expected = UPLOAD_BASELINE[slug]["rows"]
        assert abs(len(rows) - expected) <= _ROWS_TOLERANCE, (
            f"{slug} 识别行数 {len(rows)}，基线 {expected}——"
            "变多变少都要能解释，不能默认接受")

    def test_blank_total_within_baseline(self, slug):
        if not (SNAPS / f"{slug}.json").exists():
            pytest.skip(f"快照缺失：{slug}.json")
        rows = recognize_snapshot(slug)
        blank = sum(1 for r in rows if _total_of(r.fields) is None)
        expected = UPLOAD_BASELINE[slug]["blank_total"]
        assert blank <= expected + _BLANK_TOLERANCE, (
            f"{slug} 无合价行数 {blank}，基线 {expected}——识别退化了")

    def test_column_shift_within_baseline(self, slug):
        """design/34 检测到的位置映射坏掉行数不该无声变多。"""
        if not (SNAPS / f"{slug}.json").exists():
            pytest.skip(f"快照缺失：{slug}.json")
        rows = recognize_snapshot(slug)
        shift = sum(1 for r in rows if "column_shift" in r.validation_flags)
        expected = UPLOAD_BASELINE[slug]["column_shift"]
        assert shift <= expected + _BLANK_TOLERANCE, (
            f"{slug} column_shift 行数 {shift}，基线 {expected}")

    def test_total_delta_within_baseline(self, slug):
        """合价合计相对标准答案的偏差——本文件里唯一有"准确率"含义的数字。"""
        if not (SNAPS / f"{slug}.json").exists():
            pytest.skip(f"快照缺失：{slug}.json")
        rows = recognize_snapshot(slug)
        ref = read_reference(DOCS / SNAPSHOT_REFERENCE[slug])
        ref_sum = sum(parse_num(r["total_price"]) for r in ref if not r["not_quoted"])
        got_sum = sum(_total_of(r.fields) or 0 for r in rows)
        delta_pct = (got_sum - ref_sum) / ref_sum * 100 if ref_sum else 0.0
        expected = UPLOAD_BASELINE[slug]["total_delta_pct"]
        assert abs(delta_pct - expected) <= _DELTA_PCT_TOLERANCE, (
            f"{slug} 合价合计偏差 {delta_pct:.2f}%，基线 {expected}%："
            f"识别 {got_sum:,.0f} / 标答 {ref_sum:,.0f}")

    def test_silent_wrong_amounts_do_not_grow(self, slug):
        """**静默错值**——金额读错却不带 flag——不许变多。零容差。

        上面四条断言全是聚合量，跨行整体错位在它们眼里是不存在的（见
        `value_audit` 文档）。这一条是唯一能看见它的。
        """
        if not (SNAPS / f"{slug}.json").exists():
            pytest.skip(f"快照缺失：{slug}.json")
        got = value_audit(slug)
        expected = UPLOAD_BASELINE[slug]["silent"]
        detail = "".join(
            f"\n      {spec} 识别={g} 标答={w}" for spec, g, w in got["wrong_rows"][:8])
        assert got["silent"] <= expected + _SILENT_TOLERANCE, (
            f"{slug} 静默错值 {got['silent']} 行，基线 {expected}——"
            f"金额读错且系统不知道自己错了，下游没有任何机会拦截：{detail}")

    def test_row_value_accuracy_within_baseline(self, slug):
        """逐行合价正确率不许下滑，可判定行数也不许下滑。

        `judged` 一起守是有意的：把品名/规格读坏可以让一行"判不动"，于是
        `silent` 变小——不守 `judged` 的话，识别退化反而会让上一条断言变绿。
        """
        if not (SNAPS / f"{slug}.json").exists():
            pytest.skip(f"快照缺失：{slug}.json")
        got = value_audit(slug)
        base = UPLOAD_BASELINE[slug]
        assert got["judged"] >= base["judged"] - _BLANK_TOLERANCE, (
            f"{slug} 可判定行数 {got['judged']}，基线 {base['judged']}——"
            "品名或规格被读坏了，行本身还在但对不上标答")
        assert got["exact"] >= base["exact"] - _BLANK_TOLERANCE, (
            f"{slug} 逐行合价正确 {got['exact']}/{got['judged']}，"
            f"基线 {base['exact']}/{base['judged']}")


class TestUploadStageZeroModel:
    """A3/A4/B3：报价直接是 Excel/CSV，走零模型确定性解析。

    **2026-08-23 更正。** 这里原先写着「走既有的 `read_reference` 即可、不需要
    再包一层 `extract_quote_tabular`（那条路径本身是纯 pandas 读取，数字契约与
    `read_reference` 相同）」——那个假设是错的，而且**它就是缺陷能活下来的理由**：
    `read_reference` 会扫描前几行找表头，`extract_quote_tabular` 当时写死
    `header=0`。两份采购清单第一行都是项目标题，于是测试侧读得好好的、生产侧
    一律报「未识别到物料名称列。实际列名：['<整行标题>', 'Unnamed: 1', ...]」。

    教训一般化：**测试助手不得替代被测的生产函数**，哪怕看起来"契约相同"——
    看起来相同正是没人去比对的原因。下面两条现在直接调生产入口。
    """

    @pytest.mark.parametrize("case", [c for c in CASES if not c.snapshots], ids=lambda c: c.cid)
    def test_production_parser_reads_every_quote_file(self, case: Case):
        """报价文件必须能被**生产解析器**读出条目，不是被测试助手读出。"""
        from apps.api.services.ingestion.tabular_ingestion import extract_quote_tabular
        for name in case.references:
            path = DOCS / name
            if not path.exists():
                pytest.skip(f"缺文件：{name}")
            result = extract_quote_tabular(str(path), {"project_id": 0, "category": ""})
            assert result["items"], f"{name}：生产解析器读出 0 条，用户上传这份文件会直接失败"

    @pytest.mark.parametrize("case", [c for c in CASES if not c.snapshots], ids=lambda c: c.cid)
    def test_every_input_file_survives_the_classifier(self, case: Case):
        """**门口这一步也要测。** 用例表里的每份文件都得先过 `classify_tier0`。

        阶段二/三上传时类型是测试自己填好的（`_upload_quote(type=...)`），从来
        不调 `/api/intake/classify-tier0`——而真实界面第一步就是它。后果实测过：
        四份报价 CSV 在 `classify_tier0` 里返回 None、路由答"不支持的文件类型"，
        B3 用例却全绿。测试替用户做了一个真实流程里做不到的动作（跳过分类），
        绿色因此不代表用户能走通。
        """
        from apps.api.intelligence.document_classify import classify_tier0
        files = list(case.bids) + list(case.references) + ([case.tender] if case.tender else [])
        for name in files:
            path = DOCS / name
            if not path.exists():
                pytest.skip(f"缺文件：{name}")
            assert classify_tier0(str(path)) is not None, (
                f"{name}：分类返回 None——界面会显示「不支持的文件类型」，"
                "用户在第一步就被挡住，后面的阶段测得再绿也到不了")

    @pytest.mark.parametrize("case", [c for c in CASES if not c.snapshots], ids=lambda c: c.cid)
    def test_reference_itself_has_no_missing_quantity(self, case: Case):
        for name in case.references:
            rows = read_reference(DOCS / name)
            assert all(parse_num(r["qty"]) is not None for r in rows), (
                f"{name}：零模型路径下数量必须全部可解析，否则整份上传会被 "
                "`extract_quote_tabular` 判定为零有效条目而报错")


# ═══════════════════════════════════════════════════════════════════════════
# 阶段二 / 阶段三：预览与比价——走真实 HTTP API，走完整条链路
# ═══════════════════════════════════════════════════════════════════════════
#
# 阶段一只测"识别本身准不准"，不碰数据库、不碰路由。阶段二/三验证完全不同的
# 一层："识别结果能不能真的流过上传→预览→比价这条真实链路"——这正是本轮人工
# 测试抓出回归的地方（列错位检测上线后，`_gate_integrity` 漏接
# `gates_advisory`，导致预览把整份供应商拦在门外）：那类 bug 在阶段一的纯函数
# 级测试里根本不可见，必须走真实路由才能抓到。

from fastapi.testclient import TestClient  # noqa: E402

SUPPLIER_KEYWORDS = ("泰科龙", "凯硕", "绵存", "亨通", "宏胜", "浦东", "远东")


def supplier_keyword(filename: str) -> str:
    for kw in SUPPLIER_KEYWORDS:
        if kw in filename:
            return kw
    raise ValueError(f"无法从文件名识别供应商关键字：{filename}")


KEYWORD_TO_SLUG = {
    "泰科龙": "quote_taikelong", "凯硕": "quote_kaishuo", "绵存": "quote_miancun",
    "亨通": "quote_hengtong", "宏胜": "quote_hongsheng", "浦东": "quote_pudong",
    "远东": "quote_yuandong",
}


@pytest.fixture
def client(tmp_path, monkeypatch, temp_db):
    """TestClient + Paddle 快照回放（覆盖两个项目全部供应商 + 徐汇招标件）+
    免登录。跟 `test_paddle_quote_api_e2e.py` 的 `api` fixture 同一套手法，
    这里额外覆盖徐汇一侧、且不绑定某一组固定供应商。
    """
    monkeypatch.setenv("UPLOAD_DIR", str(tmp_path / "uploads"))
    import apps.api.core.config as config_mod
    config_mod._settings = None

    from apps.api.intelligence.paddle_snapshot import PaddleSnapshotReplay
    from apps.api.intelligence.pipeline import ExtractionPipeline
    from apps.api.intelligence.providers import paddle_ocr
    from apps.api.tests.test_paddle_quote_api_e2e import _UnusedTenderOnlyProvider

    keyword_to_slug = dict(KEYWORD_TO_SLUG)
    keyword_to_slug["招标"] = "tender_xuhui"
    available = {k: v for k, v in keyword_to_slug.items()
                if (SNAPS / f"{v}.json").exists()}
    replay = PaddleSnapshotReplay.from_slugs(available, snapshot_dir=SNAPS)
    monkeypatch.setattr(paddle_ocr, "submit_and_parse", replay.submit_and_parse)

    # 招标文字层直抽的封面标量调用会打到 provider.vl_extract_csv——
    # `_UnusedTenderOnlyProvider` 会抛 NotImplementedError，但
    # `extract_tender_meta` 自己 catch 住只留空（"清单才是主线"），不影响清单本身。
    import apps.api.intelligence.paddle_doc_meta as paddle_doc_meta_mod
    monkeypatch.setattr(paddle_doc_meta_mod, "get_text_client_call",
                        lambda: (lambda prompt: ""))
    monkeypatch.setattr("apps.api.main._build_pipeline",
                        lambda: ExtractionPipeline(provider=_UnusedTenderOnlyProvider()))

    from apps.api.core.security import get_current_user
    from apps.api.main import app

    app.dependency_overrides[get_current_user] = lambda: {"sub": "scenarios-e2e", "role": "管理员"}
    with TestClient(app) as c:
        yield c, replay
    app.dependency_overrides.clear()
    config_mod._settings = None


def _wait_done(client, job_id: str) -> dict:
    """`intake.py` 用线程池 inline 执行；上传响应本身不可信（拿的是调用前的
    ORM 快照），必须重新 GET 才能看到终态——真实前端也是靠轮询拿状态。"""
    r = client.get(f"/api/intake/jobs/{job_id}")
    assert r.status_code == 200, r.text
    job = r.json()
    assert job["status"] == "done", f"job {job_id} 未完成：{job}"
    return job


def _resolve_review_rows(items: list[dict]) -> list[dict]:
    """跟 `test_paddle_quote_api_e2e._resolve_review_rows` 同一套模拟——"人工在
    疑点收件箱里补全"：只补"数量和单价都读到、只是合价没被单独抽出来"的行；
    列错位行只承认（`integrity_ack`），不凭空补值——那些数字在识别产物里
    本来就不存在（design/34 §2.1）。**用于预览**（`gates_advisory=True`，
    走不到 `missing_total_requires_review`，不需要下面那个更狠的兜底）。
    """
    fixed = []
    for it in items:
        if "column_shift" in (it.get("validation_flags") or []):
            it = dict(it, integrity_ack=True)
        has_total = any(it.get(k) is not None
                        for k in ("total_price", "total_price_incl_tax", "total_price_excl_tax"))
        if has_total or it.get("not_quoted"):
            fixed.append(it)
            continue
        qty, price = it.get("qty"), it.get("unit_price")
        if qty is not None and price is not None:
            it = dict(it, total_price=round(qty * price, 2))
        fixed.append(it)
    return fixed


def _resolve_for_official_confirm(items: list[dict]) -> list[dict]:
    """正式入库用。`missing_total_requires_review` 是有意设计成**没有 ack
    字段**的闸门——"试点期采用最保守规则：单行即阻断，不用占比阈值"（亨通
    实测单行列错位造成约 2000 万误差）。系统不会代为计算，唯一的合法出口是
    人工翻看原文后手动填一个数（`total_is_manual=True`）。

    这条本身不是缺陷，是 design/33 那批空洞（泰科龙第 10 页等）该有的行为：
    没读到的钱，必须有人真的看过原文才能入账。测试模拟的是"人工确实做了这个
    动作"，不是让系统绕过它——**填的是占位值 0，不是猜测的真实金额**，因为
    测试代码不知道真实原文写的是什么数（如果知道，那就该去修识别，不是补测试）。
    这批行永远带着 `total_is_manual`，下游可查，不会被误当成识别读到的值。
    """
    fixed = []
    for it in _resolve_review_rows(items):
        has_total = any(it.get(k) is not None
                        for k in ("total_price", "total_price_incl_tax", "total_price_excl_tax"))
        if has_total or it.get("not_quoted"):
            fixed.append(it)
            continue
        fixed.append(dict(it, total_price=0.0, total_is_manual=True))
    return fixed


def _upload_quote(client, replay, *, path: Path, project_id: int,
                  supplier_id: int, category: str) -> dict:
    """上传一份报价文件（PDF 走 Paddle 回放；xlsx/csv 走零模型确定性解析），
    等它跑完，返回 job 字典。"""
    if path.suffix.lower() == ".pdf":
        replay.current = supplier_keyword(path.name)
    r = client.post("/api/intake/upload",
                    files={"file": (path.name, path.read_bytes(), "application/octet-stream")},
                    data={"type": "quote", "project_id": str(project_id),
                          "supplier_id": str(supplier_id), "category": category})
    assert r.status_code in (200, 201), f"{path.name} 上传失败：{r.text}"
    return _wait_done(client, r.json()["id"])


def _recognize_tender(client, replay, *, tender_path: Path, project_id: int) -> dict:
    """采购清单/招标文件 → 识别产物（含 `items` 与 `detected_category`）。

    **不接收 category 参数**：品类是这一步的**产出**，不是输入。之前这个函数
    收一个 category 再原样传给 `anchor_to_json`，等于测试把答案先塞进去、再从
    另一头把它取出来——自证式的假绿（`.claude/rules/tests.md`：标准答案先审计
    来源，不得循环验证）。

    PDF 走上传+识别（金桥原生文字层零模型，徐汇回落 Paddle、需 tender_xuhui
    快照）；xlsx 走 `parse_tender_xlsx` + `classify_category`，跟
    `/tender-list/preview` 路由内部完全同一套调用，不必真的过一轮 HTTP。
    """
    if tender_path.suffix.lower() == ".pdf":
        replay.current = "招标"
        r = client.post("/api/intake/upload",
                        files={"file": (tender_path.name, tender_path.read_bytes(),
                                        "application/pdf")},
                        data={"type": "tender_bidlist", "project_id": str(project_id)})
        assert r.status_code in (200, 201), f"招标文件上传失败：{r.text}"
        job = _wait_done(client, r.json()["id"])
        return job.get("result") or {}

    # xlsx：**走真实路由**。
    #
    # 2026-08-23 之前这里复刻了 `/tender-list/preview` 的内部逻辑，理由写着
    # "跟路由内部完全同一套调用，不必真的过一轮 HTTP"。这个理由是错的，而且
    # 它掩盖了一个真实缺陷：路由挑 Sheet（`pick_default_sheet`），复刻版没挑，
    # 于是徐汇那份两张表的采购清单在测试里被读成一张、在产品里也被读成一张，
    # 两边错得一模一样，测试因此永远发现不了。跟本文件
    # `TestUploadStageZeroModel` 记的是同一条教训：**测试助手不得替代被测的
    # 生产函数**——"看起来一样"正是没人去比对的原因。
    r = client.post(
        "/api/analysis/tender-list/preview",
        files={"file": (tender_path.name, tender_path.read_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, f"采购清单预览失败：{r.text}"
    data = r.json()
    return {
        "items": data.get("items") or [],
        "detected_category": data.get("detected_category") or "",
        "row_count": data.get("total") or 0,
    }


# B1 是一份**记录在案、不打算在这里修的真实数据缺口**，不是识别或代码 bug：
# `徐汇区华泾镇项目-招标文件.pdf`（11 页）第 3 页的"招标材料明细表"整行写的
# 是"详见附件1"/"详见清单"/"详见附件"——这份 PDF 本身把数量清单列成"附件1"，
# 但附件的实际内容**没有被装订进这份文件**（第 13 条"附件"目录列着"附件1：
# 数量清单"，全文 11 页止于附件目录，附件本身不在）。Paddle 识别 0 条不是
# 误判：那张表的"数量"列写的字面就是"详见附件"，不是数字，正确的行为就是
# 不产出任何锚点——总不能把占位文字当成真实数据。
# 徐汇项目的真实数量清单只存在于 `徐汇区华泾镇项目-采购清单.xlsx`（B3 用的
# 那份）。B1 因此保留为已知失败用例：一旦有一份真的装订了附件1的 PDF，去掉
# 这个 xfail 即可，不需要改任何生产代码。
_B1_KNOWN_GAP = (
    "B1：徐汇招标文件.pdf 只列出「详见附件1」，附件本身未装订进这份 PDF，"
    "Paddle 因此正确地产出 0 条锚点——不是识别缺陷，是这份 fixture 里没有可用的"
    "数量清单；徐汇的采购清单只能从 徐汇区华泾镇项目-采购清单.xlsx（B3）取得"
)

# A3/A4 用凯硕新正的报价 xlsx 当"零模型上传"时，会在对齐/官方矩阵那一步撞上
# `systematic_vat_mismatch`——**这是一个真实、已在生产代码里被记录过的限制，
# 不是本轮改动引入的**：`routes/analysis.py::tender_list_match` 算术校验只传
# `BidQuoteLine.unit_price`/`total_price` 这两个通用列（该处注释原文："unit_
# price_excl_tax 是原始列但没有配对的 total_price_excl_tax 列，因此不传它"）。
# 凯硕这份表同时有"单价(不含税)"+"合计(不含税)"和"单价(含税)"+"合价(含税)"两套
# 自洽的列；`_TABULAR_COLUMN_PATTERNS` 的 total_price 角色只有一个通用槽位，
# 选中的是"合价(含税)"，配上通用 unit_price="单价(含税)"——**这一步本身是对的、
# 自洽的**。断在下游：`BidQuoteLine` 模型没有 `total_price_excl_tax` 列，
# `unit_price_excl_tax` 因此成了"有值但没人跟它比"的孤儿列，如果哪一步误用了
# 它去跟含税合价比对，就会出这个系统性 13% 偏差。
#
# 2026-08-23 当场试过一个"优先选不含税合计列"的补丁：修好了泰科龙（只有不含税
# 单价、没有含税对照列），但**弄坏了凯硕**（它含税这一对本来自洽，补丁把
# total_price 硬扭去配不含税合计，制造出一个新的错配）——单一样本的失败模式
# 不能当成通用修法，已撤回（git 无残留）。真正的修法要给 `QuoteFact`/
# `BidQuoteLine` 加一个真正的 `total_price_excl_tax` 字段（含 Alembic 迁移），
# 这是本轮范围之外的架构改动，记在这里等安排。
# 零模型用例的 100% 对齐口径：{用例: (锚点总数, 每家报价行数)}。
#
# **这两个数是从原文数出来的，不是从系统读出来的**——这一点是这条断言全部的
# 价值所在。金桥采购清单 89 条、各家报价 89 行，一一对应。徐汇采购清单 170 条
# （矿物电缆 78 + 普通电缆 92 两张表），四家各报 136 行，正好是「矿物 1..44 +
# 普通全部」；余下 34 条是矿物电缆表尾一整块 `RTXMY-*`（数量恒为 2），四家一条
# 都没报——**这是真实的未报价，不是丢行**，逐行核对过。所以 B3 的
# `anchors_covered` 是 136 而不是 170，而每一家的 136 行必须全部落位。
_ZERO_MODEL_FULL_ALIGNMENT = {
    "A3": (89, 89),
    "B3": (170, 136),
}


def _confirm_tender_if_any(client, replay, *, case: Case, project_id: int) -> str:
    """识别招标文件/采购清单，确认成比价基准，**返回识别出的品类**。

    确认成基准这一步是产品设计里的既定动作（design/29："采购清单识别完自动
    设为比价基准"），**预览和正式比价都依赖它**：`build_preview_matrix` 是否
    派生 `quote_derived` 轴，看的是有没有一个"已确认"的 `TenderListSession`，
    不是"有没有上传过"。漏掉这一步会让 tender_anchor 用例在预览阶段被误判成
    quote_derived（本文件写测试时真的踩到过一次）。

    没有招标文件的用例（A2/A4/B2）返回空串——**这正是真实情况**：那些场景
    在当前产品里拿不到品类，后续步骤会因此失败，这是要暴露的事实不是要绕开
    的障碍。
    """
    if case.tender is None:
        return ""
    result = _recognize_tender(client, replay, tender_path=DOCS / case.tender,
                               project_id=project_id)
    items = result.get("items") or []
    category = derive_category(result)
    if not items:
        pytest.skip(_B1_KNOWN_GAP if case.cid == "B1"
                    else f"{case.tender} 识别出 0 条采购清单项")
    assert category, (
        f"{case.tender} 识别出 {len(items)} 条清单却没有品类——"
        "品类必须能从识别产物推出来，否则用户在界面上无路可走（界面没有手动"
        "选择品类的控件）")
    r = client.post("/api/analysis/tender-list/confirm", json={
        "project_id": project_id, "category": category,
        "file_name": case.tender, "anchors_json": items,
        "anchors_total": len(items), "source_type": "excel", "force": True,
    })
    assert r.status_code == 200, f"采购清单确认失败：{r.text}"
    return category


def _setup_project(client, case: Case) -> tuple[int, dict[str, int]]:
    r = client.post("/api/projects", json={"name": f"场景用例 {case.cid} - {case.project}"})
    assert r.status_code == 201, r.text
    project_id = r.json()["id"]
    supplier_ids: dict[str, int] = {}
    for bid in case.bids:
        kw = supplier_keyword(bid) if any(k in bid for k in SUPPLIER_KEYWORDS) else bid
        if kw in supplier_ids:
            continue
        r = client.post("/api/suppliers", json={"name": f"{kw}（场景{case.cid}）"})
        assert r.status_code == 201, r.text
        supplier_ids[kw] = r.json()["id"]
    return project_id, supplier_ids


# ── 品类：**必须跟应用同源，不许测试自己写死** ────────────────────────────
#
# 这里原来是 `CATEGORY_BY_PROJECT = {"金桥": "阀门", "徐汇": "电缆"}`，把品类
# 直接喂给每个用例——于是 B1/B2 一路绿灯，而真实界面上同样的文件根本走不下去：
# 徐汇招标 PDF 识别出 0 条采购清单（附件未装订，见 `_B1_KNOWN_GAP`），
# `detected_category` 为空，前端 `category` 一直是空字符串，`batch-confirm`
# 逐份拒收，用户看到的是「category 不能为空」。
#
# **测试替用户做了一个产品做不到的动作**——跟本文件早先 `integrity_ack` 那次
# 是同一类错误（测试通过 ≠ 流程能走通，见 memory/project_column_shift_design34）。
# 现在改成跟应用同一条链路取品类：招标产物的 `detected_category`。取不到就是
# 取不到，用例照实失败/跳过，不许兜底成常量。
def derive_category(tender_result: dict | None) -> str:
    """招标识别产物的 `detected_category`——有采购清单时品类的来源。

    前端 `WorkspaceView.onTenderDone` / `uploadExcel` 就是这么做的。
    """
    if not tender_result:
        return ""
    return str(tender_result.get("detected_category") or "").strip()


def derive_category_from_quotes(quote_results: list[dict]) -> str:
    """没有采购清单时，从**报价识别产物**取品类。

    A2/A4/B2 这类"无采购清单"用例（design/32 的报价派生轴）在当前产品里
    **完全走不到**：品类唯一来源是招标识别，没有招标文件就永远是空串，
    `batch-confirm` 逐份拒收。设计上支持、实现上不可达。

    这个函数读的 `detected_category` 键**报价侧目前还不产出**（招标侧才有），
    所以它现在恒返回空串、相关用例会红——这是有意为之：先让测试如实反映
    现状，再由生产代码补上这个字段把它转绿（`.claude/rules/tests.md`：
    测试通过必须如实报告，不得靠测试端兜底制造假绿）。
    """
    for r in quote_results:
        cat = str((r or {}).get("detected_category") or "").strip()
        if cat:
            return cat
    return ""


def _completeness_by_supplier(matrix: dict) -> dict[str, dict[str, int]]:
    """按供应商统计完整度，**口径与 `BidMatrix.vue::completeness` 逐字一致**。

    - `aligned`：已对齐到锚点且拿得到金额（单价或合价任一）——可比价的行
    - `priced`：其中还读到了单价的行（跟历史价比要用这个口径）

    两个数分开数，是因为它们回答的是两个问题：一个裸的 52/89 会把"没对上"
    和"对上了但缺单价"混成一个数字，读起来像系统把行弄丢了。
    """
    labels = {s["id"]: (s.get("name") or s.get("letter") or str(s["id"]))
              for s in (matrix.get("suppliers") or [])}
    out: dict[str, dict[str, int]] = {
        name: {"aligned": 0, "priced": 0, "total": len(matrix.get("rows") or [])}
        for name in labels.values()
    }
    for row in matrix.get("rows") or []:
        for cell in row.get("suppliers") or []:
            status = cell.get("cell_status")
            if status not in (None, "quoted", "aggregated"):
                continue
            name = labels.get(cell.get("submission_id") or cell.get("id"))
            if name is None or name not in out:
                continue
            if cell.get("price") is not None or cell.get("total") is not None:
                out[name]["aligned"] += 1
            if cell.get("price") is not None:
                out[name]["priced"] += 1
    return out


# 2026-08-23 实测基线，`(可比价, 有单价)`。用 `>=` 断言：变好是预期内的（改进
# 识别/对齐后要更新基线），变差必须能解释。
#
# 泰科龙 (64, 52) 的缺口是两个**已知且未修**的引擎侧缺陷叠加，不是本轮引入：
#   · 89→64：名称列读错（#44-49 闸阀被读成球阀那批），对不上锚点，design/34
#   · 64→52：第 10 页那批 Paddle 返回空白，没有单价，design/33
# 数字要真正上去得先定这两个方案，在那之前这条基线负责保证它不再恶化。
COMPLETENESS_BASELINE: dict[str, dict[str, tuple[int, int]]] = {
    "A1": {"绵存": (89, 89), "凯硕": (88, 88), "泰科龙": (64, 52)},
}


@pytest.mark.slow  # 预览链路要先把报价件识别出来
class TestPreviewStage:
    """阶段二：预览。**预览永远不拦**是产品级不变式——本轮人工测试抓到的回归
    正是这里（一行列错位就让整份供应商连预览都进不去）。断言"每一家都进得了
    预览"，不断言"没有疑点"：疑点本来就该被看见，只是不该变成拦截。
    """

    @pytest.mark.parametrize("case", CASES, ids=CASE_IDS)
    def test_every_supplier_reaches_preview(self, client, case: Case):
        cl, replay = client
        gaps = _missing(case)
        if gaps:
            pytest.skip(f"用例 {case.cid} 缺文件：{gaps}")

        project_id, supplier_ids = _setup_project(cl, case)
        category = _confirm_tender_if_any(cl, replay, case=case, project_id=project_id)

        # 先上传报价（上传本身不需要品类），再决定品类从哪来——有采购清单用
        # 采购清单的，没有就退到报价自己的 `detected_category`。这跟真实产品
        # 的取值顺序一致，测试不再自带答案。
        jobs = []
        for bid in case.bids:
            kw = supplier_keyword(bid) if any(k in bid for k in SUPPLIER_KEYWORDS) else bid
            jobs.append((kw, bid, _upload_quote(cl, replay, path=DOCS / bid,
                                                project_id=project_id,
                                                supplier_id=supplier_ids[kw],
                                                category=category)))
        if not category:
            category = derive_category_from_quotes(
                [(j.get("result") or {}) for _kw, _b, j in jobs])
        assert category, (
            f"用例 {case.cid}：没有采购清单，报价识别产物里也没有品类——"
            "用户在界面上到这一步是死路（没有手动选择品类的入口）。"
            "这条断言失败即为该产品缺口的复现")

        confirmations = []
        for kw, bid, job in jobs:
            items = (job.get("result") or {}).get("items") or []
            assert items, f"{bid}：识别出 0 条报价行，job={job}"
            confirmations.append({
                "job_id": job["id"], "supplier_id": supplier_ids[kw],
                "supplier_name": kw, "project_id": project_id, "category": category,
                "overrides": _resolve_review_rows(items),
            })

        r = cl.post("/api/analysis/bid-matrix/preview", json={
            "project_id": project_id, "category": category, "confirmations": confirmations,
        })
        assert r.status_code == 200, (
            f"用例 {case.cid}：预览被拦截，这正是「预览永远不拦」要防的回归——{r.text}")
        matrix = r.json()["matrix"]
        assert len(matrix.get("suppliers") or []) == len(case.bids), (
            f"用例 {case.cid}：{len(case.bids)} 家供应商，只有 "
            f"{len(matrix.get('suppliers') or [])} 家进了预览")
        assert matrix.get("axis_kind", "tender_anchor") == case.axis_kind, (
            f"用例 {case.cid}：预览行轴应为 {case.axis_kind}，实得 "
            f"{matrix.get('axis_kind')}")
        assert matrix.get("basis") == "preview", "预览结果必须标注 basis=preview"


@pytest.mark.slow  # 整条比价链路走 TestClient 跑完
class TestCompareStage:
    """阶段三：正式比价。这一层要验证的是 CLAUDE.md §4 的「行轴」不变式：
    `quote_derived` 只能进预览、绝不能进官方结果——不是靠约定，是靠
    `/bid-matrix` 在没有已确认采购清单时**直接拒绝**（见 `routes/analysis.py`
    的显式注释："silently falling back to legacy mode would show 449 rows of
    all-history quotes"），加上 schema 的 `_quote_derived_axis_is_preview_only`
    validator 双重兜底。
    """

    @pytest.mark.parametrize("case", [c for c in CASES if c.tender], ids=lambda c: c.cid)
    def test_tender_anchor_case_produces_official_matrix(self, client, case: Case):
        cl, replay = client
        gaps = _missing(case)
        if gaps:
            pytest.skip(f"用例 {case.cid} 缺文件：{gaps}")

        project_id, supplier_ids = _setup_project(cl, case)
        category = _confirm_tender_if_any(cl, replay, case=case, project_id=project_id)
        assert category, f"用例 {case.cid}：拿不到品类，正式比价无从谈起"

        submission_ids = []
        for bid in case.bids:
            kw = supplier_keyword(bid) if any(k in bid for k in SUPPLIER_KEYWORDS) else bid
            job = _upload_quote(cl, replay, path=DOCS / bid, project_id=project_id,
                               supplier_id=supplier_ids[kw], category=category)
            items = (job.get("result") or {}).get("items") or []
            r = cl.post("/api/quotes/batch-confirm", json={
                "job_id": job["id"], "supplier_id": supplier_ids[kw],
                "supplier_name": kw, "project_id": project_id, "category": category,
                "overrides": _resolve_for_official_confirm(items),
            })
            assert r.status_code == 200, f"{bid} 正式入库失败：{r.text}"
            submission_ids.append(r.json()["submission_id"])

        assert len(set(submission_ids)) == len(case.bids), "每家必须是独立 submission"

        r = cl.post("/api/analysis/tender-list/match", data={
            "project_id": str(project_id), "category": category,
            "supplier_ids": ",".join(str(v) for v in supplier_ids.values()),
            "submission_ids": ",".join(str(i) for i in submission_ids),
        })
        assert r.status_code == 200, f"对齐失败：{r.text}"

        # ── 零模型用例必须 100% 对齐 ─────────────────────────────────────
        # A3/B3 走的是确定性链路（Excel/CSV 解析 + 顺序/子序列直连），**没有任何
        # 模型调用**：识别不参与，语义匹配不参与。这种情况下还对不齐，一定是代码
        # 缺陷，不是"识别不准"能解释的。所以这里不设容差、不设基线，直接钉死。
        #
        # 三条真实缺陷是被这个断言逼出来的（2026-08-23）：
        #   1. 泰科龙那份只有「不含税单价 + 价税合计」，`合计(不含税)/税率/税额`
        #      三列在 `_TABULAR_COLUMN_PATTERNS` 里没有槽位、读出来就扔，于是不含税
        #      单价被迫跟含税合价配对，89 行全数 `tax_basis_suspect` → 整份被
        #      `systematic_vat_mismatch` 挡在正式比价外（本条断言之前是 skip 掉的）。
        #   2. 徐汇采购清单是两张表（矿物电缆 78 + 普通电缆 92），预览只解析
        #      `pick_default_sheet` 挑中的一张，行轴少掉一半。
        #   3. 供应商只报清单的一部分（136/170）时，顺序直连的「行数==锚点数」
        #      门禁一律拒绝，全部回落语义匹配，匹配率 58%。
        if case.cid in _ZERO_MODEL_FULL_ALIGNMENT:
            summary = r.json()
            expect_anchors, expect_rows = _ZERO_MODEL_FULL_ALIGNMENT[case.cid]
            assert summary["anchors_total"] == expect_anchors, (
                f"{case.cid}：行轴 {summary['anchors_total']} 条，应为 {expect_anchors}"
                "——采购清单的锚点数变了，先解释清楚是清单换了还是解析退化了")
            for st in summary["readiness_list"]:
                who = st["supplier_name"]
                assert st["quote_rows"] == expect_rows, (
                    f"{case.cid}/{who} 报价 {st['quote_rows']} 行，应为 {expect_rows}")
                assert st["matched_rows"] == expect_rows, (
                    f"{case.cid}/{who} 只对齐 {st['matched_rows']}/{st['quote_rows']} 行。"
                    "零模型链路对不齐 = 代码缺陷，不接受基线化")
                assert st["pending_rows"] == 0, f"{case.cid}/{who} 有 {st['pending_rows']} 行待确认"
                assert st["residue_rows"] == 0, f"{case.cid}/{who} 有 {st['residue_rows']} 行未匹配"
                assert st["validation_failed_rows"] == 0, (
                    f"{case.cid}/{who} 有 {st['validation_failed_rows']} 行校验失败")
            assert summary["residue"] == 0 and summary["low_conf"] == 0, (
                f"{case.cid}：residue={summary['residue']} low_conf={summary['low_conf']}"
                "——确定性直连不该产出低置信或无归属的行")
            assert summary["matched_quotes"] == summary["total_quotes"], (
                f"{case.cid}：{summary['matched_quotes']}/{summary['total_quotes']} 行有归属")

        r = cl.post("/api/analysis/bid-matrix", json={
            "project_id": project_id, "category": category,
            "supplier_ids": list(supplier_ids.values()), "submission_ids": submission_ids,
        })
        assert r.status_code == 200, f"官方矩阵失败：{r.text}"
        matrix = r.json()
        assert matrix.get("rows"), "官方矩阵没有行"
        assert matrix.get("axis_kind", "tender_anchor") == "tender_anchor", (
            "官方矩阵的行轴必须是 tender_anchor——本用例确认过采购清单")
        assert len(matrix.get("suppliers") or []) == len(case.bids)

        # ── 完整度基线 ───────────────────────────────────────────────────
        # 之前这里只断言"矩阵有行、列数对"，所以泰科龙的可比价行数从 89 掉到
        # 64、有单价行数掉到 52，测试一声不吭——用户在界面上看到 52/89 才发现。
        # 现在把每家的两个口径都钉住（口径与 `BidMatrix.vue::completeness` 一致）。
        counts = _completeness_by_supplier(matrix)
        expected = COMPLETENESS_BASELINE.get(case.cid)
        if expected:
            for kw, (exp_aligned, exp_priced) in expected.items():
                # 矩阵列名是 `_setup_project` 造的"泰科龙（场景A1）"，按关键字匹配
                got = next((v for name, v in counts.items() if kw in name), None)
                assert got is not None, f"{kw} 不在矩阵列里：{list(counts)}"
                assert got["aligned"] >= exp_aligned, (
                    f"{case.cid}/{kw} 可比价行数 {got['aligned']} 低于基线 "
                    f"{exp_aligned}（共 {got['total']} 行）——对齐能力退化了")
                assert got["priced"] >= exp_priced, (
                    f"{case.cid}/{kw} 有单价行数 {got['priced']} 低于基线 "
                    f"{exp_priced}（共 {got['total']} 行）——识别能力退化了")

    @pytest.mark.parametrize("case", [c for c in CASES if not c.tender], ids=lambda c: c.cid)
    def test_quote_derived_case_is_refused_by_the_official_endpoint(self, client, case: Case):
        """没有采购清单时，官方 `/bid-matrix` 必须直接拒绝，不能悄悄退化成
        某种"看起来正式"的矩阵——这正是 `quote_derived` "只能进预览"这条
        CLAUDE.md 不变式在 API 层的体现。"""
        cl, replay = client
        gaps = _missing(case)
        if gaps:
            pytest.skip(f"用例 {case.cid} 缺文件：{gaps}")

        project_id, supplier_ids = _setup_project(cl, case)

        jobs = []
        for bid in case.bids:
            kw = supplier_keyword(bid) if any(k in bid for k in SUPPLIER_KEYWORDS) else bid
            jobs.append((kw, bid, _upload_quote(cl, replay, path=DOCS / bid,
                                                project_id=project_id,
                                                supplier_id=supplier_ids[kw], category="")))
        category = derive_category_from_quotes(
            [(j.get("result") or {}) for _kw, _b, j in jobs])
        assert category, (
            f"用例 {case.cid}：无采购清单场景（design/32 报价派生轴）在当前产品里"
            "不可达——品类只能来自招标识别，没有招标文件就永远拿不到。"
            "这条断言失败即为该缺口的复现")

        submission_ids = []
        for kw, bid, job in jobs:
            items = (job.get("result") or {}).get("items") or []
            r = cl.post("/api/quotes/batch-confirm", json={
                "job_id": job["id"], "supplier_id": supplier_ids[kw],
                "supplier_name": kw, "project_id": project_id, "category": category,
                "overrides": _resolve_for_official_confirm(items),
            })
            assert r.status_code == 200, f"{bid} 正式入库失败：{r.text}"
            submission_ids.append(r.json()["submission_id"])

        r = cl.post("/api/analysis/bid-matrix", json={
            "project_id": project_id, "category": category,
            "supplier_ids": list(supplier_ids.values()), "submission_ids": submission_ids,
        })
        assert r.status_code == 409, (
            f"用例 {case.cid}：没有已确认采购清单，官方矩阵本该 409 拒绝，"
            f"实得 {r.status_code}——如果这条改成 200，等于让 quote_derived "
            f"行轴悄悄流进了官方结果")
