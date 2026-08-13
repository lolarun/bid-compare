"""Export endpoints — Excel download for all major data views."""
import io
from datetime import datetime
from urllib.parse import quote as url_quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from apps.api.core.database import get_db
from apps.api.core.enums import ROLE_ADMIN, ROLE_BUYER
from apps.api.core.security import require_role
from apps.api.core.utils import parse_id_csv
from apps.api.models.material import Material
from apps.api.models.supplier import Supplier
from apps.api.models.quote import Quote
from apps.api.models.project import Project

router = APIRouter(
    prefix="/api/export",
    tags=["export"],
    dependencies=[Depends(require_role(ROLE_ADMIN, ROLE_BUYER))],
)

# ── Shared styles ────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_GREEN_FILL = PatternFill(start_color="F6FFED", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFF2F0", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFBE6", fill_type="solid")


def _style_header(ws, col_count: int):
    """Apply header styling to row 1."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws, min_width=10, max_width=40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = length


def _to_streaming(wb: Workbook, filename: str) -> StreamingResponse:
    """Serialize workbook to a streaming download response."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # RFC 5987: use filename* with UTF-8 encoding for non-ASCII filenames
    encoded = url_quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
        },
    )


# ── 1. Dashboard / 仪表盘报表 ────────────────────────────────────────────────

@router.get("/dashboard")
def export_dashboard(db: Session = Depends(get_db)):
    """导出仪表盘报表 — 包含采购概览 + 品类统计。"""
    from apps.api.services.history.statistics import get_dashboard_summary
    summary = get_dashboard_summary(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "采购概览"

    # Summary cards
    ws.append(["指标", "数值"])
    ws.append(["累计入库材料", summary["total_materials"]])
    ws.append(["供应商数", summary["total_suppliers"]])
    ws.append(["项目数", summary["total_projects"]])
    ws.append(["报价条数", summary["total_quotes"]])
    _style_header(ws, 2)

    # Category breakdown sheet
    ws2 = wb.create_sheet("品类统计")
    cats = db.execute(
        select(Material.category, Quote.supplier_id)
        .outerjoin(Quote, Quote.material_id == Material.id)
    ).all()
    cat_stats: dict[str, dict] = {}
    for cat, sid in cats:
        if cat not in cat_stats:
            cat_stats[cat] = {"count": 0, "suppliers": set()}
        cat_stats[cat]["count"] += 1
        if sid:
            cat_stats[cat]["suppliers"].add(sid)

    ws2.append(["品类", "报价条数", "供应商数"])
    _style_header(ws2, 3)
    for cat, s in sorted(cat_stats.items(), key=lambda x: -x[1]["count"]):
        ws2.append([cat, s["count"], len(s["suppliers"])])

    _auto_width(ws)
    _auto_width(ws2)

    ts = datetime.now().strftime("%Y%m%d")
    return _to_streaming(wb, f"MEMPAS_仪表盘报表_{ts}.xlsx")


# ── 2. Suppliers / 供应商名单 ────────────────────────────────────────────────

@router.get("/suppliers")
def export_suppliers(db: Session = Depends(get_db)):
    """导出供应商名单。"""
    rows = db.scalars(select(Supplier).order_by(Supplier.id)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "供应商名单"

    ws.append(["ID", "供应商名称", "简称", "联系人", "电话", "经营品类", "中标次数", "合作评分", "备注"])
    _style_header(ws, 9)

    for s in rows:
        cats = ", ".join(s.categories) if isinstance(s.categories, list) else str(s.categories or "")
        ws.append([s.id, s.name, s.short_name, s.contact, s.phone, cats,
                    s.win_count, s.cooperation_score, s.remark])

    _auto_width(ws)
    ts = datetime.now().strftime("%Y%m%d")
    return _to_streaming(wb, f"MEMPAS_供应商名单_{ts}.xlsx")


# ── 3. Materials / 物料主数据 ────────────────────────────────────────────────

@router.get("/materials")
def export_materials(
    category: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """导出物料主数据标准库。"""
    stmt = select(Material).order_by(Material.profession, Material.category, Material.id)
    if category:
        stmt = stmt.where(Material.category == category)
    rows = db.scalars(stmt).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "物料主数据"

    ws.append(["ID", "物料编码", "标准名称", "专业", "品类", "子类", "规格", "材质", "单位", "品牌", "执行标准", "参考均价"])
    _style_header(ws, 12)

    for m in rows:
        ws.append([m.id, m.material_code, m.standard_name, m.profession,
                    m.category, m.sub_category, m.spec, m.material_type,
                    m.unit, m.brand, m.exec_standard, m.ref_price_avg])

    _auto_width(ws)
    ts = datetime.now().strftime("%Y%m%d")
    cat_suffix = f"_{category}" if category else ""
    return _to_streaming(wb, f"MEMPAS_物料主数据{cat_suffix}_{ts}.xlsx")


# ── 4. Quotes / 采购数据(历史记录) ───────────────────────────────────────────

@router.get("/quotes")
def export_quotes(
    category: str | None = Query(None),
    supplier_id: int | None = Query(None),
    project_id: int | None = Query(None),
    alert_level: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """导出采购历史数据（支持筛选条件透传）。"""
    stmt = (
        select(Quote, Material.standard_name, Material.spec, Material.category,
                 Supplier.name.label("supplier_name"), Project.name.label("project_name"))
        .outerjoin(Material, Quote.material_id == Material.id)
        .outerjoin(Supplier, Quote.supplier_id == Supplier.id)
        .outerjoin(Project, Quote.project_id == Project.id)
    )
    if category:
        stmt = stmt.where(Material.category == category)
    if supplier_id:
        stmt = stmt.where(Quote.supplier_id == supplier_id)
    if project_id:
        stmt = stmt.where(Quote.project_id == project_id)
    if alert_level:
        stmt = stmt.where(Quote.alert_level == alert_level)

    rows = db.execute(stmt.order_by(Quote.id.desc()).limit(10000)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "采购数据"

    ws.append(["ID", "物料名称", "规格", "品类", "供应商", "项目",
               "单价", "单价(不含税)", "税率", "数量", "总价",
               "偏差率", "告警", "报价日期"])
    _style_header(ws, 14)

    alert_fills = {"red": _RED_FILL, "yellow": _YELLOW_FILL, "normal": _GREEN_FILL}

    for quote, mat_name, spec, cat, sup_name, proj_name in rows:
        row_idx = ws.max_row + 1
        ws.append([
            quote.id, mat_name, spec, cat, sup_name, proj_name,
            quote.unit_price, quote.unit_price_excl_tax, quote.tax_rate,
            quote.quantity, quote.total_price,
            f"{quote.deviation_pct * 100:.1f}%" if quote.deviation_pct is not None else "",
            quote.alert_level or "normal",
            quote.quote_date.strftime("%Y-%m-%d") if quote.quote_date else "",
        ])
        # Color alert column
        fill = alert_fills.get(quote.alert_level or "normal")
        if fill:
            ws.cell(row=row_idx, column=13).fill = fill

    _auto_width(ws)
    ts = datetime.now().strftime("%Y%m%d")
    return _to_streaming(wb, f"MEMPAS_采购数据_{ts}.xlsx")


# ── 5. Bid Matrix / 比价矩阵 ────────────────────────────────────────────────

@router.get("/bid-matrix")
def export_bid_matrix(
    supplier_ids: str = Query(..., description="逗号分隔的供应商ID"),
    project_id: int | None = Query(None),
    category: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """导出横向比价矩阵为 Excel（带色标）。

    v2.5: 优先使用锚点全量矩阵（TenderListSession）；无 session 时 fallback 旧逻辑。
    """
    sids = parse_id_csv(supplier_ids, "supplier_ids")

    from apps.api.services.matrix.bid_export_service import get_bid_matrix_for_export
    if project_id and category:
        result = get_bid_matrix_for_export(db, project_id, category, sids)
    else:
        raise HTTPException(
            status_code=400,
            detail="必须同时提供 project_id 和 category 才能导出比价矩阵。",
        )

    wb = Workbook()

    # ── 摘要 sheet ────────────────────────────────────────────────────────────
    md = result.get("matrix_distribution") or {}
    if md:
        ws_sum = wb.active
        ws_sum.title = "供应商覆盖摘要"
        N_val = md.get("supplier_count", len(sids))
        anchors_total = md.get("anchors_total", 0)
        q_full = md.get("quoted_full_count", 0)
        c_full = md.get("covered_full_count", 0)
        q_ge2  = md.get("quoted_ge_2_count", 0)
        c_ge2  = md.get("covered_ge_2_count", 0)
        q_dist_raw = md.get("quoted_distribution") or {}
        c_dist_raw = md.get("covered_distribution") or {}

        def _pct(n, total):
            return f"{n/total*100:.1f}%" if total else "—"

        # Section 1: overview
        sum_rows = [
            ["供应商覆盖摘要", ""],
            ["供应商数量 (N)", N_val],
            ["锚点总数", anchors_total],
            [],
            ["── 关键指标 ──", "数量", "占比", "含义"],
            ["可比价锚点（quoted ≥2家）", q_ge2, _pct(q_ge2, anchors_total), "至少2家供应商有明确报价，可自动横向比价"],
            [f"{N_val}家完整 quoted", q_full, _pct(q_full, anchors_total), "全部供应商均有明确报价，比价最可靠"],
            [f"覆盖 ≥2家（含待确认）", c_ge2, _pct(c_ge2, anchors_total), "含 pending（待人工复核），复核后可比价潜力"],
            [f"{N_val}家完整覆盖（含待确认）", c_full, _pct(c_full, anchors_total), "含 pending，人工复核后有机会达到完整比价"],
            [],
            ["── quoted 分布（已报价家数）──", "锚点数", "占比"],
        ]
        for k in range(N_val + 1):
            cnt = q_dist_raw.get(str(k), 0)
            sum_rows.append([f"{k}/{N_val}家 quoted", cnt, _pct(cnt, anchors_total)])

        sum_rows += [
            [],
            ["── covered 分布（含待确认）──", "锚点数", "占比"],
        ]
        for k in range(N_val + 1):
            cnt = c_dist_raw.get(str(k), 0)
            sum_rows.append([f"{k}/{N_val}家 covered", cnt, _pct(cnt, anchors_total)])

        sum_rows += [
            [],
            ["说明", ""],
            ["quoted/aggregated", "供应商已报价（可直接参与比价）"],
            ["pending（待确认）", "LLM 找到候选但不确定，需人工复核后确认"],
            ["missing（未报价）", "该锚点本供应商无任何报价，无法比价"],
            ["DB 口径说明", f"covered_full 实测 ~{c_full}/90；旧版脚本推断的 85 基于错误假设已弃用"],
        ]

        _BLUE_FILL  = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
        _LBLUE_FILL = PatternFill(start_color="E6F4FF", end_color="E6F4FF", fill_type="solid")
        _GREY_HDR   = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        for r_idx, row_data in enumerate(sum_rows, start=1):
            ws_sum.append(row_data)
            cell0 = ws_sum.cell(row=r_idx, column=1)
            if row_data and str(row_data[0]).startswith("──"):
                cell0.font = Font(bold=True)
                cell0.fill = _GREY_HDR
            elif row_data and row_data[0] == "供应商覆盖摘要":
                cell0.font = Font(bold=True, size=14, color="FFFFFF")
                ws_sum.cell(row=r_idx, column=1).fill = _BLUE_FILL

        # Bold the key metrics section header
        for r_idx, row_data in enumerate(sum_rows, start=1):
            if row_data and row_data[0] == "── 关键指标 ──":
                for col in range(1, 5):
                    c = ws_sum.cell(row=r_idx, column=col)
                    c.font = Font(bold=True)
                    c.fill = _LBLUE_FILL

        ws_sum.column_dimensions["A"].width = 32
        ws_sum.column_dimensions["B"].width = 12
        ws_sum.column_dimensions["C"].width = 10
        ws_sum.column_dimensions["D"].width = 42

        # Main matrix on new sheet
        ws = wb.create_sheet("比价矩阵")
    else:
        ws = wb.active
        ws.title = "比价矩阵"

    is_anchor = result.get("anchor_matrix", False)
    suppliers = result["suppliers"]

    # Header row — anchor mode adds seq + 材质/品牌 columns
    if is_anchor:
        header = ["序号", "材料", "规格", "材质", "品牌要求", "历史均价", "合理史低"]
    else:
        header = ["材料", "规格", "历史均价", "合理史低"]
    for s in suppliers:
        header += [f"{s['letter']} {s['name']}(单价)", f"{s['letter']}(偏差)", f"{s['letter']}(状态)"]
    header += ["最低偏差", "推荐"]
    ws.append(header)
    _style_header(ws, len(header))

    # Cell fills
    _ORANGE_FILL    = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
    _GREY_FILL      = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    _OCR_FILL       = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")  # light green
    _FP_FILL        = PatternFill(start_color="FFF1F0", end_color="FFF1F0", fill_type="solid")  # light red
    alert_fills     = {"red": _RED_FILL, "yellow": _YELLOW_FILL, "normal": _GREEN_FILL}

    _STATUS_LABEL = {
        "quoted":     "",
        "aggregated": "聚合",
        "pending":    "待确认",
        "excluded":   "已排除",
        "missing":    "未报价",
        None:         "",
    }

    def _flags_label(cell: dict) -> str:
        flags = cell.get("flags") or []
        if not flags:
            return ""
        parts = []
        for f in flags:
            if f == "ocr_corrected_verified":
                parts.append("OCR纠错✓")
            elif f == "ocr_corrected":
                parts.append("OCR纠错")
            elif f.startswith("valve_type_conflict:"):
                parts.append(f"阀型冲突:{f.split(':',1)[1]}")
            elif f == "canonical_conflict":
                parts.append("规格冲突")
            elif f.startswith("ac_conflict"):
                parts.append("锚点冲突")
            elif f == "missing_without_evidence":
                parts.append("缺少证据")
            elif f.startswith("risky_candidate"):
                parts.append("候选风险")
            elif f.startswith("dup_qids"):
                parts.append("重复报价")
        return " | ".join(parts)

    for row in result["rows"]:
        if is_anchor:
            data = [
                row.get("anchor_seq", ""),
                row["material_name"],
                row.get("spec", ""),
                row.get("materials", ""),
                row.get("brand", ""),
                row["historical_avg"]["price"] if row.get("historical_avg") else "",
                row["reasonable_low"]["price"] if row.get("reasonable_low") else "",
            ]
            col_offset = 7
        else:
            data = [
                row["material_name"],
                row.get("spec", ""),
                row["historical_avg"]["price"] if row.get("historical_avg") else "",
                row["reasonable_low"]["price"] if row.get("reasonable_low") else "",
            ]
            col_offset = 4

        for cell in row["suppliers"]:
            status = cell.get("cell_status")
            # Price: show for quoted/aggregated/pending; blank for excluded/missing
            if status in (None, "quoted", "aggregated", "pending"):
                data.append(cell["price"] if cell["price"] is not None else "")
            else:
                data.append("")
            # Deviation: only for confirmed cells
            if status in (None, "quoted", "aggregated") and cell.get("deviation_pct") is not None:
                data.append(f"{cell['deviation_pct'] * 100:.1f}%")
            else:
                data.append("")
            # Status label — combine status + flags + evidence for context
            status_txt = _STATUS_LABEL.get(status, status or "")
            flag_txt = _flags_label(cell)
            evidence_txt = (cell.get("evidence") or "")[:80]
            if flag_txt:
                status_txt = f"{status_txt} [{flag_txt}]" if status_txt else f"[{flag_txt}]"
            if evidence_txt:
                status_txt = f"{status_txt} {evidence_txt}" if status_txt else evidence_txt
            data.append(status_txt)

        data.append(f"{row['min_deviation'] * 100:.1f}%" if row.get("min_deviation") is not None else "")
        data.append(row.get("recommended", ""))

        row_idx = ws.max_row + 1
        ws.append(data)

        # Color cells by status and flags
        for si, cell in enumerate(row["suppliers"]):
            status = cell.get("cell_status")
            flags  = cell.get("flags") or []
            price_col  = col_offset + si * 3 + 1
            status_col = col_offset + si * 3 + 3
            if status == "pending":
                ws.cell(row=row_idx, column=price_col).fill  = _ORANGE_FILL
                ws.cell(row=row_idx, column=status_col).fill = _ORANGE_FILL
            elif status in ("missing", "excluded"):
                ws.cell(row=row_idx, column=price_col).fill  = _GREY_FILL
                ws.cell(row=row_idx, column=status_col).fill = _GREY_FILL
            elif status in (None, "quoted", "aggregated"):
                has_fp    = any(f.startswith("valve_type_conflict") for f in flags)
                has_ocr   = "ocr_corrected_verified" in flags
                if has_fp:
                    ws.cell(row=row_idx, column=price_col).fill  = _FP_FILL
                    ws.cell(row=row_idx, column=status_col).fill = _FP_FILL
                elif has_ocr:
                    ws.cell(row=row_idx, column=price_col).fill  = _OCR_FILL
                    ws.cell(row=row_idx, column=status_col).fill = _OCR_FILL
                else:
                    alert = cell.get("alert_level", "normal")
                    fill = alert_fills.get(alert)
                    if fill:
                        ws.cell(row=row_idx, column=status_col).fill = fill

    # Totals row (quoted-only — same as backend); leading blanks match col_offset
    totals_data = (["汇总", "", "", "", "", "", ""] if is_anchor else ["汇总", "", "", ""])
    totals_map = {t["id"]: t for t in result["totals"]}
    for s in suppliers:
        t = totals_map.get(s["id"])
        totals_data.append(f"¥{t['total']:,.0f}" if t else "")
        totals_data.append(f"{t['avg_deviation'] * 100:.1f}%" if t and t.get("avg_deviation") is not None else "")
        totals_data.append("")
    totals_data += ["", ""]
    ws.append(totals_data)
    for col in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    _auto_width(ws)
    ts = datetime.now().strftime("%Y%m%d")
    cat_suffix = f"_{category}" if category else ""
    return _to_streaming(wb, f"MEMPAS_比价矩阵{cat_suffix}_{ts}.xlsx")


# ── 6. Logs / 操作日志 ──────────────────────────────────────────────────────

@router.get("/logs")
def export_logs():
    """导出操作日志（当前为占位 — 日志模块待实装后对接）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "操作日志"

    ws.append(["时间", "操作人", "模块", "操作类型", "对象", "结果", "详情"])
    _style_header(ws, 7)

    # Placeholder: 当日志模块实装后，此处从 AuditLog 表查询
    ws.append(["暂无日志数据 — 审计日志模块尚未实装", "", "", "", "", "", ""])

    _auto_width(ws)
    ts = datetime.now().strftime("%Y%m%d")
    return _to_streaming(wb, f"MEMPAS_操作日志_{ts}.xlsx")
